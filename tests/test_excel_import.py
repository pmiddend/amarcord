import datetime
from decimal import Decimal

import pytest
from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText
from openpyxl.cell.rich_text import TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.formula import DataTableFormula

from amarcord.db import orm
from amarcord.db.associated_table import AssociatedTable
from amarcord.db.attributo_id import AttributoId
from amarcord.db.beamtime_id import BeamtimeId
from amarcord.db.chemical_type import ChemicalType
from amarcord.db.excel_import import ConversionError
from amarcord.db.excel_import import ParsedRunSpreadsheet
from amarcord.db.excel_import import ParsedRunSpreadsheetRun
from amarcord.db.excel_import import RunCellResult
from amarcord.db.excel_import import SpreadsheetValidationErrors
from amarcord.db.excel_import import SpreadsheetValidationResult
from amarcord.db.excel_import import convert_value
from amarcord.db.excel_import import create_data_set_for_runs
from amarcord.db.excel_import import create_runs_from_spreadsheet
from amarcord.db.excel_import import data_sets_are_equal
from amarcord.db.excel_import import excel_import_convert_cell_to_integer
from amarcord.db.excel_import import excel_import_convert_cell_to_string
from amarcord.db.excel_import import parse_run_spreadsheet_workbook
from amarcord.db.run_external_id import RunExternalId


def test_conversion_error() -> None:
    e = ConversionError.singleton("f")
    assert len(e.error_messages) == 1

    multiple = ConversionError(["a", "b"])
    assert len(multiple.error_messages) == 2

    prepended = multiple.prepend("x")
    assert prepended.error_messages[0] == "x: a"
    assert prepended.error_messages[1] == "x: b"

    assert prepended.joined_messages() == "x: a, x: b"

    combined = e.combine(multiple)

    assert combined.error_messages == ["f", "a", "b"]


def test_convert_value() -> None:
    chemical = orm.Chemical(
        beamtime_id=BeamtimeId(1),
        name="chem1",
        responsible_person="",
        modified=datetime.datetime.now(),
        type=ChemicalType.CRYSTAL,
    )
    chemical.id = 80
    chemicals_by_name = {chemical.name: chemical}

    a = orm.Attributo(
        beamtime_id=BeamtimeId(1),
        name="test",
        description="",
        group="",
        associated_table=AssociatedTable.RUN,
        json_schema={"type": "integer"},
    )
    a.id = AttributoId(1)

    # Our attributo is of type integer, and we provide an integer cell.
    result = convert_value(a, chemicals_by_name, 1337)
    assert isinstance(result, orm.RunHasAttributoValue)
    assert result.integer_value == 1337

    # Our attributo value is "empty cell", and the type is integer
    result = convert_value(a, chemicals_by_name, "")
    assert result is None

    # Our attributo value is "string value that's an integer", and the type is integer
    result = convert_value(a, chemicals_by_name, "3")
    assert isinstance(result, orm.RunHasAttributoValue)
    assert result.integer_value == 3

    # Our attributo value is "string that's not an integer", and the type is integer
    result = convert_value(a, chemicals_by_name, "a")
    assert isinstance(result, ConversionError)

    # Our attributo value is a datetime, and the type is integer
    result = convert_value(a, chemicals_by_name, datetime.datetime.now())
    assert isinstance(result, ConversionError)

    a.json_schema = {"type": "number", "minimum": 5, "maximum": 1000}

    # Our attributo is of type decimal, and we provide a float cell.
    result = convert_value(a, chemicals_by_name, 337.5)
    assert isinstance(result, orm.RunHasAttributoValue)
    assert result.float_value == 337.5

    # Our attributo is of type decimal, and we provide a float outside the defined range.
    result = convert_value(a, chemicals_by_name, 1337.5)
    assert isinstance(result, ConversionError)

    # Our attributo is of type decimal, and we provide an empty string
    result = convert_value(a, chemicals_by_name, "")
    assert result is None

    # Our attributo is of type decimal, and we provide a float string.
    result = convert_value(a, chemicals_by_name, "337.5")
    assert isinstance(result, orm.RunHasAttributoValue)
    assert result.float_value == 337.5

    # Our attributo is of type decimal, and we provide an invalid float string.
    result = convert_value(a, chemicals_by_name, "a")
    assert isinstance(result, ConversionError)

    # Our attributo value is a datetime, and the type is decimal
    result = convert_value(a, chemicals_by_name, datetime.datetime.now())
    assert isinstance(result, ConversionError)

    a.json_schema = {"type": "integer", "format": "chemical-id"}

    # Our attributo is of type chemical, and we provide a string a valid chemical name
    result = convert_value(a, chemicals_by_name, chemical.name)
    assert isinstance(result, orm.RunHasAttributoValue)
    assert result.chemical_value == chemical.id

    # Our attributo is of type chemical, and we provide an integer
    result = convert_value(a, chemicals_by_name, chemical.id)
    assert isinstance(result, ConversionError)

    # Our attributo is of type chemical, and we provide an empty string
    result = convert_value(a, chemicals_by_name, "")
    assert result is None

    # Our attributo is of type chemical, and we provide an invalid string
    result = convert_value(a, chemicals_by_name, "invalid")
    assert isinstance(result, ConversionError)

    a.json_schema = {"type": "string"}

    # Our attributo is of type string, and we provide a string
    result = convert_value(a, chemicals_by_name, "test")
    assert isinstance(result, orm.RunHasAttributoValue)
    assert result.string_value == "test"

    # Our attributo is of type string, and we provide a number
    result = convert_value(a, chemicals_by_name, 3)
    assert isinstance(result, orm.RunHasAttributoValue)
    assert result.string_value == "3"

    result = convert_value(a, chemicals_by_name, datetime.datetime.now())
    assert isinstance(result, ConversionError)

    a.json_schema = {"type": "boolean"}
    # Our attributo is of type boolean, and we provide a boolean
    result = convert_value(a, chemicals_by_name, value=False)
    assert isinstance(result, ConversionError)

    a.json_schema = {"type": "integer", "format": "date-time"}
    # Our attributo is of type datetime, and we provide a datetime
    result = convert_value(a, chemicals_by_name, datetime.datetime.now())
    assert isinstance(result, ConversionError)

    a.json_schema = {"type": "string", "enum": ["a", "b"]}

    # Our attributo is of type choice, and we provide an empty string
    result = convert_value(a, chemicals_by_name, "")
    assert result is None

    # Our attributo is of type choice, and we provide a valid string
    result = convert_value(a, chemicals_by_name, "a")
    assert isinstance(result, orm.RunHasAttributoValue)
    assert result.string_value == "a"

    # Our attributo is of type choice, and we provide an invalid string
    result = convert_value(a, chemicals_by_name, "x")
    assert isinstance(result, ConversionError)

    # Our attributo is of type choice, and we provide a float
    result = convert_value(a, chemicals_by_name, 3.5)
    assert isinstance(result, ConversionError)


def test_convert_cell_to_integer() -> None:
    test_wb = Workbook()
    ws = test_wb.active

    assert ws is not None

    ws["A1"] = "3"
    assert excel_import_convert_cell_to_integer(ws["A1"].value) == 3

    ws["A1"] = 3
    assert excel_import_convert_cell_to_integer(ws["A1"].value) == 3

    ws["A1"] = Decimal(3)
    assert excel_import_convert_cell_to_integer(ws["A1"].value) == 3

    ws["A1"] = datetime.datetime.now()
    assert isinstance(
        excel_import_convert_cell_to_integer(ws["A1"].value), ConversionError
    )
    ws["A1"] = datetime.datetime.now().date()
    assert isinstance(
        excel_import_convert_cell_to_integer(ws["A1"].value), ConversionError
    )
    ws["A1"] = datetime.datetime.now().time()
    assert isinstance(
        excel_import_convert_cell_to_integer(ws["A1"].value), ConversionError
    )
    ws["A1"] = datetime.datetime.now() - datetime.datetime.now()
    assert isinstance(
        excel_import_convert_cell_to_integer(ws["A1"].value), ConversionError
    )
    ws["A1"] = DataTableFormula("E11")
    assert isinstance(
        excel_import_convert_cell_to_integer(ws["A1"].value), ConversionError
    )

    ws["A1"] = CellRichText("hehe", TextBlock(InlineFont(b=True), "a"))
    assert isinstance(
        excel_import_convert_cell_to_integer(ws["A1"].value), ConversionError
    )

    ws["A1"] = ArrayFormula("E2:E11", "=SUM(E2)")
    assert isinstance(
        excel_import_convert_cell_to_integer(ws["A1"].value), ConversionError
    )


def test_convert_cell_to_string() -> None:
    test_wb = Workbook()
    ws = test_wb.active

    assert ws is not None

    ws["A1"] = "3"
    assert excel_import_convert_cell_to_string(ws["A1"].value) == "3"

    ws["A1"] = 3
    assert excel_import_convert_cell_to_string(ws["A1"].value) == "3"

    ws["A1"] = Decimal(3)
    assert excel_import_convert_cell_to_string(ws["A1"].value) == "3"

    ws["A1"] = datetime.datetime.now()
    assert isinstance(
        excel_import_convert_cell_to_string(ws["A1"].value), ConversionError
    )
    ws["A1"] = datetime.datetime.now().date()
    assert isinstance(
        excel_import_convert_cell_to_string(ws["A1"].value), ConversionError
    )
    ws["A1"] = datetime.datetime.now().time()
    assert isinstance(
        excel_import_convert_cell_to_string(ws["A1"].value), ConversionError
    )
    ws["A1"] = datetime.datetime.now() - datetime.datetime.now()
    assert isinstance(
        excel_import_convert_cell_to_string(ws["A1"].value), ConversionError
    )
    ws["A1"] = DataTableFormula("E11")
    assert isinstance(
        excel_import_convert_cell_to_string(ws["A1"].value), ConversionError
    )

    ws["A1"] = CellRichText("hehe", TextBlock(InlineFont(b=True), "a"))
    assert isinstance(
        excel_import_convert_cell_to_string(ws["A1"].value), ConversionError
    )

    ws["A1"] = ArrayFormula("E2:E11", "=SUM(E2)")
    assert isinstance(
        excel_import_convert_cell_to_string(ws["A1"].value), ConversionError
    )


def test_run_cell_result() -> None:
    assert isinstance(
        RunCellResult(
            run_id=1,
            experiment_type_name="a",
            started=datetime.datetime.now(),
            stopped=None,
            files=["x"],
            custom_column_values=[],
        ).build_run(1),
        ParsedRunSpreadsheetRun,
    )
    assert isinstance(
        RunCellResult(
            run_id=None,
            experiment_type_name="a",
            started=datetime.datetime.now(),
            stopped=None,
            files=["x"],
            custom_column_values=[],
        ).build_run(1),
        ConversionError,
    )
    assert isinstance(
        RunCellResult(
            run_id=1,
            experiment_type_name=None,
            started=datetime.datetime.now(),
            stopped=None,
            files=["x"],
            custom_column_values=[],
        ).build_run(1),
        ConversionError,
    )
    assert isinstance(
        RunCellResult(
            run_id=1,
            experiment_type_name="a",
            started=None,
            stopped=None,
            files=["x"],
            custom_column_values=[],
        ).build_run(1),
        ConversionError,
    )
    assert isinstance(
        RunCellResult(
            run_id=1,
            experiment_type_name="a",
            started=datetime.datetime.now(),
            stopped=None,
            files=None,
            custom_column_values=[],
        ).build_run(1),
        ConversionError,
    )


@pytest.fixture
def base_spreadsheet() -> Workbook:
    test_wb = Workbook()
    ws = test_wb.active

    assert ws is not None

    ws["A1"] = "run id"
    ws["B1"] = "experiment type"
    ws["C1"] = "started"
    ws["D1"] = "stopped"
    ws["E1"] = "files"
    ws["F1"] = "chemical"
    ws["G1"] = "energy"

    return test_wb


def test_parse_run_spreadsheet_workbook_invalid_mandatory_header() -> None:
    test_wb = Workbook()
    ws = test_wb.active

    assert ws is not None

    # invalid header here
    ws["A1"] = "run2 id"
    ws["B1"] = "experiment type"
    # and here
    ws["C1"] = "started2"
    ws["D1"] = "stopped"
    ws["E1"] = "files"
    ws["F1"] = "chemical"
    ws["G1"] = "energy"

    result = parse_run_spreadsheet_workbook(test_wb)

    assert isinstance(result, ConversionError)


# pylint: disable=redefined-outer-name
def test_parse_run_spreadsheet_single_succesful_run(base_spreadsheet: Workbook) -> None:
    ws = base_spreadsheet.active

    assert ws is not None

    started = datetime.datetime.now()
    ws["A2"] = "1337"
    ws["B2"] = "simple"
    ws["C2"] = started  # type: ignore
    ws["E2"] = "test.h5"
    ws["F2"] = "lyso"
    ws["G2"] = "100"

    # Save to take a look
    # test_wb.save("/tmp/test.xlsx")

    assert parse_run_spreadsheet_workbook(base_spreadsheet) == ParsedRunSpreadsheet(
        custom_column_headers=["chemical", "energy"],
        runs=[
            ParsedRunSpreadsheetRun(
                external_id=1337,
                experiment_type="simple",
                started=started,
                stopped=None,
                original_row=2,
                files=["test.h5"],
                custom_column_values=["lyso", "100"],
            )
        ],
    )


# pylint: disable=redefined-outer-name
def test_parse_run_spreadsheet_run_run_id_is_not_an_integer(
    base_spreadsheet: Workbook,
) -> None:
    ws = base_spreadsheet.active

    assert ws is not None

    started = datetime.datetime.now()
    # simulate a typo in the run ID
    ws["A2"] = "1337a"
    ws["B2"] = "simple"
    ws["C2"] = started  # type: ignore
    ws["E2"] = "test.h5"
    ws["F2"] = "lyso"
    ws["G2"] = "100"

    # Save to take a look
    # test_wb.save("/tmp/test.xlsx")

    result = parse_run_spreadsheet_workbook(base_spreadsheet)
    assert isinstance(result, ConversionError)


# pylint: disable=redefined-outer-name
def test_parse_run_spreadsheet_experiment_type_invalid(
    base_spreadsheet: Workbook,
) -> None:
    ws = base_spreadsheet.active

    assert ws is not None

    started = datetime.datetime.now()
    ws["A2"] = "1337"
    # experiment type is a datetime
    ws["B2"] = datetime.datetime.now()
    ws["C2"] = started
    ws["E2"] = "test.h5"
    ws["F2"] = "lyso"
    ws["G2"] = "100"

    # Save to take a look
    # test_wb.save("/tmp/test.xlsx")

    result = parse_run_spreadsheet_workbook(base_spreadsheet)
    assert isinstance(result, ConversionError)


# pylint: disable=redefined-outer-name
def test_parse_run_spreadsheet_experiment_type_empty(
    base_spreadsheet: Workbook,
) -> None:
    ws = base_spreadsheet.active

    assert ws is not None

    started = datetime.datetime.now()
    ws["A2"] = "1337"
    # experiment type is empty
    ws["B2"] = ""
    ws["C2"] = started
    ws["E2"] = "test.h5"
    ws["F2"] = "lyso"
    ws["G2"] = "100"

    # Save to take a look
    # test_wb.save("/tmp/test.xlsx")

    result = parse_run_spreadsheet_workbook(base_spreadsheet)
    assert isinstance(result, ConversionError)


# pylint: disable=redefined-outer-name
def test_parse_run_spreadsheet_run_started_invalid(
    base_spreadsheet: Workbook,
) -> None:
    ws = base_spreadsheet.active

    assert ws is not None

    ws["A2"] = "1337"
    ws["B2"] = "simple"
    ws["C2"] = "yes"
    ws["E2"] = "test.h5"
    ws["F2"] = "lyso"
    ws["G2"] = "100"

    # Save to take a look
    # test_wb.save("/tmp/test.xlsx")

    result = parse_run_spreadsheet_workbook(base_spreadsheet)
    assert isinstance(result, ConversionError)


# pylint: disable=redefined-outer-name
def test_parse_run_spreadsheet_run_stopped_invalid(
    base_spreadsheet: Workbook,
) -> None:
    ws = base_spreadsheet.active

    assert ws is not None

    started = datetime.datetime.now()
    ws["A2"] = "1337"
    ws["B2"] = "simple"
    ws["C2"] = started
    ws["D2"] = "yes"
    ws["E2"] = "test.h5"
    ws["F2"] = "lyso"
    ws["G2"] = "100"

    # Save to take a look
    # test_wb.save("/tmp/test.xlsx")

    result = parse_run_spreadsheet_workbook(base_spreadsheet)
    assert isinstance(result, ConversionError)


# pylint: disable=redefined-outer-name
def test_parse_run_spreadsheet_files_invalid(
    base_spreadsheet: Workbook,
) -> None:
    ws = base_spreadsheet.active

    assert ws is not None

    started = datetime.datetime.now()
    ws["A2"] = "1337"
    ws["B2"] = "simple"
    ws["C2"] = started
    ws["D2"] = started
    ws["E2"] = started
    ws["F2"] = "lyso"
    ws["G2"] = "100"

    # Save to take a look
    # test_wb.save("/tmp/test.xlsx")

    result = parse_run_spreadsheet_workbook(base_spreadsheet)
    assert isinstance(result, ConversionError)


# pylint: disable=redefined-outer-name
def test_parse_run_spreadsheet_extraneous_custom_cells(
    base_spreadsheet: Workbook,
) -> None:
    ws = base_spreadsheet.active

    assert ws is not None

    started = datetime.datetime.now()
    ws["A2"] = "1337"
    ws["B2"] = "simple"
    ws["C2"] = started  # type: ignore
    ws["E2"] = "test.h5"
    ws["F2"] = "lyso"
    ws["G2"] = "100"

    # this cell shouldn't be there
    ws["H2"] = "-10"

    result = parse_run_spreadsheet_workbook(base_spreadsheet)
    assert isinstance(result, ConversionError)
    assert len(result.error_messages) == 1


def test_create_runs_from_spreadsheet_no_custom_attributi() -> None:
    beamtime_id = BeamtimeId(1)
    experiment_type_id = 1338
    experiment_type = orm.ExperimentType(name="simple", beamtime_id=beamtime_id)
    experiment_type.id = experiment_type_id
    run_external_id = 1339
    run_started = datetime.datetime.now()
    result = create_runs_from_spreadsheet(
        spreadsheet=ParsedRunSpreadsheet(
            custom_column_headers=[],
            runs=[
                ParsedRunSpreadsheetRun(
                    external_id=run_external_id,
                    experiment_type="simple",
                    started=run_started,
                    stopped=None,
                    original_row=1,
                    files=["test.h5"],
                    custom_column_values=[],
                )
            ],
        ),
        beamtime_id=beamtime_id,
        attributi=[],
        chemicals=[],
        existing_runs=[],
        experiment_types=[experiment_type],
    )
    assert isinstance(result, SpreadsheetValidationResult)
    assert len(result.runs) == 1
    first_run = result.runs[0]
    assert first_run.beamtime_id == beamtime_id
    assert first_run.experiment_type_id == experiment_type_id
    assert first_run.external_id == run_external_id
    assert first_run.started == run_started
    assert len(first_run.files) == 1
    assert first_run.files[0].glob == "test.h5"


def test_create_runs_from_spreadsheet_one_custom_int_attributo() -> None:
    beamtime_id = BeamtimeId(1)
    experiment_type_id = 1338
    experiment_type = orm.ExperimentType(name="simple", beamtime_id=beamtime_id)
    experiment_type.id = experiment_type_id
    run_external_id = 1339
    run_started = datetime.datetime.now()
    test_attributo = orm.Attributo(
        name="test",
        json_schema={"type": "integer"},
        beamtime_id=beamtime_id,
        description="",
        group="",
        associated_table=AssociatedTable.RUN,
    )
    test_attributo_id = 1400
    test_attributo.id = test_attributo_id  # type: ignore
    result = create_runs_from_spreadsheet(
        spreadsheet=ParsedRunSpreadsheet(
            custom_column_headers=["test"],
            runs=[
                ParsedRunSpreadsheetRun(
                    external_id=run_external_id,
                    experiment_type="simple",
                    started=run_started,
                    stopped=None,
                    original_row=1,
                    files=["test.h5"],
                    custom_column_values=["3"],
                )
            ],
        ),
        beamtime_id=beamtime_id,
        attributi=[test_attributo],
        chemicals=[],
        existing_runs=[],
        experiment_types=[experiment_type],
    )
    assert isinstance(result, SpreadsheetValidationResult)
    assert len(result.runs) == 1
    first_run = result.runs[0]
    assert first_run.attributo_values[0].attributo_id == test_attributo_id
    assert first_run.attributo_values[0].integer_value == 3
    assert first_run.beamtime_id == beamtime_id
    assert first_run.experiment_type_id == experiment_type_id
    assert first_run.external_id == run_external_id
    assert first_run.started == run_started


def test_create_runs_from_spreadsheet_duplicate_run_ids_inside_spreadsheet() -> None:
    beamtime_id = BeamtimeId(1)
    experiment_type_id = 1338
    experiment_type = orm.ExperimentType(name="simple", beamtime_id=beamtime_id)
    experiment_type.id = experiment_type_id
    run_external_id = 1339
    run_started = datetime.datetime.now()
    test_attributo = orm.Attributo(
        name="test",
        json_schema={"type": "integer"},
        beamtime_id=beamtime_id,
        description="",
        group="",
        associated_table=AssociatedTable.RUN,
    )
    test_attributo_id = 1400
    test_attributo.id = test_attributo_id  # type: ignore
    result = create_runs_from_spreadsheet(
        spreadsheet=ParsedRunSpreadsheet(
            custom_column_headers=["test"],
            runs=[
                ParsedRunSpreadsheetRun(
                    external_id=run_external_id,
                    experiment_type="simple",
                    started=run_started,
                    stopped=None,
                    original_row=1,
                    files=["test.h5"],
                    custom_column_values=["3"],
                ),
                ParsedRunSpreadsheetRun(
                    # see? same ID as before - not goot
                    external_id=run_external_id,
                    experiment_type="simple",
                    started=run_started,
                    stopped=None,
                    original_row=1,
                    files=["test.h5"],
                    custom_column_values=["3"],
                ),
            ],
        ),
        beamtime_id=beamtime_id,
        attributi=[test_attributo],
        chemicals=[],
        existing_runs=[],
        experiment_types=[experiment_type],
    )
    assert isinstance(result, SpreadsheetValidationErrors)


def test_data_sets_are_equal() -> None:
    assert data_sets_are_equal(
        orm.DataSet(experiment_type_id=1, attributo_values=[]),
        orm.DataSet(experiment_type_id=1, attributo_values=[]),
    )
    assert not data_sets_are_equal(
        orm.DataSet(experiment_type_id=1, attributo_values=[]),
        orm.DataSet(experiment_type_id=2, attributo_values=[]),
    )

    # Difference: one attributo each, but different ID
    assert not data_sets_are_equal(
        orm.DataSet(
            experiment_type_id=1,
            attributo_values=[
                orm.DataSetHasAttributoValue(
                    attributo_id=AttributoId(1),
                    integer_value=None,
                    float_value=None,
                    string_value="a",
                    bool_value=None,
                    datetime_value=None,
                    list_value=None,
                    chemical_value=None,
                )
            ],
        ),
        orm.DataSet(
            experiment_type_id=1,
            attributo_values=[
                orm.DataSetHasAttributoValue(
                    attributo_id=AttributoId(2),
                    integer_value=None,
                    float_value=None,
                    string_value="a",
                    bool_value=None,
                    datetime_value=None,
                    list_value=None,
                    chemical_value=None,
                )
            ],
        ),
    )

    # One attributo each, same value
    assert data_sets_are_equal(
        orm.DataSet(
            experiment_type_id=1,
            attributo_values=[
                orm.DataSetHasAttributoValue(
                    attributo_id=AttributoId(1),
                    integer_value=None,
                    float_value=None,
                    string_value="a",
                    bool_value=None,
                    datetime_value=None,
                    list_value=None,
                    chemical_value=None,
                )
            ],
        ),
        orm.DataSet(
            experiment_type_id=1,
            attributo_values=[
                orm.DataSetHasAttributoValue(
                    attributo_id=AttributoId(1),
                    integer_value=None,
                    float_value=None,
                    string_value="a",
                    bool_value=None,
                    datetime_value=None,
                    list_value=None,
                    chemical_value=None,
                )
            ],
        ),
    )

    # Difference: one attributo each, different value
    assert not data_sets_are_equal(
        orm.DataSet(
            experiment_type_id=1,
            attributo_values=[
                orm.DataSetHasAttributoValue(
                    attributo_id=AttributoId(1),
                    integer_value=None,
                    float_value=None,
                    string_value="a",
                    bool_value=None,
                    datetime_value=None,
                    list_value=None,
                    chemical_value=None,
                )
            ],
        ),
        orm.DataSet(
            experiment_type_id=1,
            attributo_values=[
                orm.DataSetHasAttributoValue(
                    attributo_id=AttributoId(1),
                    integer_value=None,
                    float_value=None,
                    string_value="b",
                    bool_value=None,
                    datetime_value=None,
                    list_value=None,
                    chemical_value=None,
                )
            ],
        ),
    )


def test_create_data_set_for_runs() -> None:
    beamtime_id = BeamtimeId(1)
    et1 = orm.ExperimentType(
        name="simple",
        beamtime_id=beamtime_id,
        attributi=[
            orm.ExperimentHasAttributo(
                attributo_id=AttributoId(1), chemical_role=ChemicalType.CRYSTAL
            )
        ],
    )
    et1.id = 1
    run1 = orm.Run(
        external_id=RunExternalId(1),
        beamtime_id=beamtime_id,
        modified=datetime.datetime.now(),
        started=datetime.datetime.now(),
        stopped=datetime.datetime.now(),
        experiment_type_id=et1.id,
        attributo_values=[
            orm.RunHasAttributoValue(
                attributo_id=AttributoId(1),
                integer_value=1,
                float_value=None,
                string_value=None,
                bool_value=None,
                datetime_value=None,
                list_value=None,
                chemical_value=None,
            )
        ],
    )
    run2 = orm.Run(
        external_id=RunExternalId(1),
        beamtime_id=beamtime_id,
        modified=datetime.datetime.now(),
        started=datetime.datetime.now(),
        stopped=datetime.datetime.now(),
        experiment_type_id=et1.id,
        attributo_values=[
            orm.RunHasAttributoValue(
                attributo_id=AttributoId(1),
                integer_value=1,
                float_value=None,
                string_value=None,
                bool_value=None,
                datetime_value=None,
                list_value=None,
                chemical_value=None,
            )
        ],
    )
    run1.experiment_type = et1
    run2.experiment_type = et1
    data_sets = create_data_set_for_runs([et1], [run1, run2], [])
    assert isinstance(data_sets, list)
    assert len(data_sets) == 1


def test_create_runs_from_spreadsheet_duplicate_file_glob() -> None:
    beamtime_id = BeamtimeId(1)
    experiment_type_id = 1338
    experiment_type = orm.ExperimentType(name="simple", beamtime_id=beamtime_id)
    experiment_type.id = experiment_type_id
    run_external_id = 1339
    run_started = datetime.datetime.now()
    result = create_runs_from_spreadsheet(
        spreadsheet=ParsedRunSpreadsheet(
            custom_column_headers=[],
            runs=[
                ParsedRunSpreadsheetRun(
                    external_id=run_external_id,
                    experiment_type="simple",
                    started=run_started,
                    stopped=None,
                    original_row=1,
                    files=["test.h5"],
                    custom_column_values=[],
                ),
                ParsedRunSpreadsheetRun(
                    external_id=run_external_id + 1,
                    experiment_type="simple",
                    started=run_started,
                    stopped=None,
                    original_row=2,
                    # see here? that's a duplicate
                    files=["test.h5"],
                    custom_column_values=[],
                ),
            ],
        ),
        beamtime_id=beamtime_id,
        attributi=[],
        chemicals=[],
        existing_runs=[],
        experiment_types=[experiment_type],
    )
    assert isinstance(result, SpreadsheetValidationResult)
    assert result.warnings


def test_create_runs_from_spreadsheet_duplicate_run_ids_outside_spreadsheet() -> None:
    beamtime_id = BeamtimeId(1)
    experiment_type_id = 1338
    experiment_type = orm.ExperimentType(name="simple", beamtime_id=beamtime_id)
    experiment_type.id = experiment_type_id
    run_external_id = 1339
    run_started = datetime.datetime.now()
    test_attributo = orm.Attributo(
        name="test",
        json_schema={"type": "integer"},
        beamtime_id=beamtime_id,
        description="",
        group="",
        associated_table=AssociatedTable.RUN,
    )
    test_attributo_id = 1400
    test_attributo.id = test_attributo_id  # type: ignore
    result = create_runs_from_spreadsheet(
        spreadsheet=ParsedRunSpreadsheet(
            custom_column_headers=["test"],
            runs=[
                ParsedRunSpreadsheetRun(
                    external_id=run_external_id,
                    experiment_type="simple",
                    started=run_started,
                    stopped=None,
                    original_row=1,
                    files=["test.h5"],
                    custom_column_values=["3"],
                )
            ],
        ),
        beamtime_id=beamtime_id,
        attributi=[test_attributo],
        chemicals=[],
        existing_runs=[
            orm.Run(
                # same ID as the one in the spreadsheet!
                external_id=RunExternalId(run_external_id),
                beamtime_id=beamtime_id,
                modified=datetime.datetime.now(),
                started=datetime.datetime.now(),
                stopped=datetime.datetime.now(),
                experiment_type_id=experiment_type_id,
                attributo_values=[
                    orm.RunHasAttributoValue(
                        attributo_id=AttributoId(test_attributo_id),
                        integer_value=1,
                        float_value=None,
                        string_value=None,
                        bool_value=None,
                        datetime_value=None,
                        list_value=None,
                        chemical_value=None,
                    )
                ],
            )
        ],
        experiment_types=[experiment_type],
    )
    assert isinstance(result, SpreadsheetValidationErrors)

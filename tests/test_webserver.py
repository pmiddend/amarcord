# working with these fixtures, this pylint error message doesn't make sense anymore
# ruff: noqa: T201

import asyncio
import os
from dataclasses import replace
from functools import partial
from io import BytesIO
from pathlib import Path
from typing import Any
from typing import AsyncGenerator
from typing import Generator
from unittest import TestCase
from zipfile import ZipFile

import aiohttp
import pytest
from aiohttp import web
from fastapi import UploadFile
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.ext.asyncio import create_async_engine

from amarcord.amici.workload_manager.dummy_workload_manager import DummyWorkloadManager
from amarcord.amici.workload_manager.job import JobMetadata
from amarcord.amici.workload_manager.job_status import JobStatus
from amarcord.amici.workload_manager.workload_manager import JobStartResult
from amarcord.cli import indexing_daemon
from amarcord.cli import merge_daemon
from amarcord.cli.indexing_daemon import (
    INDEXING_DAEMON_LONG_BREAK_DURATION_SECONDS_ENV_VAR,
)
from amarcord.cli.indexing_daemon import indexing_loop_iteration
from amarcord.cli.merge_daemon import MERGE_DAEMON_LONG_BREAK_DURATION_SECONDS_ENV_VAR
from amarcord.cli.merge_daemon import MERGE_DAEMON_SHORT_BREAK_DURATION_SECONDS_ENV_VAR
from amarcord.cli.merge_daemon import merging_loop_iteration
from amarcord.cli.webserver import app
from amarcord.db.associated_table import AssociatedTable
from amarcord.db.beamtime_id import BeamtimeId
from amarcord.db.chemical_type import ChemicalType
from amarcord.db.db_job_status import DBJobStatus
from amarcord.db.merge_model import MergeModel
from amarcord.db.merge_negative_handling import MergeNegativeHandling
from amarcord.db.merge_result import JsonMergeJobFinishedInput
from amarcord.db.merge_result import JsonMergeJobStartedInput
from amarcord.db.merge_result import JsonMergeJobStartedOutput
from amarcord.db.merge_result import JsonMergeResultFom
from amarcord.db.merge_result import JsonMergeResultInternal
from amarcord.db.merge_result import JsonMergeResultOuterShell
from amarcord.db.merge_result import JsonMergeResultShell
from amarcord.db.merge_result import JsonRefinementResultInternal
from amarcord.db.orm_utils import ATTRIBUTO_GROUP_MANUAL
from amarcord.db.orm_utils import CompressionMode
from amarcord.db.orm_utils import live_stream_image_name
from amarcord.db.orm_utils import migrate
from amarcord.db.run_internal_id import RunInternalId
from amarcord.db.scale_intensities import ScaleIntensities
from amarcord.json_schema import JSONSchemaInteger
from amarcord.json_schema import JSONSchemaString
from amarcord.web.fastapi_utils import get_orm_sessionmaker_with_url
from amarcord.web.json_models import JsonAlignDetectorGroup
from amarcord.web.json_models import JsonAttributiIdAndRole
from amarcord.web.json_models import JsonAttributo
from amarcord.web.json_models import JsonAttributoValue
from amarcord.web.json_models import JsonBeamtimeOutput
from amarcord.web.json_models import JsonBeamtimeScheduleOutput
from amarcord.web.json_models import JsonBeamtimeScheduleRowInput
from amarcord.web.json_models import JsonChangeRunExperimentType
from amarcord.web.json_models import JsonChangeRunExperimentTypeOutput
from amarcord.web.json_models import JsonCheckStandardUnitInput
from amarcord.web.json_models import JsonCheckStandardUnitOutput
from amarcord.web.json_models import JsonChemicalWithId
from amarcord.web.json_models import JsonChemicalWithoutId
from amarcord.web.json_models import JsonCopyChemicalInput
from amarcord.web.json_models import JsonCopyChemicalOutput
from amarcord.web.json_models import JsonCopyExperimentTypesInput
from amarcord.web.json_models import JsonCopyExperimentTypesOutput
from amarcord.web.json_models import JsonCreateAttributiFromSchemaInput
from amarcord.web.json_models import JsonCreateAttributiFromSchemaOutput
from amarcord.web.json_models import JsonCreateAttributiFromSchemaSingleAttributo
from amarcord.web.json_models import JsonCreateAttributoInput
from amarcord.web.json_models import JsonCreateAttributoOutput
from amarcord.web.json_models import JsonCreateChemicalOutput
from amarcord.web.json_models import JsonCreateDataSetFromRun
from amarcord.web.json_models import JsonCreateDataSetFromRunOutput
from amarcord.web.json_models import JsonCreateDataSetInput
from amarcord.web.json_models import JsonCreateDataSetOutput
from amarcord.web.json_models import JsonCreateExperimentTypeInput
from amarcord.web.json_models import JsonCreateExperimentTypeOutput
from amarcord.web.json_models import JsonCreateFileOutput
from amarcord.web.json_models import JsonCreateIndexingForDataSetInput
from amarcord.web.json_models import JsonCreateIndexingForDataSetOutput
from amarcord.web.json_models import JsonCreateLiveStreamSnapshotOutput
from amarcord.web.json_models import JsonCreateOrUpdateRun
from amarcord.web.json_models import JsonCreateOrUpdateRunOutput
from amarcord.web.json_models import JsonDeleteAttributoInput
from amarcord.web.json_models import JsonDeleteAttributoOutput
from amarcord.web.json_models import JsonDeleteChemicalInput
from amarcord.web.json_models import JsonDeleteDataSetInput
from amarcord.web.json_models import JsonDeleteDataSetOutput
from amarcord.web.json_models import JsonDeleteEventInput
from amarcord.web.json_models import JsonDeleteExperimentType
from amarcord.web.json_models import JsonDeleteExperimentTypeOutput
from amarcord.web.json_models import JsonDeleteFileInput
from amarcord.web.json_models import JsonDeleteFileOutput
from amarcord.web.json_models import JsonDeleteRunOutput
from amarcord.web.json_models import JsonEventInput
from amarcord.web.json_models import JsonEventTopLevelInput
from amarcord.web.json_models import JsonEventTopLevelOutput
from amarcord.web.json_models import JsonGeometryCopyToBeamtime
from amarcord.web.json_models import JsonGeometryCreate
from amarcord.web.json_models import JsonGeometryUpdate
from amarcord.web.json_models import JsonGeometryWithoutContent
from amarcord.web.json_models import JsonImportFinishedIndexingJobInput
from amarcord.web.json_models import JsonImportFinishedIndexingJobOutput
from amarcord.web.json_models import JsonIndexingJobUpdateOutput
from amarcord.web.json_models import JsonIndexingResultFinishSuccessfully
from amarcord.web.json_models import JsonIndexingResultFinishWithError
from amarcord.web.json_models import JsonIndexingResultStillRunning
from amarcord.web.json_models import JsonMergeJobFinishOutput
from amarcord.web.json_models import JsonMergeParameters
from amarcord.web.json_models import JsonMergeStatus
from amarcord.web.json_models import JsonPolarisation
from amarcord.web.json_models import JsonQueueMergeJobInput
from amarcord.web.json_models import JsonQueueMergeJobOutput
from amarcord.web.json_models import JsonReadAttributi
from amarcord.web.json_models import JsonReadBeamtime
from amarcord.web.json_models import JsonReadChemicals
from amarcord.web.json_models import JsonReadDataSets
from amarcord.web.json_models import JsonReadEvents
from amarcord.web.json_models import JsonReadExperimentTypes
from amarcord.web.json_models import JsonReadGeometriesForAllBeamtimes
from amarcord.web.json_models import JsonReadGeometriesForSingleBeamtime
from amarcord.web.json_models import JsonReadIndexingResultsOutput
from amarcord.web.json_models import JsonReadMergeResultsOutput
from amarcord.web.json_models import JsonReadNewAnalysisInput
from amarcord.web.json_models import JsonReadNewAnalysisOutput
from amarcord.web.json_models import JsonReadRuns
from amarcord.web.json_models import JsonReadRunsBulkInput
from amarcord.web.json_models import JsonReadRunsBulkOutput
from amarcord.web.json_models import JsonReadRunsOverview
from amarcord.web.json_models import JsonReadSingleDataSetResults
from amarcord.web.json_models import JsonReadSingleGeometryOutput
from amarcord.web.json_models import JsonRefinementResult
from amarcord.web.json_models import JsonRunFile
from amarcord.web.json_models import JsonStartRunOutput
from amarcord.web.json_models import JsonStopRunOutput
from amarcord.web.json_models import JsonUpdateAttributoConversionFlags
from amarcord.web.json_models import JsonUpdateAttributoInput
from amarcord.web.json_models import JsonUpdateAttributoOutput
from amarcord.web.json_models import JsonUpdateBeamtimeInput
from amarcord.web.json_models import JsonUpdateBeamtimeScheduleInput
from amarcord.web.json_models import JsonUpdateLiveStream
from amarcord.web.json_models import JsonUpdateRun
from amarcord.web.json_models import JsonUpdateRunOutput
from amarcord.web.json_models import JsonUpdateRunsBulkInput
from amarcord.web.json_models import JsonUpdateRunsBulkOutput
from amarcord.web.json_models import JsonUserConfigurationSingleOutput
from amarcord.web.router_files import create_file
from amarcord.web.router_indexing import indexing_job_finish_with_error
from amarcord.web.router_indexing import indexing_job_still_running
from amarcord.web.router_indexing import read_indexing_jobs
from amarcord.web.router_merging import merge_job_finished
from amarcord.web.router_merging import merge_job_started
from amarcord.web.router_merging import read_merge_jobs

IN_MEMORY_DB_URL = "sqlite+aiosqlite://"

TEST_ATTRIBUTO_NAME = "testattributo"
TEST_ATTRIBUTO_NAME2 = "testattributo2"
TEST_CHEMICAL_NAME = "chemicalname"
TEST_CHEMICAL_RESPONSIBLE_PERSON = "Rosalind Franklin"
_RUN_STRING_ATTRIBUTO_NAME = "run_string"
_RUN_INT_ATTRIBUTO_NAME = "run_int"

# Some of the tests compare ".local" time fields, which then depends
# on the time zone you're using to run the test, which is suboptimal
os.environ["AMARCORD_TZ"] = "UTC"


async def init_db(url: str) -> None:
    engine = create_async_engine(url)
    await migrate(engine)


@pytest.fixture
async def async_session(tmp_path: Path) -> AsyncGenerator[AsyncSession, None]:
    url = f"{IN_MEMORY_DB_URL}/{tmp_path}/db"
    os.environ["DB_URL"] = url
    await init_db(url)

    try:
        result = get_orm_sessionmaker_with_url(os.environ["DB_URL"])

        async with result() as session:
            print("------------------- yielding session")
            yield session
        print("------------------- yielding session done")

    except:  # noqa: S110
        pass


@pytest.fixture
def client(tmp_path: Path) -> Generator[TestClient, None, None]:
    url = f"{IN_MEMORY_DB_URL}/{tmp_path}/db"
    os.environ["DB_URL"] = url
    asyncio.run(init_db(url))
    yield TestClient(app)


def create_beamtime(client: TestClient, input_: JsonUpdateBeamtimeInput) -> BeamtimeId:
    response = JsonBeamtimeOutput(
        **client.post(
            "/api/beamtimes",
            json=input_.model_dump(),
        ).json(),
    )
    assert response.id > 0
    return BeamtimeId(response.id)


@pytest.fixture
def beamtime_id(client: TestClient) -> BeamtimeId:
    return create_beamtime(
        client,
        JsonUpdateBeamtimeInput(
            id=BeamtimeId(0),
            external_id="cool 1337",
            beamline="P12",
            proposal="BAG",
            title="Test beamtime",
            comment="comment",
            start_local=1,
            end_local=1000,
            analysis_output_path="/",
        ),
    )


@pytest.fixture
def second_beamtime_id(client: TestClient) -> BeamtimeId:
    return create_beamtime(
        client,
        JsonUpdateBeamtimeInput(
            id=BeamtimeId(0),
            external_id="second external",
            beamline="P13",
            proposal="BAG2",
            title="Test beamtime2",
            comment="comment2",
            start_local=1,
            end_local=1000,
            analysis_output_path="/",
        ),
    )


# Stored in a variable so when we want to update it in a test, we can
# reuse the old content (changing just the metadata)
_GEOMETRY_ID_CONTENT = "hehe"


@pytest.fixture
def geometry_id(client: TestClient, beamtime_id: BeamtimeId) -> int:
    response = JsonGeometryWithoutContent(
        **client.post(
            "/api/geometries",
            json=JsonGeometryCreate(
                beamtime_id=beamtime_id,
                content=_GEOMETRY_ID_CONTENT,
                name="geometry name",
            ).model_dump(),
        ).json(),
    )
    assert response.id > 0
    return response.id


@pytest.fixture
def geometry_id_with_run_string_attributo(
    client: TestClient, beamtime_id: BeamtimeId, run_string_attributo_id: int
) -> int:
    response = JsonGeometryWithoutContent(
        **client.post(
            "/api/geometries",
            json=JsonGeometryCreate(
                beamtime_id=beamtime_id,
                content="clen {{" + _RUN_STRING_ATTRIBUTO_NAME + "}}",
                name="geometry name",
            ).model_dump(),
        ).json(),
    )
    assert response.id > 0
    assert response.attributi == [run_string_attributo_id]
    return response.id


@pytest.fixture
def test_file_path() -> Path:
    # This is just some lame txt file, but it's enough for our purposes of testing. We don't need _actual_ files.
    return Path(__file__).parent / "test-file.txt"


@pytest.fixture
def test_file(client: TestClient, test_file_path: Path) -> int:
    # Upload the file
    with test_file_path.open("rb") as upload_file:
        raw_output = client.post(
            "/api/files",
            data={"description": "test description", "deduplicate": str(False)},
            files={"file": upload_file},
        )
        output = JsonCreateFileOutput(**raw_output.json())

        assert output.file_name == "test-file.txt"
        assert output.description == "test description"
        assert output.type_ == "text/plain"
        assert output.size_in_bytes == 17
        return output.id


@pytest.fixture
def second_test_file(client: TestClient, test_file_path: Path) -> int:
    # Upload the file
    with test_file_path.open("rb") as upload_file:
        raw_output = client.post(
            "/api/files",
            data={"description": "test 2 description", "deduplicate": str(False)},
            files={"file": upload_file},
        )
        output = JsonCreateFileOutput(**raw_output.json())
        return output.id


@pytest.fixture
def cell_description_attributo_id(client: TestClient, beamtime_id: BeamtimeId) -> int:
    input_ = JsonCreateAttributoInput(
        name="cell description",
        description="description",
        group=ATTRIBUTO_GROUP_MANUAL,
        associated_table=AssociatedTable.CHEMICAL,
        attributo_type_string=JSONSchemaString(type="string", enum=None),
        beamtime_id=beamtime_id,
    ).model_dump()
    response = JsonCreateAttributoOutput(
        **client.post(
            "/api/attributi",
            json=input_,
        ).json(),
    )
    assert response.id > 0
    return response.id


@pytest.fixture
def point_group_attributo_id(client: TestClient, beamtime_id: BeamtimeId) -> int:
    input_ = JsonCreateAttributoInput(
        name="point group",
        description="description",
        group=ATTRIBUTO_GROUP_MANUAL,
        associated_table=AssociatedTable.CHEMICAL,
        attributo_type_string=JSONSchemaString(type="string", enum=None),
        beamtime_id=beamtime_id,
    ).model_dump()
    response_json = client.post(
        "/api/attributi",
        json=input_,
    ).json()
    response = JsonCreateAttributoOutput(**response_json)
    return response.id


@pytest.fixture
def space_group_attributo_id(client: TestClient, beamtime_id: BeamtimeId) -> int:
    input_ = JsonCreateAttributoInput(
        name="space group",
        description="description",
        group=ATTRIBUTO_GROUP_MANUAL,
        associated_table=AssociatedTable.CHEMICAL,
        attributo_type_string=JSONSchemaString(type="string", enum=None),
        beamtime_id=beamtime_id,
    ).model_dump()
    response_json = client.post(
        "/api/attributi",
        json=input_,
    ).json()
    response = JsonCreateAttributoOutput(**response_json)
    return response.id


@pytest.fixture
def run_string_attributo_id(client: TestClient, beamtime_id: BeamtimeId) -> int:
    input_ = JsonCreateAttributoInput(
        name=_RUN_STRING_ATTRIBUTO_NAME,
        description="",
        group=ATTRIBUTO_GROUP_MANUAL,
        associated_table=AssociatedTable.RUN,
        attributo_type_string=JSONSchemaString(type="string", enum=None),
        beamtime_id=beamtime_id,
    ).model_dump()
    response = JsonCreateAttributoOutput(
        **client.post(
            "/api/attributi",
            json=input_,
        ).json(),
    )
    assert response.id > 0
    return response.id


@pytest.fixture
def run_int_attributo_id(client: TestClient, beamtime_id: BeamtimeId) -> int:
    input_ = JsonCreateAttributoInput(
        name=_RUN_INT_ATTRIBUTO_NAME,
        description="",
        group=ATTRIBUTO_GROUP_MANUAL,
        associated_table=AssociatedTable.RUN,
        attributo_type_integer=JSONSchemaInteger(type="integer", format=None),
        beamtime_id=beamtime_id,
    ).model_dump()
    response = JsonCreateAttributoOutput(
        **client.post(
            "/api/attributi",
            json=input_,
        ).json(),
    )
    assert response.id > 0
    return response.id


@pytest.fixture
def run_int_automatic_attributo_id(client: TestClient, beamtime_id: BeamtimeId) -> int:
    input_ = JsonCreateAttributoInput(
        name="run_int_automatic",
        description="",
        group="automatic",
        associated_table=AssociatedTable.RUN,
        attributo_type_integer=JSONSchemaInteger(type="integer", format=None),
        beamtime_id=beamtime_id,
    ).model_dump()
    response = JsonCreateAttributoOutput(
        **client.post(
            "/api/attributi",
            json=input_,
        ).json(),
    )
    assert response.id > 0
    return response.id


@pytest.fixture
def run_channel_1_chemical_attributo_id(
    client: TestClient,
    beamtime_id: BeamtimeId,
) -> int:
    input_ = JsonCreateAttributoInput(
        name="channel_1_chemical_id",
        description="",
        group=ATTRIBUTO_GROUP_MANUAL,
        associated_table=AssociatedTable.RUN,
        attributo_type_integer=JSONSchemaInteger(type="integer", format="chemical-id"),
        beamtime_id=beamtime_id,
    ).model_dump()
    response = JsonCreateAttributoOutput(
        **client.post(
            "/api/attributi",
            json=input_,
        ).json(),
    )
    assert response.id > 0
    return response.id


@pytest.fixture
def run_channel_1_chemical_attributo_id_in_second_beamtime(
    client: TestClient,
    second_beamtime_id: BeamtimeId,
) -> int:
    input_ = JsonCreateAttributoInput(
        name="channel_1_chemical_id",
        description="",
        group=ATTRIBUTO_GROUP_MANUAL,
        associated_table=AssociatedTable.RUN,
        attributo_type_integer=JSONSchemaInteger(type="integer", format="chemical-id"),
        beamtime_id=second_beamtime_id,
    ).model_dump()
    response = JsonCreateAttributoOutput(
        **client.post(
            "/api/attributi",
            json=input_,
        ).json(),
    )
    assert response.id > 0
    return response.id


@pytest.fixture
def chemical_experiment_type_id(
    client: TestClient,
    beamtime_id: BeamtimeId,
    run_channel_1_chemical_attributo_id: int,
) -> int:
    input_ = JsonCreateExperimentTypeInput(
        name="experiment type test",
        beamtime_id=beamtime_id,
        attributi=[
            JsonAttributiIdAndRole(
                id=run_channel_1_chemical_attributo_id,
                role=ChemicalType.CRYSTAL,
            ),
        ],
    ).model_dump()
    response = JsonCreateExperimentTypeOutput(
        **client.post(
            "/api/experiment-types",
            json=input_,
        ).json(),
    )
    assert response.id > 0
    return response.id


@pytest.fixture
def string_experiment_type_id(
    client: TestClient,
    beamtime_id: BeamtimeId,
    run_string_attributo_id: int,
) -> int:
    input_ = JsonCreateExperimentTypeInput(
        name="experiment type test",
        beamtime_id=beamtime_id,
        attributi=[
            JsonAttributiIdAndRole(
                id=run_string_attributo_id,
                role=ChemicalType.CRYSTAL,
            ),
        ],
    ).model_dump()
    response = JsonCreateExperimentTypeOutput(
        **client.post(
            "/api/experiment-types",
            json=input_,
        ).json(),
    )
    assert response.id > 0
    return response.id


# This is derliberately "90.0" instead of "90" because one of the processing steps normalizes it to 90.0, and
# if we have it as 90.0 anyways comparison is easier.
LYSO_CELL_DESCRIPTION = "tetragonal P c (79.2 79.2 38.0) (90.0 90.0 90.0)"
LYSO_POINT_GROUP = "4/mmm"
# Might not even be correct, but this is a test, who cares
LYSO_SPACE_GROUP = "P 1 21 1"


@pytest.fixture
def lyso_chemical_id(
    client: TestClient,
    beamtime_id: BeamtimeId,
    cell_description_attributo_id: int,
    point_group_attributo_id: int,
    space_group_attributo_id: int,
    test_file: int,
) -> int:
    response = JsonCreateChemicalOutput(
        **client.post(
            "/api/chemicals",
            json=JsonChemicalWithoutId(
                name=TEST_CHEMICAL_NAME,
                responsible_person=TEST_CHEMICAL_RESPONSIBLE_PERSON,
                attributi=[
                    JsonAttributoValue(
                        attributo_id=cell_description_attributo_id,
                        # Good old Lyso!
                        attributo_value_str=LYSO_CELL_DESCRIPTION,
                    ),
                    JsonAttributoValue(
                        attributo_id=point_group_attributo_id,
                        # Good old Lyso!
                        attributo_value_str=LYSO_POINT_GROUP,
                    ),
                    JsonAttributoValue(
                        attributo_id=space_group_attributo_id,
                        attributo_value_str=LYSO_SPACE_GROUP,
                    ),
                ],
                chemical_type=ChemicalType.CRYSTAL,
                file_ids=[test_file],
                beamtime_id=beamtime_id,
            ).model_dump(),
        ).json(),
    )
    assert response.id > 0
    return response.id


@pytest.fixture
def second_lyso_chemical_id(
    client: TestClient,
    beamtime_id: BeamtimeId,
    cell_description_attributo_id: int,
    point_group_attributo_id: int,
    space_group_attributo_id: int,
) -> int:
    response = JsonCreateChemicalOutput(
        **client.post(
            "/api/chemicals",
            json=JsonChemicalWithoutId(
                name=TEST_CHEMICAL_NAME + "2",
                responsible_person=TEST_CHEMICAL_RESPONSIBLE_PERSON + "2",
                attributi=[
                    JsonAttributoValue(
                        attributo_id=cell_description_attributo_id,
                        attributo_value_str="foo3",
                    ),
                    JsonAttributoValue(
                        attributo_id=point_group_attributo_id,
                        attributo_value_str="bar3",
                    ),
                    JsonAttributoValue(
                        attributo_id=space_group_attributo_id,
                        attributo_value_str="baz3",
                    ),
                ],
                chemical_type=ChemicalType.CRYSTAL,
                file_ids=[],
                beamtime_id=beamtime_id,
            ).model_dump(),
        ).json(),
    )
    assert response.id > 0
    return response.id


@pytest.fixture
def simple_run_id(
    client: TestClient,
    beamtime_id: BeamtimeId,
    run_channel_1_chemical_attributo_id: int,
    chemical_experiment_type_id: int,
    run_string_attributo_id: int,
    lyso_chemical_id: int,
) -> RunInternalId:
    external_run_id = 1000

    set_current_experiment_type(client, beamtime_id, chemical_experiment_type_id)

    create_run_response = JsonCreateOrUpdateRunOutput(
        **client.post(
            f"/api/runs/{external_run_id}",
            json=JsonCreateOrUpdateRun(
                # Important for our indexing job tests: we cannot start an offline indexing job without files (i.e. images).
                files=[JsonRunFile(id=0, glob="/tmp/test-input-file", source="raw")],  # noqa: S108
                beamtime_id=beamtime_id,
                attributi=[
                    JsonAttributoValue(
                        attributo_id=run_channel_1_chemical_attributo_id,
                        attributo_value_chemical=lyso_chemical_id,
                    ),
                    JsonAttributoValue(
                        attributo_id=run_string_attributo_id,
                        attributo_value_str="foobar",
                    ),
                ],
                create_data_set=False,
                started=1,
                stopped=None,
            ).model_dump(),
        ).json(),
    )

    assert create_run_response.run_internal_id is not None
    return create_run_response.run_internal_id


@pytest.fixture
def run_without_files_id(
    client: TestClient,
    beamtime_id: BeamtimeId,
    run_channel_1_chemical_attributo_id: int,
    chemical_experiment_type_id: int,
    lyso_chemical_id: int,
) -> RunInternalId:
    external_run_id = 1000

    set_current_experiment_type(client, beamtime_id, chemical_experiment_type_id)

    create_run_response = JsonCreateOrUpdateRunOutput(
        **client.post(
            f"/api/runs/{external_run_id}",
            json=JsonCreateOrUpdateRun(
                # Important for our indexing job tests: we cannot start an offline indexing job without files (i.e. images).
                files=[],
                beamtime_id=beamtime_id,
                attributi=[
                    JsonAttributoValue(
                        attributo_id=run_channel_1_chemical_attributo_id,
                        attributo_value_chemical=lyso_chemical_id,
                    ),
                ],
                create_data_set=False,
                started=1,
                stopped=None,
            ).model_dump(),
        ).json(),
    )

    assert create_run_response.run_internal_id is not None
    return create_run_response.run_internal_id


@pytest.fixture
def simple_data_set_id(client: TestClient, simple_run_id: RunInternalId) -> int:
    create_data_set_response = JsonCreateDataSetFromRunOutput(
        **client.post(
            "/api/data-sets/from-run",
            json=JsonCreateDataSetFromRun(
                run_internal_id=simple_run_id,
            ).model_dump(),
        ).json(),
    )

    assert create_data_set_response.data_set_id > 0
    return create_data_set_response.data_set_id


@pytest.fixture
def simple_indexing_result_id(
    client: TestClient,
    simple_data_set_id: int,
    geometry_id: int,
) -> int:
    create_indexing_response = JsonCreateIndexingForDataSetOutput(
        **client.post(
            "/api/indexing",
            json=JsonCreateIndexingForDataSetInput(
                data_set_id=simple_data_set_id,
                is_online=False,
                cell_description="",
                geometry_id=geometry_id,
                command_line="",
                source="raw",
            ).model_dump(),
        ).json(),
    )

    client.post(
        f"/api/indexing/{create_indexing_response.indexing_result_id}/success",
        json=JsonIndexingResultFinishSuccessfully(
            workload_manager_job_id=1,
            stream_file="/tmp/some-file.stream",  # noqa: S108
            program_version="",
            # More or less random values, we don't care about the specifics here
            frames=200,
            # Hit rate 50%
            hits=100,
            # Indexing rate 20%
            indexed_frames=20,
            indexed_crystals=25,
            align_detector_groups=[
                JsonAlignDetectorGroup(
                    group="all",
                    x_translation_mm=0.5,
                    y_translation_mm=-0.5,
                    z_translation_mm=-0.8,
                    x_rotation_deg=1,
                    y_rotation_deg=2,
                ),
                JsonAlignDetectorGroup(
                    group="panel0",
                    x_translation_mm=0,
                    y_translation_mm=0,
                    z_translation_mm=0,
                    x_rotation_deg=0,
                    y_rotation_deg=0,
                ),
            ],
            generated_geometry_contents="",
            unit_cell_histograms_id=None,
            latest_log="",
        ).model_dump(),
    )

    return create_indexing_response.indexing_result_id


@pytest.fixture
def indexing_result_id_with_run_string_attributo_geometry(
    client: TestClient,
    simple_data_set_id: int,
    geometry_id_with_run_string_attributo: int,
) -> int:
    create_indexing_response = JsonCreateIndexingForDataSetOutput(
        **client.post(
            "/api/indexing",
            json=JsonCreateIndexingForDataSetInput(
                data_set_id=simple_data_set_id,
                is_online=False,
                cell_description="",
                geometry_id=geometry_id_with_run_string_attributo,
                command_line="",
                source="raw",
            ).model_dump(),
        ).json(),
    )

    client.post(
        f"/api/indexing/{create_indexing_response.indexing_result_id}/success",
        json=JsonIndexingResultFinishSuccessfully(
            workload_manager_job_id=1,
            stream_file="/tmp/some-file.stream",  # noqa: S108
            program_version="",
            # More or less random values, we don't care about the specifics here
            frames=200,
            # Hit rate 50%
            hits=100,
            # Indexing rate 20%
            indexed_frames=20,
            indexed_crystals=25,
            align_detector_groups=[
                JsonAlignDetectorGroup(
                    group="all",
                    x_translation_mm=0.5,
                    y_translation_mm=-0.5,
                    z_translation_mm=-0.8,
                    x_rotation_deg=1,
                    y_rotation_deg=2,
                ),
                JsonAlignDetectorGroup(
                    group="panel0",
                    x_translation_mm=0,
                    y_translation_mm=0,
                    z_translation_mm=0,
                    x_rotation_deg=0,
                    y_rotation_deg=0,
                ),
            ],
            generated_geometry_contents="",
            unit_cell_histograms_id=None,
            latest_log="",
        ).model_dump(),
    )

    return create_indexing_response.indexing_result_id


@pytest.fixture
def run_without_files_data_set_id(
    client: TestClient,
    run_without_files_id: RunInternalId,
) -> int:
    create_data_set_response = JsonCreateDataSetFromRunOutput(
        **client.post(
            "/api/data-sets/from-run",
            json=JsonCreateDataSetFromRun(
                run_internal_id=run_without_files_id,
            ).model_dump(),
        ).json(),
    )

    assert create_data_set_response.data_set_id > 0
    return create_data_set_response.data_set_id


def test_read_single_geometry(client: TestClient, geometry_id: int) -> None:
    result_raw = client.get(f"/api/geometries/{geometry_id}/raw").text

    assert result_raw == "hehe"

    result_json = JsonReadSingleGeometryOutput(
        **client.get(f"/api/geometries/{geometry_id}").json()
    )

    assert result_json.content == "hehe"


def test_update_single_geometry_without_usage(
    client: TestClient, geometry_id: int
) -> None:
    result = JsonGeometryWithoutContent(
        **client.patch(
            f"/api/geometries/{geometry_id}",
            json=JsonGeometryUpdate(content="newcontent", name="newname").model_dump(),
        ).json()
    )

    assert result.name == "newname"

    single_result = client.get(f"/api/geometries/{geometry_id}/raw").text

    assert single_result == "newcontent"


def test_update_single_geometry_use_different_attributo(
    client: TestClient,
    geometry_id_with_run_string_attributo: int,
    run_int_attributo_id: int,
) -> None:
    result = JsonGeometryWithoutContent(
        **client.patch(
            f"/api/geometries/{geometry_id_with_run_string_attributo}",
            json=JsonGeometryUpdate(
                # Before we had _RUN_STRING_ATTRIBUTO_NAME, and now we
                # update it to the int attributo
                content="clen {{" + _RUN_INT_ATTRIBUTO_NAME + "}}",
                name="newname",
            ).model_dump(),
        ).json()
    )
    assert result.attributi == [run_int_attributo_id]


def test_update_attributo_name_which_is_used_in_geometry(
    client: TestClient,
    geometry_id_with_run_string_attributo: int,
    run_string_attributo_id: int,
) -> None:
    updated_attributo = JsonAttributo(
        id=run_string_attributo_id,
        name=_RUN_STRING_ATTRIBUTO_NAME + "new",
        description="new description",
        group="new group",
        associated_table=AssociatedTable.RUN,
        # even change the type to string here
        attributo_type_string=JSONSchemaString(type="string", enum=None),
    )

    attributo_update_response = JsonUpdateAttributoOutput(
        **client.patch(
            "/api/attributi",
            json=JsonUpdateAttributoInput(
                attributo=updated_attributo,
                conversion_flags=JsonUpdateAttributoConversionFlags(ignore_units=True),
            ).model_dump(),
        ).json(),
    )

    assert attributo_update_response.id == run_string_attributo_id

    result_raw = client.get(
        f"/api/geometries/{geometry_id_with_run_string_attributo}/raw"
    ).text

    assert "{{" + _RUN_STRING_ATTRIBUTO_NAME + "}}" not in result_raw
    assert "{{" + _RUN_STRING_ATTRIBUTO_NAME + "new}}" in result_raw


def test_delete_attributo_which_is_used_in_geometry(
    client: TestClient,
    geometry_id_with_run_string_attributo: int,  # noqa: ARG001
    run_string_attributo_id: int,
) -> None:
    response = client.request(
        "DELETE",
        "/api/attributi",
        json=JsonDeleteAttributoInput(id=run_string_attributo_id).model_dump(),
    )
    assert response.status_code // 100 == 4


def test_update_single_geometry_with_usage(
    client: TestClient,
    geometry_id: int,
    # A little bit of domain knowledge needed here:
    # simple_indexing_result_id is using geometry_id for its geometry,
    # constituting a "usage" of the geometry
    simple_indexing_result_id: int,  # noqa: ARG001
) -> None:
    response = client.patch(
        f"/api/geometries/{geometry_id}",
        json=JsonGeometryUpdate(
            content=_GEOMETRY_ID_CONTENT + "newcontent", name="newname"
        ).model_dump(),
    )

    # Should fail, because we're changing the content
    assert response.status_code // 100 == 4

    response = client.patch(
        f"/api/geometries/{geometry_id}",
        json=JsonGeometryUpdate(
            content=_GEOMETRY_ID_CONTENT, name="newname"
        ).model_dump(),
    )

    # Is fine, we're only changing metadata
    assert response.status_code // 100 == 2


def test_delete_single_geometry(client: TestClient, geometry_id: int) -> None:
    result = JsonReadGeometriesForSingleBeamtime(
        **client.delete(f"/api/geometries/{geometry_id}").json()
    )

    assert not result.geometries


def test_delete_single_geometry_with_usage(
    client: TestClient,
    geometry_id: int,
    # A little bit of domain knowledge needed here:
    # simple_indexing_result_id is using geometry_id for its geometry,
    # constituting a "usage" of the geometry
    simple_indexing_result_id: int,  # noqa: ARG001
) -> None:
    result = JsonReadGeometriesForSingleBeamtime(
        **client.delete(f"/api/geometries/{geometry_id}").json()
    )

    # Should work also, there is no test for usage in the backend yet when deleting geometries
    assert not result.geometries


def test_read_single_beamtime_geometry(
    client: TestClient, beamtime_id: BeamtimeId, geometry_id: int
) -> None:
    result = JsonReadGeometriesForSingleBeamtime(
        **client.get(f"/api/geometry-for-beamtime/{beamtime_id}").json()
    )

    assert len(result.geometries) == 1
    assert result.geometries[0].id == geometry_id


def test_copy_geometry_to_other_beamtime(
    client: TestClient,
    beamtime_id: BeamtimeId,
    second_beamtime_id: BeamtimeId,
    geometry_id: int,
) -> None:
    # First, let's make sure the "for single beamtime" request doesn't mix up beamtimes
    result_first = JsonReadGeometriesForSingleBeamtime(
        **client.get(f"/api/geometry-for-beamtime/{beamtime_id}").json()
    )
    assert len(result_first.geometries) == 1

    result_second = JsonReadGeometriesForSingleBeamtime(
        **client.get(f"/api/geometry-for-beamtime/{second_beamtime_id}").json()
    )
    assert len(result_second.geometries) == 0

    result_all = JsonReadGeometriesForAllBeamtimes(
        **client.get("/api/all-geometries").json()
    )
    assert len(result_all.geometries) == 1

    # Now copy the beamtime from "first" to "second"
    client.post(
        "/api/geometry-copy-to-beamtime",
        json=JsonGeometryCopyToBeamtime(
            geometry_id=geometry_id,
            target_beamtime_id=second_beamtime_id,
        ).model_dump(),
    )
    result_second_after_copy = JsonReadGeometriesForSingleBeamtime(
        **client.get(f"/api/geometry-for-beamtime/{second_beamtime_id}").json()
    )
    assert len(result_second_after_copy.geometries) == 1
    assert result_second_after_copy.geometries[0] != geometry_id

    result_first_after_copy = JsonReadGeometriesForSingleBeamtime(
        **client.get(f"/api/geometry-for-beamtime/{beamtime_id}").json()
    )
    assert len(result_first_after_copy.geometries) == 1


def test_copy_geometry_to_other_beamtime_when_other_beamtime_has_one_with_the_same_name(
    client: TestClient,
    beamtime_id: BeamtimeId,  # noqa: ARG001
    second_beamtime_id: BeamtimeId,
    geometry_id: int,
) -> None:
    # create another geometry, in the other beamtime
    response = JsonGeometryWithoutContent(
        **client.post(
            "/api/geometries",
            json=JsonGeometryCreate(
                # Note: same name as in the first BT
                beamtime_id=second_beamtime_id,
                content="hehe",
                name="geometry name",
            ).model_dump(),
        ).json(),
    )
    assert response.id > 0

    # Now copy the beamtime from "first" to "second"
    copy_response = client.post(
        "/api/geometry-copy-to-beamtime",
        json=JsonGeometryCopyToBeamtime(
            geometry_id=geometry_id,
            target_beamtime_id=second_beamtime_id,
        ).model_dump(),
    )

    assert copy_response.status_code // 100 == 4


def test_read_all_geometries(client: TestClient, geometry_id: int) -> None:
    result = JsonReadGeometriesForAllBeamtimes(
        **client.get("/api/all-geometries").json()
    )

    assert len(result.geometries) == 1
    assert result.geometries[0].id == geometry_id


def test_read_single_beamtime(client: TestClient, beamtime_id: BeamtimeId) -> None:
    beamtime = JsonBeamtimeOutput(**client.get(f"/api/beamtimes/{beamtime_id}").json())

    assert beamtime == JsonBeamtimeOutput(
        id=beamtime_id,
        external_id="cool 1337",
        beamline="P12",
        proposal="BAG",
        title="Test beamtime",
        comment="comment",
        start=1,
        start_local=1,
        end=1000,
        end_local=1000,
        chemical_names=[],
        analysis_output_path="/",
    )


def test_read_single_chemical_names_beamtime(
    client: TestClient,
    beamtime_id: BeamtimeId,
    lyso_chemical_id: int,  # noqa: ARG001
) -> None:
    beamtimes = JsonReadBeamtime(**client.get("/api/beamtimes").json()).beamtimes

    assert beamtimes[0] == JsonBeamtimeOutput(
        id=beamtime_id,
        external_id="cool 1337",
        beamline="P12",
        proposal="BAG",
        title="Test beamtime",
        comment="comment",
        start=1,
        start_local=1,
        end=1000,
        end_local=1000,
        chemical_names=[TEST_CHEMICAL_NAME],
        analysis_output_path="/",
    )


def test_update_random_beamtime(client: TestClient, beamtime_id: BeamtimeId) -> None:
    response = JsonBeamtimeOutput(
        **client.patch(
            "/api/beamtimes",
            json=JsonUpdateBeamtimeInput(
                id=beamtime_id,
                external_id="cool 13372",
                beamline="P122",
                proposal="BAG2",
                title="Test beamtime2",
                comment="comment2",
                start_local=2,
                end_local=1002,
                analysis_output_path="/",
            ).model_dump(),
        ).json(),
    )
    assert response.id == beamtime_id
    beamtimes = JsonReadBeamtime(**client.get("/api/beamtimes").json()).beamtimes
    assert len(beamtimes) == 1

    assert beamtimes[0] == JsonBeamtimeOutput(
        id=beamtime_id,
        external_id="cool 13372",
        beamline="P122",
        proposal="BAG2",
        title="Test beamtime2",
        comment="comment2",
        start=2,
        start_local=2,
        end=1002,
        end_local=1002,
        chemical_names=[],
        analysis_output_path="/",
    )


def test_random_beamtime_creation_works(
    client: TestClient,
    beamtime_id: BeamtimeId,
) -> None:
    response = JsonReadBeamtime(**client.get("/api/beamtimes").json())
    assert len(response.beamtimes) == 1
    first_beamtime = response.beamtimes[0]
    assert first_beamtime == JsonBeamtimeOutput(
        id=beamtime_id,
        external_id="cool 1337",
        beamline="P12",
        proposal="BAG",
        title="Test beamtime",
        comment="comment",
        start=1,
        start_local=1,
        end=1000,
        end_local=1000,
        chemical_names=[],
        analysis_output_path="/",
    )


def test_chemical_string_attributo_creation_works_with_one_beamtime(
    client: TestClient,
    cell_description_attributo_id: int,
    beamtime_id: BeamtimeId,
) -> None:
    response = JsonReadAttributi(**client.get(f"/api/attributi/{beamtime_id}").json())
    assert len(response.attributi) == 1
    first_attributo = response.attributi[0]
    assert first_attributo == JsonAttributo(
        id=cell_description_attributo_id,
        associated_table=AssociatedTable.CHEMICAL,
        description="description",
        group="manual",
        name="cell description",
        attributo_type_string=JSONSchemaString(type="string", enum=None),
    )


def test_chemical_string_attributo_creation_works_in_presence_of_second_beamtime(
    client: TestClient,
    cell_description_attributo_id: int,
    beamtime_id: BeamtimeId,
    second_beamtime_id: BeamtimeId,
) -> None:
    # We have a second beam time simply to demonstrate that creating an attributo in one beam time and then requesting
    # attributi from a different beam time works.
    response = JsonReadAttributi(**client.get(f"/api/attributi/{beamtime_id}").json())
    assert len(response.attributi) == 1
    first_attributo = response.attributi[0]
    assert first_attributo == JsonAttributo(
        id=cell_description_attributo_id,
        associated_table=AssociatedTable.CHEMICAL,
        description="description",
        group="manual",
        name="cell description",
        attributo_type_string=JSONSchemaString(type="string", enum=None),
    )

    # Second beamtime shouldn't have attributi
    assert not JsonReadAttributi(
        **client.get(f"/api/attributi/{second_beamtime_id}").json(),
    ).attributi


def _mock_string_attributo_value(aid: int, s: str) -> JsonAttributoValue:
    return JsonAttributoValue(
        attributo_id=aid,
        attributo_value_str=s,
        attributo_value_bool=None,
        attributo_value_float=None,
        attributo_value_int=None,
        attributo_value_list_bool=None,
        attributo_value_list_float=None,
        attributo_value_list_str=None,
    )


def read_chemicals(client: TestClient, beamtime_id: BeamtimeId) -> JsonReadChemicals:
    return JsonReadChemicals(**client.get(f"/api/chemicals/{beamtime_id}").json())


def test_chemical_creation(
    client: TestClient,
    lyso_chemical_id: int,
    cell_description_attributo_id: int,
    point_group_attributo_id: int,
    space_group_attributo_id: int,
    beamtime_id: BeamtimeId,
    test_file: int,
) -> None:
    """
    Test if the creation of the 'lyso_chemical_id' actually worked, by checking the return value of the list chemicals function
    """
    response = read_chemicals(client, beamtime_id)
    assert len(response.chemicals) == 1
    first_chemical = response.chemicals[0]
    assert first_chemical.id == lyso_chemical_id
    assert first_chemical.attributi == [
        _mock_string_attributo_value(
            cell_description_attributo_id,
            LYSO_CELL_DESCRIPTION,
        ),
        _mock_string_attributo_value(point_group_attributo_id, LYSO_POINT_GROUP),
        _mock_string_attributo_value(space_group_attributo_id, LYSO_SPACE_GROUP),
    ]
    assert first_chemical.beamtime_id == beamtime_id
    assert first_chemical.chemical_type == ChemicalType.CRYSTAL
    assert first_chemical.name == "chemicalname"
    assert first_chemical.responsible_person == "Rosalind Franklin"
    assert len(first_chemical.files) == 1
    assert first_chemical.files[0].id == test_file


def test_chemical_creation_and_deletion(
    client: TestClient,
    beamtime_id: BeamtimeId,
    lyso_chemical_id: int,
    second_lyso_chemical_id: int,
) -> None:
    # can't use "client.delete" since that doesn't get JSON as an input
    client.request(
        "DELETE",
        "/api/chemicals",
        json=JsonDeleteChemicalInput(id=lyso_chemical_id).model_dump(),
    )

    response = read_chemicals(client, beamtime_id)

    assert len(response.chemicals) == 1
    assert response.chemicals[0].id == second_lyso_chemical_id


def test_chemical_creation_and_update(
    client: TestClient,
    lyso_chemical_id: int,
    cell_description_attributo_id: int,
    point_group_attributo_id: int,
    space_group_attributo_id: int,
    beamtime_id: BeamtimeId,
    second_test_file: int,
) -> None:
    """
    Take the 'lyso_chemical_id' and update it, then check if the list contains
    just one element and it's the up-to-date version.
    """
    # Here, we are assuming some knowledge of how the lyso chemical
    # was created in the fixture (see the attributo values and the
    # files).
    new_cell_description = "foo2"
    new_point_group = "bar2"
    new_space_group = "baz2"
    patched_chemical_json = JsonChemicalWithId(
        id=lyso_chemical_id,
        attributi=[
            _mock_string_attributo_value(
                cell_description_attributo_id,
                new_cell_description,
            ),
            _mock_string_attributo_value(point_group_attributo_id, new_point_group),
            _mock_string_attributo_value(space_group_attributo_id, new_space_group),
        ],
        beamtime_id=beamtime_id,
        chemical_type=ChemicalType.CRYSTAL,
        # The first chemical has "test_file" as files attached to it.
        # Here, we swap out the files, and see if that works.
        file_ids=[second_test_file],
        # new name
        name="chemicalname2",
        # new responsible person
        responsible_person="Rosalind Franklin2",
    )

    single_response = JsonCreateChemicalOutput(
        **client.patch(
            "/api/chemicals", json=patched_chemical_json.model_dump()
        ).json(),
    )
    assert single_response.id == lyso_chemical_id

    response = read_chemicals(client, beamtime_id)

    # still only one chemical => the update didn't accidentally create another chemical
    assert len(response.chemicals) == 1
    chemical = response.chemicals[0]
    assert chemical.name == "chemicalname2"
    # still just one file, but a different one!
    assert len(chemical.files) == 1
    assert chemical.files[0].id == second_test_file
    # point group, space group, cell description
    assert len(chemical.attributi) == 3
    assert set(a.attributo_value_str for a in chemical.attributi) == set(
        (new_cell_description, new_point_group, new_space_group),
    )


def test_chemical_creation_with_invalid_attributo_value(
    client: TestClient,
    cell_description_attributo_id: int,
    point_group_attributo_id: int,  # noqa: ARG001
    beamtime_id: BeamtimeId,
) -> None:
    """
    Create a chemical, and try to use an invalid attributo value. This will test if values are actually tested against the schema
    """
    response = client.post(
        "/api/chemicals",
        json=JsonChemicalWithoutId(
            name=TEST_CHEMICAL_NAME,
            responsible_person=TEST_CHEMICAL_RESPONSIBLE_PERSON,
            attributi=[
                JsonAttributoValue(
                    attributo_id=cell_description_attributo_id,
                    # cell description is definitely not an integer, this should lead to failure
                    attributo_value_int=3,
                ),
            ],
            chemical_type=ChemicalType.CRYSTAL,
            file_ids=[],
            beamtime_id=beamtime_id,
        ).model_dump(),
    )
    assert response.status_code // 100 == 4


def test_chemical_copy_from_other_beamtime_with_copy_attributi(
    client: TestClient,
    cell_description_attributo_id: int,  # noqa: ARG001
    point_group_attributo_id: int,  # noqa: ARG001
    space_group_attributo_id: int,  # noqa: ARG001
    lyso_chemical_id: int,
    beamtime_id: BeamtimeId,  # noqa: ARG001
    second_beamtime_id: BeamtimeId,
) -> None:
    input_ = JsonCopyChemicalInput(
        chemical_id=lyso_chemical_id,
        target_beamtime_id=second_beamtime_id,
        create_attributi=True,
    ).model_dump()
    response = JsonCopyChemicalOutput(
        **client.post(
            "/api/copy-chemical",
            json=input_,
        ).json(),
    )
    assert response.new_chemical_id > 0

    chemicals_in_second_beamtime = read_chemicals(client, second_beamtime_id)

    # Crude to just assume we have one chemical. We should also test for the actual contents.
    assert len(chemicals_in_second_beamtime.chemicals) == 1

    attributi_response = JsonReadAttributi(
        **client.get(f"/api/attributi/{second_beamtime_id}").json(),
    )
    # A bit crude to hard-code 2 here, but we are expecting point group, cell description and space group
    assert len(attributi_response.attributi) == 3


def test_create_and_delete_event_without_live_stream_and_files(
    client: TestClient,
    beamtime_id: BeamtimeId,
    second_beamtime_id: BeamtimeId,
) -> None:
    create_event_response = JsonEventTopLevelOutput(
        **client.post(
            "/api/events",
            json=JsonEventTopLevelInput(
                beamtime_id=beamtime_id,
                event=JsonEventInput(
                    # The "level" has to be "user" here because we're
                    # testing read_runs below, which only includes
                    # user messages.
                    source="mysource",
                    text="mytext",
                    level="user",
                    file_ids=[],
                ),
                with_live_stream=False,
            ).model_dump(),
        ).json(),
    )
    assert create_event_response.id > 0

    read_events_output = JsonReadEvents(
        **client.get(f"/api/events/{beamtime_id}").json(),
    )
    assert len(read_events_output.events) == 1
    e = read_events_output.events[0]
    assert e.id == create_event_response.id
    assert e.source == "mysource"
    assert e.text == "mytext"

    assert not JsonReadEvents(
        **client.get(f"/api/events/{second_beamtime_id}").json(),
    ).events

    # can't use "client.delete" since that doesn't get JSON as an input
    client.request(
        "DELETE",
        "/api/events",
        json=JsonDeleteEventInput(id=create_event_response.id).model_dump(),
    )

    read_events_output_after_deletion = JsonReadEvents(
        **client.get(f"/api/events/{beamtime_id}").json(),
    )
    assert not read_events_output_after_deletion.events


def test_upload_and_retrieve_file(client: TestClient) -> None:
    test_file = Path(__file__).parent / "test-file.txt"

    # Upload the file
    with test_file.open("rb") as upload_file:
        raw_output = client.post(
            "/api/files",
            data={"description": "test description", "deduplicate": str(False)},
            files={"file": upload_file},
        )
        print(raw_output)
        output = JsonCreateFileOutput(**raw_output.json())

        assert output.file_name == "test-file.txt"
        assert output.description == "test description"
        assert output.type_ == "text/plain"
        assert output.size_in_bytes == 17

    # Retrieve the contents (not the metadata)
    with test_file.open("rb") as upload_file:
        assert client.get(f"/api/files/{output.id}").content == upload_file.read()


def test_upload_and_retrieve_file_with_compression(client: TestClient) -> None:
    test_file = Path(__file__).parent / "big-test-file.txt"

    # Upload the file
    with test_file.open("rb") as upload_file:
        raw_output = client.post(
            "/api/files",
            data={
                "description": "test description",
                "deduplicate": str(False),
                # Explicitly enable compression (default is auto,
                # which uses the file size as a condition)
                "compress": "on",
            },
            files={"file": upload_file},
        )
        print(raw_output)
        output = JsonCreateFileOutput(**raw_output.json())

        assert output.type_ == "text/plain"
        assert output.size_in_bytes == 3198
        assert output.size_in_bytes_compressed == 53

    # Retrieve the contents (not the metadata)
    with test_file.open("rb") as upload_file:
        assert client.get(f"/api/files/{output.id}").content == upload_file.read()


def test_create_event_with_file(
    client: TestClient,
    beamtime_id: BeamtimeId,
    test_file: int,
) -> None:
    JsonEventTopLevelOutput(
        **client.post(
            "/api/events",
            json=JsonEventTopLevelInput(
                beamtime_id=beamtime_id,
                event=JsonEventInput(
                    source="mysource",
                    text="mytext",
                    level="user",
                    file_ids=[test_file],
                ),
                with_live_stream=False,
            ).model_dump(),
        ).json(),
    )

    read_events_output = JsonReadEvents(
        **client.get(f"/api/events/{beamtime_id}").json(),
    )
    assert len(read_events_output.events) == 1
    assert len(read_events_output.events[0].files) == 1
    assert read_events_output.events[0].files[0].id == test_file


def test_create_event_with_live_stream(
    client: TestClient,
    beamtime_id: BeamtimeId,
    test_file_path: Path,
) -> None:
    # Upload the file: for the live stream, it's important to use the
    # proper file name, as that's the criterion for a file to be a
    # live stream (yes, ugly, I know).
    with test_file_path.open("rb") as upload_file:
        raw_output = client.post(
            "/api/files",
            data={"description": "test description", "deduplicate": str(False)},
            # To pass a file name as well as the file contents, httpx uses a tuple.
            files={"file": (live_stream_image_name(beamtime_id), upload_file)},
        )
        file_id = JsonCreateFileOutput(**raw_output.json()).id

    JsonEventTopLevelOutput(
        **client.post(
            "/api/events",
            json=JsonEventTopLevelInput(
                beamtime_id=beamtime_id,
                event=JsonEventInput(
                    source="mysource",
                    level="user",
                    text="mytext",
                    file_ids=[],
                ),
                with_live_stream=True,
            ).model_dump(),
        ).json(),
    )

    # Read events -- assume we "magically" got a file attached to the event - a copy of our live stream image
    # (the actual live stream image doesn't work, because it gets updated in-place)
    read_events_output = JsonReadEvents(
        **client.get(f"/api/events/{beamtime_id}").json(),
    )
    assert len(read_events_output.events) == 1
    assert len(read_events_output.events[0].files) == 1
    assert read_events_output.events[0].files[0].id > file_id


def test_create_experiment_type(
    client: TestClient,
    beamtime_id: BeamtimeId,
    chemical_experiment_type_id: int,
) -> None:
    read_ets_output = JsonReadExperimentTypes(
        **client.get(f"/api/experiment-types/{beamtime_id}").json(),
    )
    assert len(read_ets_output.experiment_types) == 1
    ets = read_ets_output.experiment_types
    assert ets[0].id == chemical_experiment_type_id


def test_copy_experiment_types_from_other_beamtime_where_attributi_are_missing(
    client: TestClient,
    beamtime_id: BeamtimeId,
    chemical_experiment_type_id: int,  # noqa: ARG001
    second_beamtime_id: BeamtimeId,
) -> None:
    copy_ets_output = client.post(
        "/api/copy-experiment-types",
        json=JsonCopyExperimentTypesInput(
            from_beamtime=beamtime_id,
            to_beamtime=second_beamtime_id,
        ).model_dump(),
    )
    assert copy_ets_output.status_code == 400


def test_copy_experiment_types_from_other_beamtime_where_attributi_are_present(
    client: TestClient,
    beamtime_id: BeamtimeId,
    chemical_experiment_type_id: int,  # noqa: ARG001
    run_channel_1_chemical_attributo_id_in_second_beamtime: int,  # noqa: ARG001
    second_beamtime_id: BeamtimeId,
) -> None:
    copy_ets_output = JsonCopyExperimentTypesOutput(
        **client.post(
            "/api/copy-experiment-types",
            json=JsonCopyExperimentTypesInput(
                from_beamtime=beamtime_id,
                to_beamtime=second_beamtime_id,
            ).model_dump(),
        ).json(),
    )
    assert len(copy_ets_output.to_beamtime_experiment_type_ids) == 1

    read_ets_output = JsonReadExperimentTypes(
        **client.get(f"/api/experiment-types/{second_beamtime_id}").json(),
    )

    assert len(read_ets_output.experiment_types) == 1


def test_create_or_update_run_fails_without_experiment_type(
    client: TestClient,
    beamtime_id: BeamtimeId,
    # we want the experiment type to be created, but not used here
    chemical_experiment_type_id: int,  # noqa: ARG001
    run_string_attributo_id: int,
) -> None:
    # Let's make the external run ID deliberately high
    external_run_id = 1000

    response = client.post(
        f"/api/runs/{external_run_id}",
        json=JsonCreateOrUpdateRun(
            beamtime_id=beamtime_id,
            files=[],
            attributi=[
                JsonAttributoValue(
                    attributo_id=run_string_attributo_id,
                    attributo_value_str="foo",
                ),
            ],
            create_data_set=False,
            started=1,
            stopped=None,
        ).model_dump(),
    )

    # Doesn't work, because we haven't set the current experiment type yet
    assert response.status_code == 400


def set_current_experiment_type(
    client: TestClient,
    beamtime_id: BeamtimeId,
    id_: int,
) -> None:
    option_set_result = JsonUserConfigurationSingleOutput(
        **client.patch(
            f"/api/user-config/{beamtime_id}/current-experiment-type-id/{id_}",
        ).json(),
    )

    assert option_set_result.value_int == id_


def set_auto_pilot(client: TestClient, beamtime_id: BeamtimeId, enabled: bool) -> None:
    option_set_result = JsonUserConfigurationSingleOutput(
        **client.patch(f"/api/user-config/{beamtime_id}/auto-pilot/{enabled}").json(),
    )

    assert option_set_result.value_bool == enabled


def enable_crystfel_online(client: TestClient, beamtime_id: BeamtimeId) -> None:
    option_set_result = JsonUserConfigurationSingleOutput(
        **client.patch(f"/api/user-config/{beamtime_id}/online-crystfel/True").json(),
    )

    assert option_set_result.value_bool


def test_create_and_update_run_adding_some_files_later(
    client: TestClient,
    beamtime_id: BeamtimeId,
    chemical_experiment_type_id: int,
    run_string_attributo_id: int,  # noqa: ARG001
) -> None:
    # Set the experiment type (otherwise creating a run will fail - see above)
    set_current_experiment_type(client, beamtime_id, chemical_experiment_type_id)

    # Let's make the external run ID deliberately high
    external_run_id = 1000

    # Create the run and check the result
    response = JsonCreateOrUpdateRunOutput(
        **client.post(
            f"/api/runs/{external_run_id}",
            json=JsonCreateOrUpdateRun(
                beamtime_id=beamtime_id,
                # Initially no files
                files=[],
                attributi=[],
                create_data_set=False,
                started=1,
                stopped=None,
            ).model_dump(),
        ).json(),
    )

    assert response.run_created

    # Now same request, but with some files!
    response = JsonCreateOrUpdateRunOutput(
        **client.post(
            f"/api/runs/{external_run_id}",
            json=JsonCreateOrUpdateRun(
                beamtime_id=beamtime_id,
                # Initially no files
                files=[JsonRunFile(id=0, source="h5", glob="hehe")],
                attributi=[],
                # Keep started/stopped by setting to None
                create_data_set=False,
                started=None,
                stopped=None,
            ).model_dump(),
        ).json(),
    )

    assert not response.run_created

    read_runs_output = JsonReadRuns(**client.get(f"/api/runs/{beamtime_id}").json())

    assert len(read_runs_output.runs) == 1
    assert len(read_runs_output.runs[0].files) == 1


def test_create_and_update_run_after_setting_experiment_type_no_crystfel_online(
    client: TestClient,
    beamtime_id: BeamtimeId,
    chemical_experiment_type_id: int,
    run_string_attributo_id: int,
) -> None:
    # Set the experiment type (otherwise creating a run will fail - see above)
    set_current_experiment_type(client, beamtime_id, chemical_experiment_type_id)

    # Let's make the external run ID deliberately high
    external_run_id = 1000

    run_string_attributo_value = "foo"
    # Create the run and check the result
    response = JsonCreateOrUpdateRunOutput(
        **client.post(
            f"/api/runs/{external_run_id}",
            json=JsonCreateOrUpdateRun(
                beamtime_id=beamtime_id,
                files=[],
                attributi=[
                    JsonAttributoValue(
                        attributo_id=run_string_attributo_id,
                        attributo_value_str=run_string_attributo_value,
                    ),
                ],
                create_data_set=False,
                started=1,
                stopped=None,
            ).model_dump(),
        ).json(),
    )

    assert response.indexing_result_id is None
    assert response.run_created
    assert response.run_internal_id is not None and response.run_internal_id > 0
    assert not response.error_message

    read_runs_output = JsonReadRuns(**client.get(f"/api/runs/{beamtime_id}").json())

    assert len(read_runs_output.runs) == 1
    run = read_runs_output.runs[0]
    assert run.external_id == external_run_id
    assert run.stopped is None
    assert run.started == 1
    assert run.experiment_type_id == chemical_experiment_type_id
    assert len(run.attributi) == 1
    attributo = run.attributi[0]
    assert attributo.attributo_id == run_string_attributo_id
    assert attributo.attributo_value_str == run_string_attributo_value

    # Now, update the run
    response = JsonCreateOrUpdateRunOutput(
        **client.post(
            f"/api/runs/{external_run_id}",
            json=JsonCreateOrUpdateRun(
                beamtime_id=beamtime_id,
                files=[],
                attributi=[
                    JsonAttributoValue(
                        attributo_id=run_string_attributo_id,
                        # Update it with a simple "2" at the end!
                        attributo_value_str=run_string_attributo_value + "2",
                    ),
                ],
                create_data_set=False,
                started=1,
                # and signal a stop
                stopped=2,
            ).model_dump(),
        ).json(),
    )

    assert response.indexing_result_id is None
    assert not response.run_created
    assert not response.error_message
    assert (
        response.run_internal_id is not None
        and response.run_internal_id > 0
        and response.run_internal_id == read_runs_output.runs[0].id
    )

    read_runs_output = JsonReadRuns(**client.get(f"/api/runs/{beamtime_id}").json())

    assert len(read_runs_output.runs) == 1
    run = read_runs_output.runs[0]
    assert run.external_id == external_run_id
    assert run.stopped == 2
    assert run.started == 1
    assert run.experiment_type_id == chemical_experiment_type_id
    assert len(run.attributi) == 1
    attributo = run.attributi[0]
    assert attributo.attributo_id == run_string_attributo_id
    # repeat our updated "2" again
    assert attributo.attributo_value_str == run_string_attributo_value + "2"


def test_create_and_update_run_after_setting_experiment_type_crystfel_online(
    client: TestClient,
    beamtime_id: BeamtimeId,
    chemical_experiment_type_id: int,
    run_channel_1_chemical_attributo_id: int,
    lyso_chemical_id: int,
) -> None:
    # Set the experiment type (otherwise creating a run will fail - see above)
    set_current_experiment_type(client, beamtime_id, chemical_experiment_type_id)
    enable_crystfel_online(client, beamtime_id)

    # Let's make the external run ID deliberately high
    external_run_id = 1000

    # Create the run and check the result
    response = JsonCreateOrUpdateRunOutput(
        **client.post(
            f"/api/runs/{external_run_id}",
            json=JsonCreateOrUpdateRun(
                beamtime_id=beamtime_id,
                files=[],
                attributi=[
                    JsonAttributoValue(
                        attributo_id=run_channel_1_chemical_attributo_id,
                        attributo_value_chemical=lyso_chemical_id,
                    ),
                ],
                create_data_set=False,
                started=1,
                stopped=None,
            ).model_dump(),
        ).json(),
    )

    assert response.run_created
    assert response.run_internal_id is not None and response.run_internal_id > 0
    assert not response.error_message
    assert response.indexing_result_id is not None and response.indexing_result_id > 0

    # Next, test the "read indexing jobs" request with the status parameter
    read_indexing_results_response = JsonReadIndexingResultsOutput(
        **client.get(
            f"/api/indexing?status={DBJobStatus.RUNNING.value}&beamtimeId={beamtime_id}",
        ).json(),
    )
    # Wrong status (running), so no results
    assert not read_indexing_results_response.indexing_jobs

    read_indexing_results_response = JsonReadIndexingResultsOutput(
        **client.get(
            f"/api/indexing?status={DBJobStatus.QUEUED.value}&beamtimeId={beamtime_id}",
        ).json(),
    )
    assert len(read_indexing_results_response.indexing_jobs) == 1


def test_create_run_and_import_external_indexing_result(
    client: TestClient,
    beamtime_id: BeamtimeId,
    chemical_experiment_type_id: int,
) -> None:
    # Set the experiment type (otherwise creating a run will fail - see above)
    set_current_experiment_type(client, beamtime_id, chemical_experiment_type_id)

    # Let's make the external run ID deliberately high
    external_run_id = 1000

    # Create the run and check the result
    response = JsonCreateOrUpdateRunOutput(
        **client.post(
            f"/api/runs/{external_run_id}",
            json=JsonCreateOrUpdateRun(
                beamtime_id=beamtime_id,
                files=[],
                # Deliberately empty list of attributi - we want to
                # show that we can still create an indexing result
                # just fine, even without a chemical and so on
                attributi=[],
                create_data_set=False,
                started=1,
                stopped=None,
            ).model_dump(),
        ).json(),
    )

    assert response.run_created
    assert response.run_internal_id is not None and response.run_internal_id > 0
    assert response.indexing_result_id is None

    # Create the run and check the result
    indexing_creation_response = JsonImportFinishedIndexingJobOutput(
        **client.post(
            "/api/indexing/import",
            json=JsonImportFinishedIndexingJobInput(
                is_online=False,
                cell_description="",
                command_line="--multi",
                source="raw",
                run_internal_id=response.run_internal_id,
                stream_file="/tmp/stream.file",  # noqa: S108
                program_version="0.11.1",
                frames=10,
                hits=2,
                indexed_frames=3,
                align_detector_groups=[],
                geometry_contents="/tmp/geom",  # noqa: S108
                generated_geometry_file=None,
                job_log="test log",
            ).model_dump(),
        ).json()
    )

    # Next, test the "read indexing jobs" request with the status parameter
    read_indexing_results_response = JsonReadIndexingResultsOutput(
        **client.get(
            f"/api/indexing?status={DBJobStatus.DONE.value}&beamtimeId={beamtime_id}",
        ).json(),
    )
    assert len(read_indexing_results_response.indexing_jobs) == 1
    assert (
        read_indexing_results_response.indexing_jobs[0].id
        == indexing_creation_response.indexing_result_id
    )


def test_create_and_delete_run_after_setting_experiment_type_crystfel_online(
    client: TestClient,
    beamtime_id: BeamtimeId,
    chemical_experiment_type_id: int,
    run_channel_1_chemical_attributo_id: int,
    lyso_chemical_id: int,
) -> None:
    # Set the experiment type (otherwise creating a run will fail - see above)
    set_current_experiment_type(client, beamtime_id, chemical_experiment_type_id)
    enable_crystfel_online(client, beamtime_id)

    # Let's make the external run ID deliberately high
    external_run_id = 1000

    # Create the run and check the result
    client.post(
        f"/api/runs/{external_run_id}",
        json=JsonCreateOrUpdateRun(
            beamtime_id=beamtime_id,
            files=[],
            attributi=[
                JsonAttributoValue(
                    attributo_id=run_channel_1_chemical_attributo_id,
                    attributo_value_chemical=lyso_chemical_id,
                ),
            ],
            create_data_set=False,
            started=1,
            stopped=None,
        ).model_dump(),
    )
    client.post(
        f"/api/runs/{external_run_id+1}",
        json=JsonCreateOrUpdateRun(
            beamtime_id=beamtime_id,
            files=[],
            attributi=[
                JsonAttributoValue(
                    attributo_id=run_channel_1_chemical_attributo_id,
                    attributo_value_chemical=lyso_chemical_id,
                ),
            ],
            create_data_set=False,
            started=1,
            stopped=None,
        ).model_dump(),
    )

    # ...and remove the run again
    assert JsonDeleteRunOutput(
        **client.request("DELETE", f"/api/runs/{beamtime_id}/{external_run_id}").json(),
    ).result

    read_runs_output = JsonReadRuns(**client.get(f"/api/runs/{beamtime_id}").json())

    assert len(read_runs_output.runs) == 1
    assert read_runs_output.runs[0].external_id == external_run_id + 1


def test_create_and_update_run_with_patch(
    client: TestClient,
    beamtime_id: BeamtimeId,
    chemical_experiment_type_id: int,
    run_channel_1_chemical_attributo_id: int,
    run_string_attributo_id: int,
    lyso_chemical_id: int,
) -> None:
    set_current_experiment_type(client, beamtime_id, chemical_experiment_type_id)

    # Let's make the external run ID deliberately high
    external_run_id = 1000

    update_run_raw_output = client.post(
        f"/api/runs/{external_run_id}",
        json=JsonCreateOrUpdateRun(
            # two files initially, so we can modify and add and see what happens
            files=[
                JsonRunFile(id=0, source="h5", glob="hehe"),
                JsonRunFile(id=0, source="raw", glob="hoho"),
            ],
            beamtime_id=beamtime_id,
            attributi=[
                # we don't even mention the second run attributo here, since we're going to add it later and test if that works
                JsonAttributoValue(
                    attributo_id=run_channel_1_chemical_attributo_id,
                    attributo_value_chemical=lyso_chemical_id,
                ),
            ],
            create_data_set=False,
            started=1,
            stopped=None,
        ).model_dump(),
    ).json()

    # Create the run and check the result
    create_response = JsonCreateOrUpdateRunOutput(**update_run_raw_output)
    assert create_response.run_internal_id is not None
    assert len(create_response.files) == 2

    # this is arbitrary, it could also be files[1], but we have to assume something here
    assert create_response.files[0].glob == "hehe"

    new_files: list[JsonRunFile] = [
        # modify the first file and add a new one - remove the second one implicitly by not including it here
        JsonRunFile(
            id=create_response.files[0].id,
            glob=create_response.files[0].glob + "modified",
            source=create_response.files[0].source,
        ),
        JsonRunFile(id=0, source="raw2", glob="oh"),
    ]

    update_response = JsonUpdateRunOutput(
        **client.patch(
            "/api/runs",
            json=JsonUpdateRun(
                id=create_response.run_internal_id,
                experiment_type_id=chemical_experiment_type_id,
                attributi=[
                    JsonAttributoValue(
                        attributo_id=run_string_attributo_id,
                        attributo_value_str="some string",
                    ),
                ],
                files=new_files,
            ).model_dump(),
        ).json(),
    )

    assert update_response.result
    assert len(update_response.files) == 2
    assert update_response.files[0].glob == "hehemodified"
    assert update_response.files[1].glob == "oh"

    read_runs_output = JsonReadRuns(**client.get(f"/api/runs/{beamtime_id}").json())

    # Updating a run, you can specify attributi to update. However,
    # the attributes you specify are only the attributes that are
    # updated. All other attributes should stay the same.
    assert len(read_runs_output.runs[0].attributi) == 2
    cell_description_attributo = next(
        iter(
            a
            for a in read_runs_output.runs[0].attributi
            if a.attributo_id == run_string_attributo_id
        ),
        None,
    )
    assert cell_description_attributo is not None
    assert cell_description_attributo.attributo_value_str == "some string"
    # check if we still have two files
    assert len(read_runs_output.runs[0].files) == 2


def test_create_and_stop_run(
    client: TestClient,
    beamtime_id: BeamtimeId,
    chemical_experiment_type_id: int,
    run_channel_1_chemical_attributo_id: int,
    lyso_chemical_id: int,
) -> None:
    set_current_experiment_type(client, beamtime_id, chemical_experiment_type_id)

    # Let's make the external run ID deliberately high
    external_run_id = 1000

    # Create the run and check the result
    create_response = JsonCreateOrUpdateRunOutput(
        **client.post(
            f"/api/runs/{external_run_id}",
            json=JsonCreateOrUpdateRun(
                files=[],
                beamtime_id=beamtime_id,
                attributi=[
                    # we don't even mention the second run attributo here, since we're going to add it later and test if that works
                    JsonAttributoValue(
                        attributo_id=run_channel_1_chemical_attributo_id,
                        attributo_value_chemical=lyso_chemical_id,
                    ),
                ],
                create_data_set=False,
                started=1,
                stopped=None,
            ).model_dump(),
        ).json(),
    )

    assert create_response.run_created

    client.get(f"/api/runs/stop-latest/{beamtime_id}")

    read_runs_output = JsonReadRuns(**client.get(f"/api/runs/{beamtime_id}").json())

    assert read_runs_output.runs[0].stopped is not None


def test_update_indexing_job(
    client: TestClient,
    beamtime_id: BeamtimeId,
    chemical_experiment_type_id: int,
    run_channel_1_chemical_attributo_id: int,
    lyso_chemical_id: int,
) -> None:
    # Set the experiment type (otherwise creating a run will fail - see above)
    set_current_experiment_type(client, beamtime_id, chemical_experiment_type_id)
    # Enable CrystFEL online so an indexing job will be created
    enable_crystfel_online(client, beamtime_id)

    # Let's make the external run ID deliberately high
    external_run_id = 1000

    # Create the run and check the result
    create_run_response = JsonCreateOrUpdateRunOutput(
        **client.post(
            f"/api/runs/{external_run_id}",
            json=JsonCreateOrUpdateRun(
                files=[],
                beamtime_id=beamtime_id,
                attributi=[
                    JsonAttributoValue(
                        attributo_id=run_channel_1_chemical_attributo_id,
                        attributo_value_chemical=lyso_chemical_id,
                    ),
                ],
                create_data_set=False,
                started=1,
                stopped=None,
            ).model_dump(),
        ).json(),
    )

    assert create_run_response.run_internal_id is not None
    assert (
        create_run_response.indexing_result_id is not None
        and create_run_response.indexing_result_id > 0
    )

    update_indexing_job_response = JsonIndexingJobUpdateOutput(
        **client.post(
            f"/api/indexing/{create_run_response.indexing_result_id}/success",
            json=JsonIndexingResultFinishSuccessfully(
                workload_manager_job_id=1,
                stream_file="/tmp/some-file.stream",  # noqa: S108
                program_version="",
                # More or less random values, we don't care about the specifics here
                frames=200,
                # Hit rate 50%
                hits=100,
                # Indexing rate 20%
                indexed_frames=20,
                indexed_crystals=25,
                align_detector_groups=[
                    JsonAlignDetectorGroup(
                        group="all",
                        x_translation_mm=0.5,
                        y_translation_mm=-0.5,
                        z_translation_mm=-0.8,
                        x_rotation_deg=1,
                        y_rotation_deg=2,
                    ),
                    JsonAlignDetectorGroup(
                        group="panel0",
                        x_translation_mm=0,
                        y_translation_mm=0,
                        z_translation_mm=0,
                        x_rotation_deg=0,
                        y_rotation_deg=0,
                    ),
                ],
                generated_geometry_contents="",
                unit_cell_histograms_id=None,
                latest_log="",
            ).model_dump(),
        ).json(),
    )

    assert update_indexing_job_response.result

    # The result of this indexing you can query in various places. One of the most prominent ones is the "read runs" call.
    # For that to work, however, we need a data set.
    #
    # The test will be "unnecessarily" long now, but let's test the "create data set from run" feature just now
    create_data_set_response = JsonCreateDataSetFromRunOutput(
        **client.post(
            "/api/data-sets/from-run",
            json=JsonCreateDataSetFromRun(
                run_internal_id=create_run_response.run_internal_id,
            ).model_dump(),
        ).json(),
    )

    assert create_data_set_response.data_set_id > 0

    read_runs_output = JsonReadRunsOverview(
        **client.get(f"/api/runs-overview/{beamtime_id}").json(),
    )
    assert read_runs_output.foms_for_this_data_set is not None
    first_ds = read_runs_output.foms_for_this_data_set.data_set
    assert first_ds.id == create_data_set_response.data_set_id
    assert first_ds.attributi == [
        JsonAttributoValue(
            attributo_id=run_channel_1_chemical_attributo_id,
            attributo_value_chemical=lyso_chemical_id,
        ),
    ]
    assert first_ds.experiment_type_id == chemical_experiment_type_id
    summary = read_runs_output.foms_for_this_data_set.fom
    assert summary is not None
    assert summary.hit_rate == pytest.approx(50, 0.01)
    assert summary.indexing_rate == pytest.approx(20, 0.01)

    # Another place is the analysis view
    analysis_response = JsonReadNewAnalysisOutput(
        **client.post(
            "/api/analysis/analysis-results",
            json=JsonReadNewAnalysisInput(
                attributi_filter=[
                    JsonAttributoValue(
                        attributo_id=run_channel_1_chemical_attributo_id,
                        attributo_value_chemical=lyso_chemical_id,
                    ),
                ],
                beamtime_id=beamtime_id,
                merge_status=JsonMergeStatus.BOTH,
            ).model_dump(),
        ).json(),
    )

    assert len(analysis_response.filtered_data_sets) == 1
    assert (
        analysis_response.filtered_data_sets[0].experiment_type_id
        == chemical_experiment_type_id
    )
    assert (
        analysis_response.filtered_data_sets[0].id
        == create_data_set_response.data_set_id
    )

    # Another place is the analysis view
    single_data_set_result = JsonReadSingleDataSetResults(
        **client.get(
            f"/api/analysis/single-data-set/{beamtime_id}/{analysis_response.filtered_data_sets[0].id}",
        ).json(),
    )

    assert len(single_data_set_result.data_set.indexing_results) == 1


def test_indexing_result_with_two_equal_parameter(
    client: TestClient,
    beamtime_id: BeamtimeId,
    chemical_experiment_type_id: int,
    run_channel_1_chemical_attributo_id: int,  # noqa: ARG001
    lyso_chemical_id: int,  # noqa: ARG001
    simple_run_id: int,  # noqa: ARG001
    simple_data_set_id: int,
    geometry_id: int,
) -> None:
    # Set the experiment type (otherwise creating a run will fail - see above)
    set_current_experiment_type(client, beamtime_id, chemical_experiment_type_id)

    create_indexing_response = JsonCreateIndexingForDataSetOutput(
        **client.post(
            "/api/indexing",
            json=JsonCreateIndexingForDataSetInput(
                data_set_id=simple_data_set_id,
                is_online=False,
                cell_description="",
                geometry_id=geometry_id,
                command_line="",
                source="raw",
            ).model_dump(),
        ).json(),
    )

    ir_id = create_indexing_response.indexing_result_id
    ip_id = create_indexing_response.indexing_parameters_id

    finish_with_error_response = JsonIndexingJobUpdateOutput(
        **client.post(
            f"/api/indexing/{create_indexing_response.indexing_result_id}/finish-with-error",
            json=JsonIndexingResultFinishWithError(
                error_message="",
                latest_log="",
                workload_manager_job_id=1,
            ).model_dump(),
        ).json()
    )
    assert finish_with_error_response.result

    create_indexing_response_later = JsonCreateIndexingForDataSetOutput(
        **client.post(
            "/api/indexing",
            json=JsonCreateIndexingForDataSetInput(
                data_set_id=simple_data_set_id,
                is_online=False,
                cell_description="",
                geometry_id=geometry_id,
                command_line="",
                source="raw",
            ).model_dump(),
        ).json(),
    )

    new_ir_id = create_indexing_response_later.indexing_result_id
    new_ip_id = create_indexing_response_later.indexing_parameters_id

    # This is the important part: we get a new indexing parameters
    # object (due to laziness: we could just re-use an old one)
    assert new_ip_id != ip_id
    assert new_ir_id != ir_id

    # To start a merge job, we have to finish the indexing result first.
    client.post(
        f"/api/indexing/{new_ir_id}/success",
        json=JsonIndexingResultFinishSuccessfully(
            workload_manager_job_id=1,
            stream_file="/tmp/some-file.stream",  # noqa: S108
            program_version="",
            # More or less random values, we don't care about the specifics here
            frames=200,
            # Hit rate 50%
            hits=100,
            # Indexing rate 20%
            indexed_frames=20,
            indexed_crystals=25,
            align_detector_groups=[
                JsonAlignDetectorGroup(
                    group="all",
                    x_translation_mm=0.5,
                    y_translation_mm=-0.5,
                    z_translation_mm=-0.8,
                    x_rotation_deg=1,
                    y_rotation_deg=2,
                ),
            ],
            generated_geometry_contents="",
            unit_cell_histograms_id=None,
            latest_log="",
        ).model_dump(),
    )

    # Now queue a merge job. The bug we encountered before lead to
    # this merge job not showing up in the analysis results. So let's
    # see if this situation is fixed now.
    queue_merge_job_response = JsonQueueMergeJobOutput(
        **client.post(
            "/api/merging",
            json=JsonQueueMergeJobInput(
                strict_mode=False,
                data_set_id=simple_data_set_id,
                # New indexing parameters ID! Important. We should still see the old ID and jobs in the analysis result later
                indexing_parameters_id=new_ip_id,
                merge_parameters=JsonMergeParameters(
                    cell_description=LYSO_CELL_DESCRIPTION,
                    point_group=LYSO_POINT_GROUP,
                    space_group=LYSO_SPACE_GROUP,
                    merge_model=MergeModel.UNITY,
                    scale_intensities=ScaleIntensities.OFF,
                    post_refinement=False,
                    iterations=3,
                    polarisation=JsonPolarisation(angle=30, percent=50),
                    negative_handling=MergeNegativeHandling.IGNORE,
                    start_after=None,
                    stop_after=None,
                    rel_b=1.0,
                    no_pr=False,
                    force_bandwidth=None,
                    force_radius=None,
                    force_lambda=None,
                    no_delta_cc_half=False,
                    max_adu=None,
                    min_measurements=1,
                    logs=False,
                    min_res=None,
                    push_res=None,
                    w=None,
                    ambigator_command_line="",
                ),
            ).model_dump(),
        ).json(),
    )
    assert queue_merge_job_response.merge_result_id > 0

    # Finally, our analysis view
    single_data_set_result = JsonReadSingleDataSetResults(
        **client.get(
            f"/api/analysis/single-data-set/{beamtime_id}/{simple_data_set_id}",
        ).json(),
    )

    assert len(single_data_set_result.data_set.indexing_results) == 1
    ir_and_mr = single_data_set_result.data_set.indexing_results[0]
    assert ir_and_mr.parameters.id == ip_id
    assert len(ir_and_mr.indexing_results) == 2
    assert set(ir.id for ir in ir_and_mr.indexing_results) == set([ir_id, new_ir_id])
    assert len(ir_and_mr.merge_results) == 1


def test_analysis_view_with_single_data_set_directly_returns_results(
    client: TestClient,
    beamtime_id: BeamtimeId,
    chemical_experiment_type_id: int,
    run_channel_1_chemical_attributo_id: int,
    lyso_chemical_id: int,
) -> None:
    # Set the experiment type (otherwise creating a run will fail - see above)
    set_current_experiment_type(client, beamtime_id, chemical_experiment_type_id)

    create_data_set_response = JsonCreateDataSetOutput(
        **client.post(
            "/api/data-sets",
            json=JsonCreateDataSetInput(
                experiment_type_id=chemical_experiment_type_id,
                attributi=[
                    JsonAttributoValue(
                        attributo_id=run_channel_1_chemical_attributo_id,
                        # Good old Lyso!
                        attributo_value_chemical=lyso_chemical_id,
                    ),
                ],
            ).model_dump(),
        ).json(),
    )

    assert create_data_set_response.id > 0

    analysis_response = JsonReadNewAnalysisOutput(
        **client.post(
            "/api/analysis/analysis-results",
            json=JsonReadNewAnalysisInput(
                # Here we specify no filters, this is special for this test!
                attributi_filter=[],
                beamtime_id=beamtime_id,
                merge_status=JsonMergeStatus.BOTH,
            ).model_dump(),
        ).json(),
    )

    assert len(analysis_response.filtered_data_sets) == 1


def test_change_run_experiment_type(
    client: TestClient,
    beamtime_id: BeamtimeId,
    chemical_experiment_type_id: int,
    string_experiment_type_id: int,
    run_channel_1_chemical_attributo_id: int,
    lyso_chemical_id: int,
) -> None:
    set_current_experiment_type(client, beamtime_id, chemical_experiment_type_id)

    # Let's make the external run ID deliberately high
    external_run_id = 1000

    create_response = JsonCreateOrUpdateRunOutput(
        **client.post(
            f"/api/runs/{external_run_id}",
            json=JsonCreateOrUpdateRun(
                files=[],
                beamtime_id=beamtime_id,
                attributi=[
                    # we don't even mention the second run attributo
                    # here, since we're going to add it later and test
                    # if that works
                    JsonAttributoValue(
                        attributo_id=run_channel_1_chemical_attributo_id,
                        attributo_value_chemical=lyso_chemical_id,
                    ),
                ],
                create_data_set=False,
                started=1,
                stopped=None,
            ).model_dump(),
        ).json(),
    )
    assert create_response.run_created

    # To check if just one run has a different experiment type ID in
    # the end, create another run that isn't changed
    external_run_id_2 = external_run_id + 1
    second_create_response = JsonCreateOrUpdateRunOutput(
        **client.post(
            f"/api/runs/{external_run_id_2}",
            json=JsonCreateOrUpdateRun(
                files=[],
                beamtime_id=beamtime_id,
                attributi=[
                    JsonAttributoValue(
                        attributo_id=run_channel_1_chemical_attributo_id,
                        attributo_value_chemical=lyso_chemical_id,
                    ),
                ],
                create_data_set=False,
                started=2,
                stopped=None,
            ).model_dump(),
        ).json(),
    )
    assert second_create_response.run_created
    assert create_response.run_internal_id is not None

    output = JsonChangeRunExperimentTypeOutput(
        **client.post(
            "/api/experiment-types/change-for-run",
            json=JsonChangeRunExperimentType(
                run_internal_id=create_response.run_internal_id,
                experiment_type_id=string_experiment_type_id,
            ).model_dump(),
        ).json(),
    )

    assert output.result

    read_runs_output = JsonReadRuns(**client.get(f"/api/runs/{beamtime_id}").json())

    # Runs are sorted by time, descending, so the second element will be our run
    assert read_runs_output.runs[1].id == create_response.run_internal_id


def test_create_and_delete_data_set(
    client: TestClient,
    beamtime_id: BeamtimeId,
    chemical_experiment_type_id: int,
    lyso_chemical_id: int,
    run_channel_1_chemical_attributo_id: int,
) -> None:
    create_response = JsonCreateDataSetOutput(
        **client.post(
            "/api/data-sets",
            json=JsonCreateDataSetInput(
                experiment_type_id=chemical_experiment_type_id,
                attributi=[
                    JsonAttributoValue(
                        attributo_id=run_channel_1_chemical_attributo_id,
                        # Good old Lyso!
                        attributo_value_chemical=lyso_chemical_id,
                    ),
                ],
            ).model_dump(),
        ).json(),
    )

    assert create_response.id > 0

    # Now retrieve the data set list
    data_sets_response = JsonReadDataSets(
        **client.get(f"/api/data-sets/{beamtime_id}").json(),
    )
    assert len(data_sets_response.data_sets) == 1
    assert data_sets_response.data_sets[0].id == create_response.id
    assert (
        data_sets_response.data_sets[0].experiment_type_id
        == chemical_experiment_type_id
    )
    assert data_sets_response.data_sets[0].attributi == [
        JsonAttributoValue(
            attributo_id=run_channel_1_chemical_attributo_id,
            # Good old Lyso!
            attributo_value_chemical=lyso_chemical_id,
        ),
    ]

    # ...and remove the data set again
    assert JsonDeleteDataSetOutput(
        **client.request(
            "DELETE",
            "/api/data-sets",
            json=JsonDeleteDataSetInput(id=create_response.id).model_dump(),
        ).json(),
    ).result

    assert not JsonReadDataSets(
        **client.get(f"/api/data-sets/{beamtime_id}").json(),
    ).data_sets


def test_queue_merge_job_with_point_and_space_group_inferred(
    client: TestClient,
    beamtime_id: BeamtimeId,
    chemical_experiment_type_id: int,
    run_channel_1_chemical_attributo_id: int,
    lyso_chemical_id: int,
) -> None:
    # Set the experiment type (otherwise creating a run will fail - see above)
    set_current_experiment_type(client, beamtime_id, chemical_experiment_type_id)
    # Enable CrystFEL online so an indexing job will be created
    enable_crystfel_online(client, beamtime_id)

    # Let's make the external run ID deliberately high
    external_run_id = 1000

    # Create the run and check the result
    create_run_response = JsonCreateOrUpdateRunOutput(
        **client.post(
            f"/api/runs/{external_run_id}",
            json=JsonCreateOrUpdateRun(
                files=[],
                beamtime_id=beamtime_id,
                attributi=[
                    JsonAttributoValue(
                        attributo_id=run_channel_1_chemical_attributo_id,
                        attributo_value_chemical=lyso_chemical_id,
                    ),
                ],
                create_data_set=False,
                started=1,
                stopped=None,
            ).model_dump(),
        ).json(),
    )

    assert create_run_response.run_internal_id is not None
    assert (
        create_run_response.indexing_result_id is not None
        and create_run_response.indexing_result_id > 0
    )

    assert JsonIndexingJobUpdateOutput(
        **client.post(
            f"/api/indexing/{create_run_response.indexing_result_id}/success",
            json=JsonIndexingResultFinishSuccessfully(
                stream_file="/tmp/some-file.stream",  # noqa: S108
                program_version="",
                workload_manager_job_id=1,
                # More or less random values, we don't care about the specifics here
                frames=200,
                # Hit rate 50%
                hits=100,
                # Indexing rate 20%
                indexed_frames=20,
                indexed_crystals=25,
                align_detector_groups=[
                    JsonAlignDetectorGroup(
                        group="all",
                        x_translation_mm=0.5,
                        y_translation_mm=-0.5,
                        z_translation_mm=-0.8,
                        x_rotation_deg=1,
                        y_rotation_deg=2,
                    ),
                    JsonAlignDetectorGroup(
                        group="panel0",
                        x_translation_mm=0,
                        y_translation_mm=0,
                        z_translation_mm=0,
                        x_rotation_deg=0,
                        y_rotation_deg=0,
                    ),
                ],
                unit_cell_histograms_id=None,
                generated_geometry_contents="",
                latest_log="",
            ).model_dump(),
        ).json(),
    ).result

    # The result of this indexing you can query in various places. One
    # of the most prominent ones is the "read runs" call. For that to
    # work, however, we need a data set.
    #
    # The test will be "unnecessarily" long now, but let's test the
    # "create data set from run" feature just now
    create_data_set_response = JsonCreateDataSetFromRunOutput(
        **client.post(
            "/api/data-sets/from-run",
            json=JsonCreateDataSetFromRun(
                run_internal_id=create_run_response.run_internal_id,
            ).model_dump(),
        ).json(),
    )

    assert create_data_set_response.data_set_id > 0

    queue_merge_job_response = JsonQueueMergeJobOutput(
        **client.post(
            "/api/merging",
            json=JsonQueueMergeJobInput(
                strict_mode=False,
                data_set_id=create_data_set_response.data_set_id,
                # Literally random stuff here, doesn't matter.
                indexing_parameters_id=1,
                merge_parameters=JsonMergeParameters(
                    cell_description=LYSO_CELL_DESCRIPTION,
                    # Deliberately left blank
                    point_group="",
                    # ...or None so we can test if this is inferred from the chemical/indexing result
                    space_group=None,
                    merge_model=MergeModel.UNITY,
                    scale_intensities=ScaleIntensities.OFF,
                    post_refinement=False,
                    iterations=3,
                    polarisation=JsonPolarisation(angle=30, percent=50),
                    negative_handling=MergeNegativeHandling.IGNORE,
                    start_after=None,
                    stop_after=None,
                    rel_b=1.0,
                    no_pr=False,
                    force_bandwidth=None,
                    force_radius=None,
                    force_lambda=None,
                    no_delta_cc_half=False,
                    max_adu=None,
                    min_measurements=1,
                    logs=False,
                    min_res=None,
                    push_res=None,
                    w=None,
                    ambigator_command_line="",
                ),
            ).model_dump(),
        ).json(),
    )
    assert queue_merge_job_response.merge_result_id > 0

    queued_merge_results = JsonReadMergeResultsOutput(
        **client.get(f"/api/merging?status={DBJobStatus.QUEUED.value}").json(),
    )

    assert len(queued_merge_results.merge_jobs) == 1
    assert queued_merge_results.merge_jobs[0].parameters.point_group == LYSO_POINT_GROUP
    assert queued_merge_results.merge_jobs[0].parameters.space_group == LYSO_SPACE_GROUP


def test_queue_then_start_then_finish_merge_job(
    client: TestClient,
    beamtime_id: BeamtimeId,
    chemical_experiment_type_id: int,
    run_channel_1_chemical_attributo_id: int,
    lyso_chemical_id: int,
    test_file: int,
) -> None:
    # Set the experiment type (otherwise creating a run will fail - see above)
    set_current_experiment_type(client, beamtime_id, chemical_experiment_type_id)
    # Enable CrystFEL online so an indexing job will be created
    enable_crystfel_online(client, beamtime_id)

    # Let's make the external run ID deliberately high
    external_run_id = 1000

    # Create the run and check the result
    create_run_response = JsonCreateOrUpdateRunOutput(
        **client.post(
            f"/api/runs/{external_run_id}",
            json=JsonCreateOrUpdateRun(
                files=[],
                beamtime_id=beamtime_id,
                attributi=[
                    JsonAttributoValue(
                        attributo_id=run_channel_1_chemical_attributo_id,
                        attributo_value_chemical=lyso_chemical_id,
                    ),
                ],
                create_data_set=False,
                started=1,
                stopped=None,
            ).model_dump(),
        ).json(),
    )

    assert create_run_response.run_internal_id is not None
    assert (
        create_run_response.indexing_result_id is not None
        and create_run_response.indexing_result_id > 0
    )

    assert JsonIndexingJobUpdateOutput(
        **client.post(
            f"/api/indexing/{create_run_response.indexing_result_id}/success",
            json=JsonIndexingResultFinishSuccessfully(
                stream_file="/tmp/some-file.stream",  # noqa: S108
                program_version="",
                workload_manager_job_id=1,
                # More or less random values, we don't care about the specifics here
                frames=200,
                # Hit rate 50%
                hits=100,
                # Indexing rate 20%
                indexed_frames=20,
                indexed_crystals=25,
                align_detector_groups=[
                    JsonAlignDetectorGroup(
                        group="all",
                        x_translation_mm=0.5,
                        y_translation_mm=-0.5,
                        z_translation_mm=-0.8,
                        x_rotation_deg=1,
                        y_rotation_deg=2,
                    ),
                    JsonAlignDetectorGroup(
                        group="panel0",
                        x_translation_mm=0,
                        y_translation_mm=0,
                        z_translation_mm=0,
                        x_rotation_deg=0,
                        y_rotation_deg=0,
                    ),
                ],
                unit_cell_histograms_id=None,
                generated_geometry_contents="",
                latest_log="",
            ).model_dump(),
        ).json(),
    ).result

    # The result of this indexing you can query in various places. One
    # of the most prominent ones is the "read runs" call. For that to
    # work, however, we need a data set.
    #
    # The test will be "unnecessarily" long now, but let's test the
    # "create data set from run" feature just now
    create_data_set_response = JsonCreateDataSetFromRunOutput(
        **client.post(
            "/api/data-sets/from-run",
            json=JsonCreateDataSetFromRun(
                run_internal_id=create_run_response.run_internal_id,
            ).model_dump(),
        ).json(),
    )

    assert create_data_set_response.data_set_id > 0

    queue_merge_job_response = JsonQueueMergeJobOutput(
        **client.post(
            "/api/merging",
            json=JsonQueueMergeJobInput(
                strict_mode=False,
                data_set_id=create_data_set_response.data_set_id,
                # Literally random stuff here, doesn't matter.
                indexing_parameters_id=1,
                merge_parameters=JsonMergeParameters(
                    cell_description=LYSO_CELL_DESCRIPTION,
                    point_group=LYSO_POINT_GROUP,
                    space_group=LYSO_SPACE_GROUP,
                    merge_model=MergeModel.UNITY,
                    scale_intensities=ScaleIntensities.OFF,
                    post_refinement=False,
                    iterations=3,
                    polarisation=JsonPolarisation(angle=30, percent=50),
                    negative_handling=MergeNegativeHandling.IGNORE,
                    start_after=None,
                    stop_after=None,
                    rel_b=1.0,
                    no_pr=False,
                    force_bandwidth=None,
                    force_radius=None,
                    force_lambda=None,
                    no_delta_cc_half=False,
                    max_adu=None,
                    min_measurements=1,
                    logs=False,
                    min_res=None,
                    push_res=None,
                    w=None,
                    ambigator_command_line="",
                ),
            ).model_dump(),
        ).json(),
    )
    assert queue_merge_job_response.merge_result_id > 0

    # Get the queued merge jobs, should be one:
    queued_merge_results = JsonReadMergeResultsOutput(
        **client.get(f"/api/merging?status={DBJobStatus.QUEUED.value}").json(),
    )
    assert len(queued_merge_results.merge_jobs) == 1

    # Now simulate that some daemon actually started the merge job
    start_merge_job_response = JsonMergeJobStartedOutput(
        **client.post(
            f"/api/merging/{queue_merge_job_response.merge_result_id}/start",
            # job ID is the SLURM (or similar) job ID, so it's random
            json=JsonMergeJobStartedInput(job_id=1337, time=10).model_dump(),
        ).json(),
    )
    assert start_merge_job_response.time > 0

    # Get the running merge jobs, should be one:
    running_merge_results = JsonReadMergeResultsOutput(
        **client.get(f"/api/merging?status={DBJobStatus.RUNNING.value}").json(),
    )
    assert len(running_merge_results.merge_jobs) == 1

    merge_result = JsonMergeResultInternal(
        mtz_file_id=test_file,
        ambigator_fg_graph_file_id=None,
        fom=JsonMergeResultFom(
            # Again, more or less completely random stuff here
            snr=1.0,
            wilson=None,
            ln_k=None,
            discarded_reflections=1,
            one_over_d_from=1.1,
            one_over_d_to=1.2,
            redundancy=1.3,
            completeness=1.4,
            measurements_total=10,
            reflections_total=11,
            reflections_possible=12,
            r_split=1.5,
            r1i=1.6,
            r2=1.7,
            cc=1.8,
            ccstar=1.9,
            ccano=2.0,
            crdano=2.1,
            rano=2.2,
            rano_over_r_split=2.3,
            d1sig=2.4,
            d2sig=2.5,
            outer_shell=JsonMergeResultOuterShell(
                resolution=2.6,
                ccstar=2.7,
                r_split=2.8,
                cc=2.9,
                unique_reflections=20,
                completeness=3.0,
                redundancy=3.1,
                snr=3.2,
                min_res=3.3,
                max_res=3.4,
            ),
        ),
        detailed_foms=[
            JsonMergeResultShell(
                one_over_d_centre=4.0,
                nref=3,
                d_over_a=4.1,
                min_res=4.2,
                max_res=4.3,
                cc=4.4,
                ccstar=4.5,
                r_split=4.6,
                reflections_possible=4,
                completeness=4.7,
                measurements=5,
                redundancy=4.8,
                snr=4.9,
                mean_i=5.0,
            ),
        ],
        refinement_results=[
            JsonRefinementResultInternal(
                # The ID is ignored when ingesting the result into the
                # DB, but we are using a comparison operator down in
                # this test, where we have to specify an ID
                id=1,
                pdb_file_id=test_file,
                mtz_file_id=test_file,
                r_free=0.5,
                r_work=0.6,
                rms_bond_angle=180.0,
                rms_bond_length=2.0,
            ),
        ],
    )

    # Now finish merge job
    finish_merge_job_response = JsonMergeJobFinishOutput(
        **client.post(
            f"/api/merging/{queue_merge_job_response.merge_result_id}/finish",
            json=JsonMergeJobFinishedInput(
                error=None, result=merge_result, latest_log=None
            ).model_dump(),
        ).json(),
    )

    assert finish_merge_job_response.result

    # Get the analysis view and find our result!
    analysis_response = JsonReadNewAnalysisOutput(
        **client.post(
            "/api/analysis/analysis-results",
            json=JsonReadNewAnalysisInput(
                attributi_filter=[
                    JsonAttributoValue(
                        attributo_id=run_channel_1_chemical_attributo_id,
                        attributo_value_chemical=lyso_chemical_id,
                    ),
                ],
                beamtime_id=beamtime_id,
                merge_status=JsonMergeStatus.BOTH,
            ).model_dump(),
        ).json(),
    )

    assert len(analysis_response.filtered_data_sets) == 1

    # Another place is the analysis view
    single_data_set_result = JsonReadSingleDataSetResults(
        **client.get(
            f"/api/analysis/single-data-set/{beamtime_id}/{analysis_response.filtered_data_sets[0].id}",
        ).json(),
    )

    first_ds = single_data_set_result.data_set
    assert len(first_ds.indexing_results) == 1
    first_ir = first_ds.indexing_results[0]
    assert len(first_ir.merge_results) == 1
    first_mr = first_ir.merge_results[0]
    assert first_mr.id == queue_merge_job_response.merge_result_id
    assert first_mr.runs == [str(external_run_id)]
    assert first_mr.parameters == JsonMergeParameters(
        point_group=LYSO_POINT_GROUP,
        space_group=LYSO_SPACE_GROUP,
        cell_description=LYSO_CELL_DESCRIPTION,
        negative_handling=MergeNegativeHandling.IGNORE,
        merge_model=MergeModel.UNITY,
        scale_intensities=ScaleIntensities.OFF,
        post_refinement=False,
        iterations=3,
        polarisation=JsonPolarisation(angle=30, percent=50),
        start_after=None,
        stop_after=None,
        rel_b=1.0,
        no_pr=False,
        force_bandwidth=None,
        force_radius=None,
        force_lambda=None,
        no_delta_cc_half=False,
        max_adu=None,
        min_measurements=1,
        logs=False,
        min_res=None,
        push_res=None,
        w=None,
        ambigator_command_line="",
    )
    assert len(first_mr.refinement_results) == 1
    refinement_result_id = first_mr.refinement_results[0].id
    assert first_mr.refinement_results[0] == JsonRefinementResult(
        id=refinement_result_id,
        merge_result_id=queue_merge_job_response.merge_result_id,
        pdb_file_id=test_file,
        mtz_file_id=test_file,
        r_free=0.5,
        r_work=0.6,
        rms_bond_angle=180.0,
        rms_bond_length=2.0,
    )
    assert first_mr.state_queued is None
    assert first_mr.state_error is None
    assert first_mr.state_running is None
    assert first_mr.state_done is not None
    assert first_mr.state_done.result == merge_result


def test_start_run(
    client: TestClient,
    beamtime_id: BeamtimeId,
    chemical_experiment_type_id: int,
) -> None:
    # Set the experiment type (otherwise creating a run will fail - see above)
    set_current_experiment_type(client, beamtime_id, chemical_experiment_type_id)

    # Let's make the external run ID deliberately high
    external_run_id = 1000

    # Create the run and check the result
    start_run_response = JsonStartRunOutput(
        **client.get(f"/api/runs/{external_run_id}/start/{beamtime_id}").json(),
    )

    assert start_run_response.run_internal_id is not None
    read_runs_output = JsonReadRuns(**client.get(f"/api/runs/{beamtime_id}").json())

    assert len(read_runs_output.runs) == 1
    assert read_runs_output.runs[0].id == start_run_response.run_internal_id


def test_start_two_runs_and_enable_auto_pilot(
    client: TestClient,
    beamtime_id: BeamtimeId,
    chemical_experiment_type_id: int,
    # this attributo is marked as manual
    run_string_attributo_id: int,
) -> None:
    # Set the experiment type (otherwise creating a run will fail - see above)
    set_current_experiment_type(client, beamtime_id, chemical_experiment_type_id)
    set_auto_pilot(client, beamtime_id, enabled=True)

    # Let's make the external run ID deliberately high
    external_run_id = 1000

    # Start a normal run, and finish it.
    first_run_id = JsonStartRunOutput(
        **client.get(f"/api/runs/{external_run_id}/start/{beamtime_id}").json(),
    ).run_internal_id
    # This value we set to the first run, and since it's manual and
    # the autopilot is on, we expect it to be set in the next run as
    # well
    string_value = "testvalue"
    # Set the manual attribute:
    update_run_result = client.patch(
        "/api/runs",
        json=JsonUpdateRun(
            id=first_run_id,
            experiment_type_id=chemical_experiment_type_id,
            attributi=[
                JsonAttributoValue(
                    attributo_id=run_string_attributo_id,
                    attributo_value_str=string_value,
                ),
            ],
            files=[],
        ).model_dump(),
    )
    assert update_run_result.status_code // 100 == 2
    stop_result = client.get(f"/api/runs/stop-latest/{beamtime_id}")
    assert stop_result.status_code // 100 == 2
    assert JsonStopRunOutput(**stop_result.json()).result

    # Now start a second run, providing no attributo values
    second_external_run_id = 1001
    second_run_id = JsonStartRunOutput(
        **client.get(f"/api/runs/{second_external_run_id}/start/{beamtime_id}").json(),
    ).run_internal_id

    read_runs_output = JsonReadRuns(**client.get(f"/api/runs/{beamtime_id}").json())

    assert len(read_runs_output.runs) == 2
    assert read_runs_output.runs[0].id == second_run_id
    assert len(read_runs_output.runs[0].attributi) == 1
    assert read_runs_output.runs[0].attributi[0].attributo_value_str == string_value
    assert read_runs_output.runs[1].stopped is not None

    # Another test: create a third run. This was added later to simulate a bug: in the
    # web server, we cannot use "one or none" to get the latest run: it will return more than one run!
    third_external_run_id = 1002
    third_run_id = JsonStartRunOutput(
        **client.get(f"/api/runs/{third_external_run_id}/start/{beamtime_id}").json(),
    ).run_internal_id

    assert third_run_id > second_run_id

    fourth_external_run_id = 1003
    # Now, we add a run, and provide the manual attributo directly. In
    # this case, auto pilot shouldn't do anything and use the new attributo value
    fourth_string_value = string_value + "new"
    fourth_run_response = JsonCreateOrUpdateRunOutput(
        **client.post(
            f"/api/runs/{fourth_external_run_id}",
            json=JsonCreateOrUpdateRun(
                files=[],
                beamtime_id=beamtime_id,
                attributi=[
                    JsonAttributoValue(
                        attributo_id=run_string_attributo_id,
                        attributo_value_str=fourth_string_value,
                    ),
                ],
                create_data_set=False,
            ).model_dump(),
        ).json(),
    )
    assert fourth_run_response.run_created

    read_runs_output_after_last_run = JsonReadRuns(
        **client.get(f"/api/runs/{beamtime_id}").json(),
    )
    assert len(read_runs_output_after_last_run.runs) == 4

    fourth_attributi = read_runs_output_after_last_run.runs[0].attributi
    fourth_manual_attributo = [
        x for x in fourth_attributi if x.attributo_id == run_string_attributo_id
    ]
    assert fourth_manual_attributo
    assert fourth_manual_attributo[0].attributo_value_str == string_value + "new"


def test_start_two_runs_and_enable_auto_pilot_using_create_or_update_run(
    client: TestClient,
    beamtime_id: BeamtimeId,
    chemical_experiment_type_id: int,
    run_int_automatic_attributo_id: int,
    # this attributo is marked as manual
    run_string_attributo_id: int,
) -> None:
    # Set the experiment type (otherwise creating a run will fail - see above)
    set_current_experiment_type(client, beamtime_id, chemical_experiment_type_id)
    set_auto_pilot(client, beamtime_id, enabled=True)

    # Let's make the external run ID deliberately high
    external_run_id = 1000

    # Start a normal run, and finish it.
    first_run_id = JsonStartRunOutput(
        **client.get(f"/api/runs/{external_run_id}/start/{beamtime_id}").json(),
    ).run_internal_id
    # This value we set to the first run, and since it's manual and
    # the autopilot is on, we expect it to be set in the next run as
    # well
    string_value = "testvalue"
    # Set the manual attribute:
    update_run_result = client.patch(
        "/api/runs",
        json=JsonUpdateRun(
            id=first_run_id,
            experiment_type_id=chemical_experiment_type_id,
            attributi=[
                JsonAttributoValue(
                    attributo_id=run_string_attributo_id,
                    attributo_value_str=string_value,
                ),
                JsonAttributoValue(
                    attributo_id=run_int_automatic_attributo_id,
                    attributo_value_int=1337,
                ),
            ],
            files=[],
        ).model_dump(),
    )
    assert update_run_result.status_code // 100 == 2
    stop_result = client.get(f"/api/runs/stop-latest/{beamtime_id}")
    assert stop_result.status_code // 100 == 2
    assert JsonStopRunOutput(**stop_result.json()).result

    # Now start a second run, providing no attributo values
    second_external_run_id = 2000
    second_run_response = JsonCreateOrUpdateRunOutput(
        **client.post(
            f"/api/runs/{second_external_run_id}",
            json=JsonCreateOrUpdateRun(
                files=[],
                beamtime_id=beamtime_id,
                # neither automatic nor manual attributi are present here, deliberately. We expect the automatic one to
                # not appear, but the manual one should.
                attributi=[],
                create_data_set=False,
            ).model_dump(),
        ).json(),
    )
    assert second_run_response.run_created
    assert second_run_response.indexing_result_id is None
    assert second_run_response.error_message is None

    read_runs_output = JsonReadRuns(**client.get(f"/api/runs/{beamtime_id}").json())

    assert len(read_runs_output.runs) == 2
    assert read_runs_output.runs[0].id == second_run_response.run_internal_id
    assert len(read_runs_output.runs[0].attributi) == 1
    assert read_runs_output.runs[0].attributi[0].attributo_id == run_string_attributo_id
    assert read_runs_output.runs[0].attributi[0].attributo_value_str == string_value


def test_start_run_no_experiment_type_set(
    client: TestClient,
    beamtime_id: BeamtimeId,
) -> None:
    # Let's make the external run ID deliberately high
    external_run_id = 1000

    # Create the run and check the result
    start_run_response = client.get(f"/api/runs/{external_run_id}/start/{beamtime_id}")

    assert start_run_response.status_code == 400


def test_read_and_update_runs_bulk(
    client: TestClient,
    beamtime_id: BeamtimeId,
    chemical_experiment_type_id: int,
    string_experiment_type_id: int,
    run_string_attributo_id: int,
    run_int_attributo_id: int,
    run_channel_1_chemical_attributo_id: int,
) -> None:
    # Set the experiment type (otherwise creating a run will fail - see above)
    set_current_experiment_type(client, beamtime_id, chemical_experiment_type_id)

    run_definitions = [
        {"id": 1000, "string-attributo": "foo", "int-attributo": 3},
        {"id": 1001, "string-attributo": "bar", "int-attributo": 4},
        {"id": 1002, "string-attributo": None, "int-attributo": 4},
        {"id": 1003, "string-attributo": "bar", "int-attributo": None},
    ]

    run_internal_ids: list[int] = []

    for run_definition in run_definitions:
        response = JsonCreateOrUpdateRunOutput(
            **client.post(
                f"/api/runs/{run_definition['id']}",
                json=JsonCreateOrUpdateRun(
                    beamtime_id=beamtime_id,
                    attributi=(
                        [
                            JsonAttributoValue(
                                attributo_id=run_string_attributo_id,
                                attributo_value_str=run_definition["string-attributo"],  # type: ignore
                            ),
                        ]
                        if run_definition["string-attributo"] is not None
                        else []
                    )
                    + (
                        [
                            JsonAttributoValue(
                                attributo_id=run_int_attributo_id,
                                attributo_value_int=run_definition["int-attributo"],  # type: ignore
                            ),
                        ]
                        if run_definition["int-attributo"] is not None
                        else []
                    ),
                    create_data_set=False,
                    started=1,
                    stopped=None,
                    files=[],
                ).model_dump(),
            ).json(),
        )
        assert response.run_created
        assert response.run_internal_id is not None
        run_internal_ids.append(response.run_internal_id)

    read_output = JsonReadRunsBulkOutput(
        **client.post(
            "/api/runs-bulk",
            json=JsonReadRunsBulkInput(
                beamtime_id=beamtime_id,
                external_run_ids=[x["id"] for x in run_definitions],  # type: ignore
            ).model_dump(),
        ).json(),
    )

    assert read_output.experiment_type_ids == [chemical_experiment_type_id]
    assert read_output.experiment_types

    # Annoying redirect to Python's "unittest" framework, but it's a very elegant solution to the problem
    # of comparing two lists of elements that are not hashable (as BaseModel, right now, is)
    tc = TestCase()
    tc.maxDiff = None
    bulk_values_chemical: list[JsonAttributoValue] = [
        y
        for x in read_output.attributi_values
        if x.attributo_id == run_channel_1_chemical_attributo_id
        for y in x.values
    ]
    assert not bulk_values_chemical
    bulk_values_string = [
        y
        for x in read_output.attributi_values
        if x.attributo_id == run_string_attributo_id
        for y in x.values
    ]
    tc.assertCountEqual(
        bulk_values_string,
        [
            JsonAttributoValue(
                attributo_id=run_string_attributo_id,
                attributo_value_str="foo",
            ),
            JsonAttributoValue(
                attributo_id=run_string_attributo_id,
                attributo_value_str="bar",
            ),
        ],
    )
    bulk_values_int = [
        y
        for x in read_output.attributi_values
        if x.attributo_id == run_int_attributo_id
        for y in x.values
    ]
    tc.assertCountEqual(
        bulk_values_int,
        [
            JsonAttributoValue(
                attributo_id=run_int_attributo_id,
                attributo_value_int=3,
            ),
            JsonAttributoValue(
                attributo_id=run_int_attributo_id,
                attributo_value_int=4,
            ),
        ],
    )

    # Let's first just update one of the attributi: the string one, and set it to something completely different, so
    # there's a change in every run. Also, let's change the experiment type
    write_output = JsonUpdateRunsBulkOutput(
        **client.patch(
            "/api/runs-bulk",
            json=JsonUpdateRunsBulkInput(
                beamtime_id=beamtime_id,
                external_run_ids=[x["id"] for x in run_definitions],  # type: ignore
                attributi=[
                    JsonAttributoValue(
                        attributo_id=run_string_attributo_id,
                        attributo_value_str="qux",
                    ),
                ],
                new_experiment_type_id=string_experiment_type_id,
            ).model_dump(),
        ).json(),
    )

    assert write_output.result

    read_runs_output = JsonReadRuns(**client.get(f"/api/runs/{beamtime_id}").json())

    for run in read_runs_output.runs:
        assert run.experiment_type_id == string_experiment_type_id

        # First test: did our string update succeed?
        string_attributo = next(
            iter(
                [a for a in run.attributi if a.attributo_id == run_string_attributo_id],
            ),
            None,
        )
        assert string_attributo is not None
        assert string_attributo.attributo_value_str == "qux"

        # Second test: is the run "the same"? Check that the
        # previously unset attribute for the chemical ID is still not
        # there.
        channel_1_attributo = next(
            iter(
                [
                    a
                    for a in run.attributi
                    if a.attributo_id == run_channel_1_chemical_attributo_id
                ],
            ),
            None,
        )
        assert channel_1_attributo is None


def test_create_file(client: TestClient, test_file_path: Path) -> None:
    file_description = "test description"
    # Upload the file
    with test_file_path.open("rb") as upload_file:
        raw_output = client.post(
            "/api/files",
            files={"file": upload_file},
            data={"description": file_description, "deduplicate": str(False)},
        )
        print(raw_output.json())
        output = JsonCreateFileOutput(**raw_output.json())

        assert output.file_name.endswith(".txt")
        assert output.description == file_description
        assert output.type_ == "text/plain"
        assert output.size_in_bytes == 17
        assert output.id > 0

    # # Retrieve the contents (not the metadata)
    with test_file_path.open("rb") as upload_file:
        assert client.get(f"/api/files/{output.id}").content == upload_file.read()


def test_create_file_with_deduplication(
    client: TestClient,
    test_file_path: Path,
) -> None:
    # Upload the file once
    with test_file_path.open("rb") as upload_file:
        first_upload = JsonCreateFileOutput(
            **client.post(
                "/api/files",
                files={"file": upload_file},
                data={"description": "test", "deduplicate": str(False)},
            ).json(),
        )

    # Upload again, check if the same file is returned due to deduplication
    with test_file_path.open("rb") as upload_file:
        second_upload = JsonCreateFileOutput(
            **client.post(
                "/api/files",
                files={"file": upload_file},
                data={"description": "test", "deduplicate": str(True)},
            ).json(),
        )

    assert first_upload.id == second_upload.id


def test_create_file_simple(client: TestClient, test_file_path: Path) -> None:
    # Upload the file
    with test_file_path.open("rb") as upload_file:
        raw_output = client.post(
            "/api/files/simple/txt",
            # Important: don't use "files=" here since we're sending the data without multipart shenenigans
            content=upload_file.read(),
        )
        print(raw_output.json())
        output = JsonCreateFileOutput(**raw_output.json())

        assert output.file_name.endswith(".txt")
        assert output.description == ""
        assert output.type_ == "text/plain"
        assert output.size_in_bytes == 17
        assert output.id > 0

    # Retrieve the contents (not the metadata)
    with test_file_path.open("rb") as upload_file:
        assert client.get(f"/api/files/{output.id}").content == upload_file.read()


def test_read_and_update_user_config(
    client: TestClient,
    beamtime_id: BeamtimeId,
) -> None:
    result_ap = JsonUserConfigurationSingleOutput(
        **client.get(f"/api/user-config/{beamtime_id}/auto-pilot").json(),
    )
    assert result_ap.value_bool is not None
    assert result_ap.value_int is None

    result_ap = JsonUserConfigurationSingleOutput(
        **client.patch(f"/api/user-config/{beamtime_id}/auto-pilot/True").json(),
    )
    assert result_ap.value_bool

    result_ap = JsonUserConfigurationSingleOutput(
        **client.patch(f"/api/user-config/{beamtime_id}/auto-pilot/False").json(),
    )
    assert not result_ap.value_bool

    result_co = JsonUserConfigurationSingleOutput(
        **client.get(f"/api/user-config/{beamtime_id}/online-crystfel").json(),
    )
    assert result_co.value_bool is not None
    assert result_co.value_int is None

    result_co = JsonUserConfigurationSingleOutput(
        **client.patch(f"/api/user-config/{beamtime_id}/online-crystfel/True").json(),
    )
    assert result_co.value_bool

    result_co = JsonUserConfigurationSingleOutput(
        **client.patch(f"/api/user-config/{beamtime_id}/online-crystfel/False").json(),
    )
    assert not result_co.value_bool


def test_read_experiment_types(
    client: TestClient,
    beamtime_id: BeamtimeId,
    chemical_experiment_type_id: int,
) -> None:
    result = JsonReadExperimentTypes(
        **client.get(f"/api/experiment-types/{beamtime_id}").json(),
    )

    assert len(result.experiment_types) == 1
    assert result.experiment_types[0].id == chemical_experiment_type_id


def test_delete_experiment_types(
    client: TestClient,
    beamtime_id: BeamtimeId,
    chemical_experiment_type_id: int,
    run_channel_1_chemical_attributo_id: int,
    lyso_chemical_id: int,
) -> None:
    # to really test the experiment type deletion thing, let's create a data set for the experiment type as well
    create_response = JsonCreateDataSetOutput(
        **client.post(
            "/api/data-sets",
            json=JsonCreateDataSetInput(
                experiment_type_id=chemical_experiment_type_id,
                attributi=[
                    JsonAttributoValue(
                        attributo_id=run_channel_1_chemical_attributo_id,
                        # Good old Lyso!
                        attributo_value_chemical=lyso_chemical_id,
                    ),
                ],
            ).model_dump(),
        ).json(),
    )

    assert create_response.id > 0

    result = JsonDeleteExperimentTypeOutput(
        **client.request(
            "DELETE",
            "/api/experiment-types",
            json=JsonDeleteExperimentType(id=chemical_experiment_type_id).model_dump(),
        ).json(),
    )

    assert result.result

    assert not JsonReadExperimentTypes(
        **client.get(f"/api/experiment-types/{beamtime_id}").json(),
    ).experiment_types

    # Now retrieve the data set list, should be empty
    data_sets_response = JsonReadDataSets(
        **client.get(f"/api/data-sets/{beamtime_id}").json(),
    )
    assert not data_sets_response.data_sets


def test_delete_file(client: TestClient, test_file: int) -> None:
    result = JsonDeleteFileOutput(
        **client.request(
            "DELETE",
            "/api/files",
            json=JsonDeleteFileInput(id=test_file).model_dump(),
        ).json(),
    )
    assert result.id == test_file

    assert client.get(f"/api/files/{test_file}").status_code == 404


def test_update_beamtime_schedule(
    client: TestClient,
    beamtime_id: BeamtimeId,
    lyso_chemical_id: int,
) -> None:
    schedule_rows = [
        JsonBeamtimeScheduleRowInput(
            users="users",
            date="2025-04-30",
            shift="10:00-11:00",
            comment="comment",
            td_support="td_support",
            chemicals=[lyso_chemical_id],
        ),
        JsonBeamtimeScheduleRowInput(
            users="users2",
            date="2025-04-30",
            shift="11:00-12:00",
            comment="comment2",
            td_support="td_support2",
            chemicals=[],
        ),
    ]
    update_result = JsonBeamtimeScheduleOutput(
        **client.post(
            "/api/schedule",
            json=JsonUpdateBeamtimeScheduleInput(
                beamtime_id=beamtime_id,
                schedule=schedule_rows,
            ).model_dump(),
        ).json(),
    )

    for original, result in zip(schedule_rows, update_result.schedule, strict=False):
        assert original.date == result.date
        assert original.users == result.users
        assert original.shift == result.shift
        assert result.start is not None
        assert result.start_local is not None
        assert result.stop is not None
        assert result.stop_local is not None

    get_response = JsonBeamtimeScheduleOutput(
        **client.get(f"/api/schedule/{beamtime_id}").json(),
    )
    for original, result in zip(schedule_rows, get_response.schedule, strict=False):
        assert original.date == result.date
        assert original.users == result.users
        assert original.shift == result.shift


def test_create_live_stream_snapshot(
    client: TestClient,
    beamtime_id: BeamtimeId,
    test_file_path: Path,
) -> None:
    # Upload the file: for the live stream, it's important to use the proper file name, as that's the criterion for a file to be
    # a live stream (yes, ugly, I know).
    with test_file_path.open("rb") as upload_file:
        raw_output = client.post(
            "/api/files",
            data={"description": "test description", "deduplicate": str(False)},
            # To pass a file name as well as the file contents, httpx uses a tuple.
            files={"file": (f"live-stream-image-{beamtime_id}", upload_file)},
        )
        file = JsonCreateFileOutput(**raw_output.json())

    snapshot_response = JsonCreateLiveStreamSnapshotOutput(
        **client.get(f"/api/live-stream/snapshot/{beamtime_id}").json(),
    )

    assert snapshot_response.id != file.id
    assert snapshot_response.size_in_bytes == file.size_in_bytes


def test_update_live_stream(
    client: TestClient,
    beamtime_id: BeamtimeId,
    test_file_path: Path,
) -> None:
    # Upload the file: for the live stream, it's important to use the proper file name, as that's the criterion for a file to be
    # a live stream (yes, ugly, I know).
    with test_file_path.open("rb") as upload_file:
        raw_output = client.post(
            "/api/files",
            data={"description": "test description", "deduplicate": str(False)},
            # To pass a file name as well as the file contents, httpx uses a tuple.
            files={"file": (f"live-stream-image-{beamtime_id}", upload_file)},
        )
        original_file = JsonCreateFileOutput(**raw_output.json())

        raw_output = client.post(
            f"/api/live-stream/{beamtime_id}",
            # no file name needed here, it's implicit taht it's "the live stream"
            files={"file": upload_file},
        )
        new_file = JsonUpdateLiveStream(**raw_output.json())

        assert new_file.id == original_file.id


def test_check_standard_unit(client: TestClient) -> None:
    # First, test normalization
    output = JsonCheckStandardUnitOutput(
        **client.post(
            "/api/unit",
            json=JsonCheckStandardUnitInput(input="mm").model_dump(),
        ).json(),
    )
    assert output.error is None
    assert output.normalized == "1 millimeter"

    # Now try an invalid unit
    output = JsonCheckStandardUnitOutput(
        **client.post(
            "/api/unit",
            json=JsonCheckStandardUnitInput(input="bananas").model_dump(),
        ).json(),
    )
    assert output.error is not None
    assert output.normalized is None

    # Finally, empty string is also an important edge case
    output = JsonCheckStandardUnitOutput(
        **client.post(
            "/api/unit",
            json=JsonCheckStandardUnitInput(input="").model_dump(),
        ).json(),
    )
    assert output.error is not None
    assert output.normalized is None


def test_update_attributo(client: TestClient, beamtime_id: BeamtimeId) -> None:
    attributo_response = JsonCreateAttributoOutput(
        **client.post(
            "/api/attributi",
            json=JsonCreateAttributoInput(
                name="some attributo",
                description="description",
                group=ATTRIBUTO_GROUP_MANUAL,
                associated_table=AssociatedTable.CHEMICAL,
                attributo_type_string=JSONSchemaString(type="string", enum=None),
                beamtime_id=beamtime_id,
            ).model_dump(),
        ).json(),
    )
    assert attributo_response.id > 0

    chemical_response = JsonCreateChemicalOutput(
        **client.post(
            "/api/chemicals",
            json=JsonChemicalWithoutId(
                name=TEST_CHEMICAL_NAME,
                responsible_person=TEST_CHEMICAL_RESPONSIBLE_PERSON,
                attributi=[
                    JsonAttributoValue(
                        attributo_id=attributo_response.id,
                        # this is a string, but it encodes an integer, so we can change the type further down to "int"
                        attributo_value_str="1",
                    ),
                ],
                chemical_type=ChemicalType.CRYSTAL,
                file_ids=[],
                beamtime_id=beamtime_id,
            ).model_dump(),
        ).json(),
    )
    assert chemical_response.id > 0

    updated_attributo = JsonAttributo(
        id=attributo_response.id,
        # Change all attributes, except the associated table
        name="new name",
        description="new description",
        group="new group",
        associated_table=AssociatedTable.CHEMICAL,
        # even change the type to int here
        attributo_type_integer=JSONSchemaInteger(type="integer", format=None),
    )

    attributo_update_response = JsonUpdateAttributoOutput(
        **client.patch(
            "/api/attributi",
            json=JsonUpdateAttributoInput(
                attributo=updated_attributo,
                conversion_flags=JsonUpdateAttributoConversionFlags(ignore_units=True),
            ).model_dump(),
        ).json(),
    )
    assert attributo_update_response.id > 0

    # After updating, check that the attributi list is updated.
    read_attributi_response = JsonReadAttributi(
        **client.get(f"/api/attributi/{beamtime_id}").json(),
    )
    assert len(read_attributi_response.attributi) == 1

    assert read_attributi_response.attributi[0] == updated_attributo

    # Also, check that the chemicals list is proper.
    chemical_list = read_chemicals(client, beamtime_id)
    assert len(chemical_list.chemicals) == 1
    assert len(chemical_list.chemicals[0].attributi) == 1
    assert chemical_list.chemicals[0].attributi[0].attributo_value_str is None
    assert chemical_list.chemicals[0].attributi[0].attributo_value_int == 1


def test_delete_attributo(client: TestClient, beamtime_id: BeamtimeId) -> None:
    attributo_response = JsonCreateAttributoOutput(
        **client.post(
            "/api/attributi",
            json=JsonCreateAttributoInput(
                name="some attributo",
                description="description",
                group=ATTRIBUTO_GROUP_MANUAL,
                associated_table=AssociatedTable.CHEMICAL,
                attributo_type_string=JSONSchemaString(type="string", enum=None),
                beamtime_id=beamtime_id,
            ).model_dump(),
        ).json(),
    )
    assert attributo_response.id > 0

    chemical_response = JsonCreateChemicalOutput(
        **client.post(
            "/api/chemicals",
            json=JsonChemicalWithoutId(
                name=TEST_CHEMICAL_NAME,
                responsible_person=TEST_CHEMICAL_RESPONSIBLE_PERSON,
                attributi=[
                    JsonAttributoValue(
                        attributo_id=attributo_response.id,
                        # this is a string, but it encodes an integer, so we can change the type further down to "int"
                        attributo_value_str="1",
                    ),
                ],
                chemical_type=ChemicalType.CRYSTAL,
                file_ids=[],
                beamtime_id=beamtime_id,
            ).model_dump(),
        ).json(),
    )
    assert chemical_response.id > 0

    output = JsonDeleteAttributoOutput(
        **client.request(
            "DELETE",
            "/api/attributi",
            json=JsonDeleteAttributoInput(id=attributo_response.id).model_dump(),
        ).json(),
    )
    assert output.id == attributo_response.id
    read_attributi_response = JsonReadAttributi(
        **client.get(f"/api/attributi/{beamtime_id}").json(),
    )
    assert not read_attributi_response.attributi


def test_download_spreadsheet(
    client: TestClient,
    chemical_experiment_type_id: int,
    run_string_attributo_id: int,
    beamtime_id: BeamtimeId,
) -> None:
    # Set the experiment type (otherwise creating a run will fail - see above)
    set_current_experiment_type(client, beamtime_id, chemical_experiment_type_id)

    # Let's make the external run ID deliberately high
    external_run_id = 1000

    run_string_attributo_value = "foo"
    # Create the run and check the result
    create_run_response = JsonCreateOrUpdateRunOutput(
        **client.post(
            f"/api/runs/{external_run_id}",
            json=JsonCreateOrUpdateRun(
                files=[],
                beamtime_id=beamtime_id,
                attributi=[
                    JsonAttributoValue(
                        attributo_id=run_string_attributo_id,
                        attributo_value_str=run_string_attributo_value,
                    ),
                ],
                create_data_set=False,
                started=1,
                stopped=None,
            ).model_dump(),
        ).json(),
    )
    assert (
        create_run_response.run_internal_id is not None
        and create_run_response.run_internal_id > 0
    )

    response = client.get(f"/api/{beamtime_id}/spreadsheet.zip")
    print(response)
    with ZipFile(BytesIO(response.content)) as zipf:
        names = zipf.namelist()
        assert len(names) == 1
        assert names[0].endswith(".xlsx")

        with zipf.open(names[0]) as sheet:
            # This deviation with BytesIO(read) is apparently necessary. Bad.
            wb = load_workbook(filename=BytesIO(sheet.read()))

            ws = wb.active

            assert ws is not None

            # This test is highly incomplete. Let's add more conditions once we encounter bugs in this thing (yes, that's how testing works!)
            assert ws.title == "Runs"
            assert ws["A1"].value == "ID"
            assert ws["A2"].value == external_run_id


def test_create_run_attributi_from_schema(
    client: TestClient,
    beamtime_id: BeamtimeId,
) -> None:
    first_attributo_name = "first_attributo"
    second_attributo_name = "second_attributo"
    response = JsonCreateAttributiFromSchemaOutput(
        **client.post(
            "/api/attributi/schema",
            json=JsonCreateAttributiFromSchemaInput(
                beamtime_id=BeamtimeId(beamtime_id),
                attributi_schema=[
                    JsonCreateAttributiFromSchemaSingleAttributo(
                        attributo_name=first_attributo_name,
                        attributo_type=JSONSchemaInteger(type="integer", format=None),
                        description=first_attributo_name + "description",
                    ),
                    JsonCreateAttributiFromSchemaSingleAttributo(
                        attributo_name=second_attributo_name,
                        attributo_type=JSONSchemaString(type="string", enum=None),
                        description=second_attributo_name + "description",
                    ),
                ],
            ).model_dump(),
        ).json(),
    )
    assert response.created_attributi == 2

    read_response = JsonReadAttributi(
        **client.get(f"/api/attributi/{beamtime_id}").json(),
    )
    assert len(read_response.attributi) == 2
    assert set(a.name for a in read_response.attributi) == set(
        (first_attributo_name, second_attributo_name),
    )
    assert set(a.description for a in read_response.attributi) == set(
        (first_attributo_name + "description", second_attributo_name + "description"),
    )

    int_attributo = next(
        iter(a for a in read_response.attributi if a.name == first_attributo_name),
    )
    assert int_attributo.attributo_type_integer is not None
    string_attributo = next(
        iter(a for a in read_response.attributi if a.name == second_attributo_name),
    )
    assert string_attributo.attributo_type_string is not None


def test_create_two_compatible_data_sets(
    client: TestClient,
    chemical_experiment_type_id: int,
    lyso_chemical_id: int,
    run_channel_1_chemical_attributo_id: int,
) -> None:
    create_response = JsonCreateDataSetOutput(
        **client.post(
            "/api/data-sets",
            json=JsonCreateDataSetInput(
                experiment_type_id=chemical_experiment_type_id,
                attributi=[
                    JsonAttributoValue(
                        attributo_id=run_channel_1_chemical_attributo_id,
                        # Good old Lyso!
                        attributo_value_chemical=lyso_chemical_id,
                    ),
                ],
            ).model_dump(),
        ).json(),
    )

    assert create_response.id > 0

    second_create_response = client.post(
        "/api/data-sets",
        json=JsonCreateDataSetInput(
            experiment_type_id=chemical_experiment_type_id,
            attributi=[
                JsonAttributoValue(
                    attributo_id=run_channel_1_chemical_attributo_id,
                    # Good old Lyso!
                    attributo_value_chemical=lyso_chemical_id,
                ),
            ],
        ).model_dump(),
    )
    assert second_create_response.status_code // 100 == 4


async def read_indexing_jobs_wrapper(
    tmp_path: Path,
    request: web.Request,
) -> web.Response:
    url = f"{IN_MEMORY_DB_URL}/{tmp_path}/db"
    os.environ["DB_URL"] = url
    await init_db(url)

    result = get_orm_sessionmaker_with_url(os.environ["DB_URL"])

    async with result() as web_async_session:
        status = request.query.get("status")
        beamtime_id = request.query.get("beamtimeId")
        result = await read_indexing_jobs(
            status=DBJobStatus(status) if status is not None else None,
            beamtimeId=int(beamtime_id) if beamtime_id is not None else None,
            withFiles=request.query.get("withFiles") == "True",
            session=web_async_session,
        )
        return web.json_response(result.model_dump())


async def read_merge_jobs_wrapper(tmp_path: Path, request: web.Request) -> web.Response:
    url = f"{IN_MEMORY_DB_URL}/{tmp_path}/db"
    os.environ["DB_URL"] = url
    await init_db(url)

    result = get_orm_sessionmaker_with_url(os.environ["DB_URL"])

    async with result() as web_async_session:
        status = request.query.get("status")
        result = await read_merge_jobs(
            status=DBJobStatus(status) if status is not None else None,
            session=web_async_session,
        )
        print(result)
        return web.json_response(result.model_dump())


async def create_file_wrapper(tmp_path: Path, request: web.Request) -> web.Response:
    url = f"{IN_MEMORY_DB_URL}/{tmp_path}/db"
    os.environ["DB_URL"] = url
    await init_db(url)

    result = get_orm_sessionmaker_with_url(os.environ["DB_URL"])

    async with result() as web_async_session:
        data = await request.post()
        description = data["description"]
        assert isinstance(description, str)
        deduplicate = data["deduplicate"]
        assert isinstance(deduplicate, str)
        file = data["file"]
        assert isinstance(file, web.FileField)
        print(f"creating file {file}")
        result = await create_file(
            file=UploadFile(filename=file.filename, file=file.file),
            description=description,
            deduplicate=deduplicate,
            compress=CompressionMode.COMPRESS_AUTO.value,
            session=web_async_session,
        )
        return web.json_response(result.model_dump())


async def indexing_job_still_running_wrapper(
    tmp_path: Path,
    request: web.Request,
) -> web.Response:
    url = f"{IN_MEMORY_DB_URL}/{tmp_path}/db"
    os.environ["DB_URL"] = url
    await init_db(url)

    result = get_orm_sessionmaker_with_url(os.environ["DB_URL"])

    async with result() as web_async_session:
        result = await indexing_job_still_running(
            indexingResultId=int(request.match_info["indexingResultId"]),
            json_result=JsonIndexingResultStillRunning(**await request.json()),
            session=web_async_session,
        )
        return web.json_response(result.model_dump())


async def merge_job_started_wrapper(
    tmp_path: Path,
    request: web.Request,
) -> web.Response:
    url = f"{IN_MEMORY_DB_URL}/{tmp_path}/db"
    os.environ["DB_URL"] = url
    await init_db(url)

    result = get_orm_sessionmaker_with_url(os.environ["DB_URL"])

    async with result() as web_async_session:
        result = await merge_job_started(
            mergeResultId=int(request.match_info["mergeResultId"]),
            json_result=JsonMergeJobStartedInput(**await request.json()),
            session=web_async_session,
        )
        return web.json_response(result.model_dump())


async def merge_job_finished_wrapper(
    tmp_path: Path,
    request: web.Request,
) -> web.Response:
    url = f"{IN_MEMORY_DB_URL}/{tmp_path}/db"
    os.environ["DB_URL"] = url
    await init_db(url)

    result = get_orm_sessionmaker_with_url(os.environ["DB_URL"])

    async with result() as web_async_session:
        result = await merge_job_finished(
            mergeResultId=int(request.match_info["mergeResultId"]),
            json_result=JsonMergeJobFinishedInput(**await request.json()),
            session=web_async_session,
        )
        return web.json_response(result.model_dump())


async def indexing_job_finish_with_error_wrapper(
    tmp_path: Path,
    request: web.Request,
) -> web.Response:
    url = f"{IN_MEMORY_DB_URL}/{tmp_path}/db"
    os.environ["DB_URL"] = url
    await init_db(url)

    result = get_orm_sessionmaker_with_url(os.environ["DB_URL"])

    async with result() as web_async_session:
        result = await indexing_job_finish_with_error(
            indexingResultId=int(request.match_info["indexingResultId"]),
            json_result=JsonIndexingResultFinishWithError(**await request.json()),
            session=web_async_session,
        )
        return web.json_response(result.model_dump())


@pytest.fixture
async def daemon_session(aiohttp_client: Any, tmp_path: Path) -> aiohttp.ClientSession:
    """Create a test aiohttp HTTP server (one that doesn't run on a
    port on the host) and connect it to the real FastAPI handler
    functions in order to have semi-realistic testing of both the web
    server and the daemons
    """
    app = web.Application()
    app.router.add_get("/api/indexing", partial(read_indexing_jobs_wrapper, tmp_path))
    app.router.add_get("/api/merging", partial(read_merge_jobs_wrapper, tmp_path))
    app.router.add_post(
        "/api/merging/{mergeResultId}/start",
        partial(merge_job_started_wrapper, tmp_path),
    )
    app.router.add_post(
        "/api/merging/{mergeResultId}/finish",
        partial(merge_job_finished_wrapper, tmp_path),
    )
    app.router.add_post("/api/files", partial(create_file_wrapper, tmp_path))
    app.router.add_post(
        "/api/indexing/{indexingResultId}/still-running",
        partial(indexing_job_still_running_wrapper, tmp_path),
    )
    app.router.add_post(
        "/api/indexing/{indexingResultId}/finish-with-error",
        partial(indexing_job_finish_with_error_wrapper, tmp_path),
    )
    return await aiohttp_client(app)


async def test_indexing_daemon_start_job_but_then_vanish_from_workload_manager(
    client: TestClient,
    daemon_session: aiohttp.ClientSession,
    simple_data_set_id: int,
    geometry_id: int,
) -> None:
    os.environ[INDEXING_DAEMON_LONG_BREAK_DURATION_SECONDS_ENV_VAR] = "0.01"
    client.post(
        "/api/indexing",
        json=JsonCreateIndexingForDataSetInput(
            data_set_id=simple_data_set_id,
            is_online=False,
            cell_description="",
            geometry_id=geometry_id,
            command_line="",
            source="raw",
        ).model_dump(),
    )

    workload_manager = DummyWorkloadManager()
    args = indexing_daemon.Arguments()
    # Important for the test scenario: no absolute paths allowed, so this is "mockable"
    args.amarcord_url = ""
    # Also not really important, we don't actually start CrystFEL
    args.crystfel_path = Path("/usr/bin")

    workload_manager.job_start_results.append(
        JobStartResult(job_id=1337, metadata=JobMetadata({})),
    )
    # start the job
    await indexing_loop_iteration(
        workload_manager=workload_manager,
        session=daemon_session,
        args=args,
        start_new_jobs=True,
        online_workload_manager=None,
    )

    # Remove the job from the workload manager
    workload_manager.jobs.clear()

    await indexing_loop_iteration(
        workload_manager=workload_manager,
        online_workload_manager=None,
        session=daemon_session,
        args=args,
        start_new_jobs=False,
    )

    # Now get the indexing job from the DB and check that its status is indeed failed
    # Check the DB again
    indexing_jobs_result = JsonReadIndexingResultsOutput(
        **client.get("/api/indexing").json(),
    )
    assert len(indexing_jobs_result.indexing_jobs) == 1
    assert indexing_jobs_result.indexing_jobs[0].job_status == DBJobStatus.DONE
    assert indexing_jobs_result.indexing_jobs[0].started is not None
    assert indexing_jobs_result.indexing_jobs[0].stopped is not None


async def test_indexing_daemon_start_job_with_run_that_is_missing_files(
    client: TestClient,
    daemon_session: aiohttp.ClientSession,
    run_without_files_data_set_id: int,
    geometry_id: int,
) -> None:
    os.environ[INDEXING_DAEMON_LONG_BREAK_DURATION_SECONDS_ENV_VAR] = "0.01"
    client.post(
        "/api/indexing",
        json=JsonCreateIndexingForDataSetInput(
            data_set_id=run_without_files_data_set_id,
            is_online=False,
            cell_description="",
            geometry_id=geometry_id,
            command_line="",
            source="raw",
        ).model_dump(),
    )

    workload_manager = DummyWorkloadManager()
    args = indexing_daemon.Arguments()
    # Important for the test scenario: no absolute paths allowed, so this is "mockable"
    args.amarcord_url = ""
    # Also not really important, we don't actually start CrystFEL
    args.crystfel_path = Path("/usr/bin")

    workload_manager.job_start_results.append(
        JobStartResult(job_id=1337, metadata=JobMetadata({})),
    )

    # Now start jobs
    await indexing_loop_iteration(
        workload_manager=workload_manager,
        online_workload_manager=None,
        session=daemon_session,
        args=args,
        start_new_jobs=True,
    )

    assert not workload_manager.job_starts
    assert not list(await workload_manager.list_jobs())


async def test_geometry_creation_with_invalid_attributo_names(
    client: TestClient,
    beamtime_id: BeamtimeId,
) -> None:
    geometry_with_invalid_attributo = """clen = {{invalid_name}}"""

    response = client.post(
        "/api/geometries",
        json=JsonGeometryCreate(
            beamtime_id=beamtime_id,
            content=geometry_with_invalid_attributo,
            name="geometry name",
        ).model_dump(),
    )

    assert response.status_code == 400


async def test_geometry_creation_with_valid_attributo_name(
    client: TestClient,
    beamtime_id: BeamtimeId,
    run_string_attributo_id: int,  # noqa: ARG001
) -> None:
    geometry_with_run_string_attributo = """clen = {{run_string}}"""

    response = JsonGeometryWithoutContent(
        **client.post(
            "/api/geometries",
            json=JsonGeometryCreate(
                beamtime_id=beamtime_id,
                content=geometry_with_run_string_attributo,
                name="geometry name",
            ).model_dump(),
        ).json()
    )

    assert response.id > 0


async def test_geometry_with_usage_in_result(
    client: TestClient,
    beamtime_id: BeamtimeId,
    geometry_id: int,
    simple_indexing_result_id: int,  # noqa: ARG001
) -> None:
    # We have a geometry in the indexing parameters corresponding to
    # the simple result we pass here. So we should have usages in the list of geometries
    result = JsonReadGeometriesForSingleBeamtime(
        **client.get(f"/api/geometry-for-beamtime/{beamtime_id}").json()
    )

    assert len(result.geometries) == 1
    assert result.geometries[0].id == geometry_id
    assert len(result.geometry_with_usage) == 1
    assert result.geometry_with_usage[0].geometry_id == geometry_id
    assert result.geometry_with_usage[0].usages == 1


async def test_geometry_raw_with_variables(
    client: TestClient,
    geometry_id_with_run_string_attributo: int,
    indexing_result_id_with_run_string_attributo_geometry: int,
) -> None:
    result_raw = client.get(
        f"/api/geometries/{geometry_id_with_run_string_attributo}/raw?indexingResultId={indexing_result_id_with_run_string_attributo_geometry}"
    ).text

    assert result_raw == "clen foobar"


async def test_merge_daemon(
    client: TestClient,
    daemon_session: aiohttp.ClientSession,
    simple_data_set_id: int,
    # the indexing result must be created, but we create the merge result indirectly by specifying the data set ID
    simple_indexing_result_id: int,  # noqa: ARG001
) -> None:
    os.environ[MERGE_DAEMON_LONG_BREAK_DURATION_SECONDS_ENV_VAR] = "0.01"
    os.environ[MERGE_DAEMON_SHORT_BREAK_DURATION_SECONDS_ENV_VAR] = "0.01"
    workload_manager = DummyWorkloadManager()
    args = merge_daemon.Arguments()
    # Important for the test scenario: no absolute paths allowed, so this is "mockable"
    args.amarcord_url = ""
    # Also not really important, we don't actually start CrystFEL
    args.crystfel_path = Path("/usr/bin")
    args.ccp4_path = ""

    queue_merge_job_response = JsonQueueMergeJobOutput(
        **client.post(
            "/api/merging",
            json=JsonQueueMergeJobInput(
                strict_mode=False,
                data_set_id=simple_data_set_id,
                # Literally random stuff here, doesn't matter.
                indexing_parameters_id=1,
                merge_parameters=JsonMergeParameters(
                    cell_description=LYSO_CELL_DESCRIPTION,
                    point_group=LYSO_POINT_GROUP,
                    space_group=LYSO_SPACE_GROUP,
                    merge_model=MergeModel.UNITY,
                    scale_intensities=ScaleIntensities.OFF,
                    post_refinement=False,
                    iterations=3,
                    polarisation=JsonPolarisation(angle=30, percent=50),
                    negative_handling=MergeNegativeHandling.IGNORE,
                    start_after=None,
                    stop_after=None,
                    rel_b=1.0,
                    no_pr=False,
                    force_bandwidth=None,
                    force_radius=None,
                    force_lambda=None,
                    no_delta_cc_half=False,
                    max_adu=None,
                    min_measurements=1,
                    logs=False,
                    min_res=None,
                    push_res=None,
                    w=None,
                    ambigator_command_line="",
                ),
            ).model_dump(),
        ).json(),
    )

    workload_manager.job_start_results.append(
        JobStartResult(job_id=1337, metadata=JobMetadata({})),
    )

    # One iteration, should start the job on the workload manager
    await merging_loop_iteration(
        daemon_session,
        workload_manager,
        args,
        zombie_job_times={},
    )

    assert workload_manager.job_starts

    queued_merge_results = JsonReadMergeResultsOutput(
        **client.get(f"/api/merging?status={DBJobStatus.RUNNING.value}").json(),
    )

    assert len(queued_merge_results.merge_jobs) == 1
    assert (
        queued_merge_results.merge_jobs[0].id
        == queue_merge_job_response.merge_result_id
    )

    # Remove the job from the workload manager
    workload_manager.jobs.clear()

    # Since we have the "zombie" mechanic, we actually wait a bit for
    # the job to complete, so here we set the zombie job times (noting
    # when a job was first not seen anymore) as empty and expect the
    # job to still be running.
    await merging_loop_iteration(
        daemon_session,
        workload_manager,
        args,
        zombie_job_times={},
    )

    queued_merge_results = JsonReadMergeResultsOutput(
        **client.get(f"/api/merging?status={DBJobStatus.RUNNING.value}").json(),
    )

    assert len(queued_merge_results.merge_jobs) == 1

    # Now we set the job to be "long gone"
    await merging_loop_iteration(
        daemon_session,
        workload_manager,
        args,
        zombie_job_times={1337: 0.0},
    )

    queued_merge_results = JsonReadMergeResultsOutput(
        **client.get(f"/api/merging?status={DBJobStatus.RUNNING.value}").json(),
    )

    assert not queued_merge_results.merge_jobs
    # indexing_jobs_result = JsonReadIndexingResultsOutput(
    #     **client.get("/api/indexing").json()
    # )
    # assert len(indexing_jobs_result.indexing_jobs) == 1
    # assert indexing_jobs_result.indexing_jobs[0].job_status == DBJobStatus.QUEUED
    # assert indexing_jobs_result.indexing_jobs[0].stopped is None

    # print("second iteration, should start a job now")

    # # Now start jobs
    # await indexing_loop_iteration(
    #     workload_manager, daemon_session, args, start_new_jobs=True
    # )

    # assert len(workload_manager.job_starts) == 1
    # assert len(list(await workload_manager.list_jobs())) == 1
    # workload_manager.job_starts.clear()

    # # Check the DB again
    # indexing_jobs_result = JsonReadIndexingResultsOutput(
    #     **client.get("/api/indexing").json()
    # )
    # assert len(indexing_jobs_result.indexing_jobs) == 1
    # assert indexing_jobs_result.indexing_jobs[0].job_status == DBJobStatus.RUNNING
    # assert indexing_jobs_result.indexing_jobs[0].started is not None
    # assert indexing_jobs_result.indexing_jobs[0].stopped is None

    # print("third iteration, should _not_ start a job again")

    # # To be sure: another start iteration shouldn't do anything now
    # await indexing_loop_iteration(
    #     workload_manager, daemon_session, args, start_new_jobs=True
    # )

    # assert not workload_manager.job_starts
    # assert len(list(await workload_manager.list_jobs())) == 1

    # # Again, to be sure: another update shouldn't do anything
    # await indexing_loop_iteration(
    #     workload_manager, daemon_session, args, start_new_jobs=False
    # )

    # # Now we just assume the job we just started failed on the workload manager (i.e. SLURM)
    # workload_manager.jobs[0] = replace(
    #     # we say successful, but if the job didn't produce a result in time, that doesn't matter
    #     workload_manager.jobs[0],
    #     status=JobStatus.SUCCESSFUL,
    # )

    # print(
    #     "fourth iteration, job should be marked as failed, because it quit unexpectedly"
    # )

    # await indexing_loop_iteration(
    #     workload_manager, daemon_session, args, start_new_jobs=False
    # )

    # # Now get the indexing job from the DB and check that its status is indeed failed
    # # Check the DB again
    # indexing_jobs_result = JsonReadIndexingResultsOutput(
    #     **client.get("/api/indexing").json()
    # )
    # assert len(indexing_jobs_result.indexing_jobs) == 1
    # assert indexing_jobs_result.indexing_jobs[0].job_status == DBJobStatus.DONE
    # assert indexing_jobs_result.indexing_jobs[0].started is not None
    # assert indexing_jobs_result.indexing_jobs[0].stopped is not None


async def test_indexing_daemon_start_job_but_then_fail_unexpectedly(
    client: TestClient,
    daemon_session: aiohttp.ClientSession,
    simple_data_set_id: int,
    geometry_id: int,
) -> None:
    os.environ[INDEXING_DAEMON_LONG_BREAK_DURATION_SECONDS_ENV_VAR] = "0.01"
    create_response = JsonCreateIndexingForDataSetOutput(
        **client.post(
            "/api/indexing",
            json=JsonCreateIndexingForDataSetInput(
                data_set_id=simple_data_set_id,
                is_online=False,
                cell_description="",
                geometry_id=geometry_id,
                command_line="",
                source="raw",
            ).model_dump(),
        ).json(),
    )

    assert len(create_response.jobs_started_run_external_ids) == 1
    assert create_response.data_set_id == simple_data_set_id
    assert create_response.indexing_parameters_id > 0

    workload_manager = DummyWorkloadManager()
    args = indexing_daemon.Arguments()
    # Important for the test scenario: no absolute paths allowed, so this is "mockable"
    args.amarcord_url = ""
    # Also not really important, we don't actually start CrystFEL
    args.crystfel_path = Path("/usr/bin")

    # One iteration without starting jobs => shouldn't start jobs!
    await indexing_loop_iteration(
        workload_manager=workload_manager,
        online_workload_manager=None,
        session=daemon_session,
        args=args,
        start_new_jobs=False,
    )

    print(
        "first iteration, should not do anything because it's just an update iteration",
    )

    assert not workload_manager.job_starts

    workload_manager.job_start_results.append(
        JobStartResult(job_id=1337, metadata=JobMetadata({})),
    )

    indexing_jobs_result = JsonReadIndexingResultsOutput(
        **client.get("/api/indexing").json(),
    )
    assert len(indexing_jobs_result.indexing_jobs) == 1
    assert indexing_jobs_result.indexing_jobs[0].job_status == DBJobStatus.QUEUED
    assert indexing_jobs_result.indexing_jobs[0].stopped is None

    print("second iteration, should start a job now")

    # Now start jobs
    await indexing_loop_iteration(
        workload_manager=workload_manager,
        online_workload_manager=None,
        session=daemon_session,
        args=args,
        start_new_jobs=True,
    )

    assert len(workload_manager.job_starts) == 1
    assert len(list(await workload_manager.list_jobs())) == 1
    workload_manager.job_starts.clear()

    # Check the DB again
    indexing_jobs_result = JsonReadIndexingResultsOutput(
        **client.get("/api/indexing").json(),
    )
    assert len(indexing_jobs_result.indexing_jobs) == 1
    assert indexing_jobs_result.indexing_jobs[0].job_status == DBJobStatus.RUNNING
    assert indexing_jobs_result.indexing_jobs[0].started is not None
    assert indexing_jobs_result.indexing_jobs[0].stopped is None

    print("third iteration, should _not_ start a job again")

    # To be sure: another start iteration shouldn't do anything now
    await indexing_loop_iteration(
        workload_manager=workload_manager,
        online_workload_manager=None,
        session=daemon_session,
        args=args,
        start_new_jobs=True,
    )

    assert not workload_manager.job_starts
    assert len(list(await workload_manager.list_jobs())) == 1

    # Again, to be sure: another update shouldn't do anything
    await indexing_loop_iteration(
        workload_manager=workload_manager,
        online_workload_manager=None,
        session=daemon_session,
        args=args,
        start_new_jobs=True,
    )

    # Now we just assume the job we just started failed on the workload manager (i.e. SLURM)
    workload_manager.jobs[0] = replace(
        # we say successful, but if the job didn't produce a result in time, that doesn't matter
        workload_manager.jobs[0],
        status=JobStatus.SUCCESSFUL,
    )

    print(
        "fourth iteration, job should be marked as failed, because it quit unexpectedly",
    )

    await indexing_loop_iteration(
        workload_manager=workload_manager,
        online_workload_manager=None,
        session=daemon_session,
        args=args,
        start_new_jobs=False,
    )

    # Now get the indexing job from the DB and check that its status is indeed failed
    # Check the DB again
    indexing_jobs_result = JsonReadIndexingResultsOutput(
        **client.get("/api/indexing").json(),
    )
    assert len(indexing_jobs_result.indexing_jobs) == 1
    assert indexing_jobs_result.indexing_jobs[0].job_status == DBJobStatus.DONE
    assert indexing_jobs_result.indexing_jobs[0].started is not None
    assert indexing_jobs_result.indexing_jobs[0].stopped is not None

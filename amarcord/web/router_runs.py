import datetime
import re
from dataclasses import dataclass
from io import BytesIO
from typing import Annotated
from typing import Any
from typing import Generator
from typing import Iterable
from typing import Mapping

import structlog
from fastapi import APIRouter
from fastapi import Depends
from fastapi import HTTPException
from fastapi import Response
from fastapi import UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl import load_workbook
from sqlalchemy import true
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from sqlalchemy.sql import delete
from sqlalchemy.sql import select

from amarcord.cli.crystfel_index import coparse_cell_description
from amarcord.db import orm
from amarcord.db.associated_table import AssociatedTable
from amarcord.db.attributi import local_int_to_utc_datetime
from amarcord.db.attributi import run_matches_dataset
from amarcord.db.attributi import schema_dict_to_attributo_type
from amarcord.db.attributi import utc_datetime_to_local_int
from amarcord.db.attributi import utc_datetime_to_utc_int
from amarcord.db.attributi import utc_int_to_utc_datetime
from amarcord.db.attributo_id import AttributoId
from amarcord.db.attributo_type import AttributoType
from amarcord.db.attributo_value import AttributoValue
from amarcord.db.beamtime_id import BeamtimeId
from amarcord.db.db_job_status import DBJobStatus
from amarcord.db.event_log_level import EventLogLevel
from amarcord.db.excel_import import ConversionError
from amarcord.db.excel_import import SpreadsheetValidationErrors
from amarcord.db.excel_import import create_data_set_for_runs
from amarcord.db.excel_import import create_runs_from_spreadsheet
from amarcord.db.excel_import import parse_run_spreadsheet_workbook
from amarcord.db.indexing_result import IndexingResultSummary
from amarcord.db.indexing_result import empty_indexing_fom
from amarcord.db.orm_utils import ATTRIBUTO_GROUP_MANUAL
from amarcord.db.orm_utils import data_sets_are_equal
from amarcord.db.orm_utils import default_online_indexing_parameters
from amarcord.db.orm_utils import determine_run_indexing_metadata
from amarcord.db.orm_utils import duplicate_run_attributo
from amarcord.db.orm_utils import live_stream_image_name
from amarcord.db.orm_utils import retrieve_latest_config
from amarcord.db.orm_utils import retrieve_latest_run
from amarcord.db.orm_utils import run_has_attributo_to_data_set_has_attributo
from amarcord.db.orm_utils import validate_json_attributo_return_error
from amarcord.db.run_external_id import RunExternalId
from amarcord.db.run_internal_id import RunInternalId
from amarcord.filter_expression import FilterInput
from amarcord.filter_expression import FilterParseError
from amarcord.filter_expression import compile_run_filter
from amarcord.util import get_local_tz
from amarcord.web.constants import DATE_FORMAT
from amarcord.web.fastapi_utils import encode_data_set_attributo_value
from amarcord.web.fastapi_utils import encode_run_attributo_value
from amarcord.web.fastapi_utils import event_has_date
from amarcord.web.fastapi_utils import get_orm_db
from amarcord.web.fastapi_utils import json_attributo_to_run_orm_attributo
from amarcord.web.fastapi_utils import run_has_started_date
from amarcord.web.fastapi_utils import safe_create_new_event
from amarcord.web.fastapi_utils import update_attributi_from_json
from amarcord.web.json_models import JsonAttributoBulkValue
from amarcord.web.json_models import JsonAttributoValue
from amarcord.web.json_models import JsonCreateOrUpdateRun
from amarcord.web.json_models import JsonCreateOrUpdateRunOutput
from amarcord.web.json_models import JsonDataSet
from amarcord.web.json_models import JsonDataSetWithFom
from amarcord.web.json_models import JsonDeleteRunOutput
from amarcord.web.json_models import JsonIndexingStatistic
from amarcord.web.json_models import JsonLiveStream
from amarcord.web.json_models import JsonReadRuns
from amarcord.web.json_models import JsonReadRunsBulkInput
from amarcord.web.json_models import JsonReadRunsBulkOutput
from amarcord.web.json_models import JsonReadRunsOverview
from amarcord.web.json_models import JsonRun
from amarcord.web.json_models import JsonRunAnalysisIndexingResult
from amarcord.web.json_models import JsonRunFile
from amarcord.web.json_models import JsonRunsBulkImportInfo
from amarcord.web.json_models import JsonRunsBulkImportOutput
from amarcord.web.json_models import JsonStartRunOutput
from amarcord.web.json_models import JsonStopRunOutput
from amarcord.web.json_models import JsonUpdateRun
from amarcord.web.json_models import JsonUpdateRunOutput
from amarcord.web.json_models import JsonUpdateRunsBulkInput
from amarcord.web.json_models import JsonUpdateRunsBulkOutput
from amarcord.web.router_attributi import encode_attributo
from amarcord.web.router_chemicals import encode_chemical
from amarcord.web.router_data_sets import encode_orm_data_set_to_json
from amarcord.web.router_events import encode_event
from amarcord.web.router_experiment_types import encode_experiment_type
from amarcord.web.router_indexing import encode_indexing_fom_to_json
from amarcord.web.router_indexing import fom_for_indexing_result
from amarcord.web.router_indexing import summary_from_foms
from amarcord.web.router_user_configuration import encode_user_configuration

logger = structlog.stdlib.get_logger(__name__)
router = APIRouter()
_SHIFT_RE = re.compile(r"(\d{2}):(\d{2})-(\d{2}):(\d{2})")


def extract_runs_and_event_dates(
    runs: Iterable[orm.Run],
    events: Iterable[orm.EventLog],
) -> list[str]:
    set_of_dates: set[str] = set()
    for run in runs:
        started_date = run.started
        set_of_dates.add(started_date.strftime(DATE_FORMAT))
    for event in events:
        set_of_dates.add(event.created.strftime(DATE_FORMAT))

    return sorted(set_of_dates, reverse=True)


def indexing_fom_for_run(
    indexing_results_for_runs: Mapping[RunInternalId, list[orm.IndexingResult]],
    run: orm.Run,
) -> IndexingResultSummary:
    indexing_results_for_run = sorted(
        [
            (ir.id, fom_for_indexing_result(ir))
            for ir in indexing_results_for_runs.get(run.id, [])
            if not ir.job_error
        ],
        key=lambda ir: ir[0],
        reverse=True,
    )
    if indexing_results_for_run:
        return indexing_results_for_run[0][1]
    return empty_indexing_fom


@router.get(
    "/api/runs/{runExternalId}/start/{beamtimeId}",
    tags=["runs"],
    response_model_exclude_defaults=True,
)
async def start_run(
    runExternalId: RunExternalId,  # noqa: N803
    beamtimeId: BeamtimeId,  # noqa: N803
    session: Annotated[AsyncSession, Depends(get_orm_db)],
) -> JsonStartRunOutput:
    async with session.begin():
        latest_config = await retrieve_latest_config(session, beamtimeId)
        experiment_type_id = latest_config.current_experiment_type_id
        if experiment_type_id is None:
            raise HTTPException(
                status_code=400,
                detail="Cannot create run, no experiment type set!",
            )
        new_run = orm.Run(
            external_id=runExternalId,
            experiment_type_id=experiment_type_id,
            beamtime_id=beamtimeId,
            started=datetime.datetime.now(datetime.timezone.utc),
            stopped=None,
            modified=datetime.datetime.now(datetime.timezone.utc),
        )
        if latest_config.auto_pilot:
            latest_run = await retrieve_latest_run(session, beamtimeId)
            if latest_run is not None:
                for latest_run_attributo in latest_run.attributo_values:
                    if (
                        await latest_run_attributo.awaitable_attrs.attributo
                    ).group == ATTRIBUTO_GROUP_MANUAL:
                        new_run.attributo_values.append(
                            duplicate_run_attributo(latest_run_attributo),
                        )
        session.add(new_run)
        await session.flush()
        return JsonStartRunOutput(run_internal_id=new_run.id)


@router.get(
    "/api/runs/stop-latest/{beamtimeId}",
    tags=["runs"],
    response_model_exclude_defaults=True,
)
async def stop_latest_run(
    beamtimeId: BeamtimeId,  # noqa: N803
    session: Annotated[AsyncSession, Depends(get_orm_db)],
) -> JsonStopRunOutput:
    async with session.begin():
        latest_run = await retrieve_latest_run(session, beamtimeId)

        if latest_run is not None:
            latest_run.stopped = datetime.datetime.now(datetime.timezone.utc)
            await session.commit()
            return JsonStopRunOutput(result=True)

        return JsonStopRunOutput(result=False)


async def _create_data_set_for_run(
    session: AsyncSession,
    latest_config: orm.UserConfiguration,
    run: orm.Run,
) -> None:
    current_experiment_type = (
        await latest_config.awaitable_attrs.current_experiment_type
    )
    new_data_set = orm.DataSet(experiment_type_id=current_experiment_type.id)
    for et_attributo in current_experiment_type.attributi:
        for run_attributo in run.attributo_values:
            if run_attributo.attributo_id == et_attributo.attributo_id:
                new_data_set.attributo_values.append(
                    run_has_attributo_to_data_set_has_attributo(run_attributo),
                )
    if len(new_data_set.attributo_values) != len(current_experiment_type.attributi):
        for existing_attributo in run.experiment_type.attributi:
            found = False
            for data_set_attributo in new_data_set.attributo_values:
                if data_set_attributo.attributo_id == existing_attributo.attributo_id:
                    found = True
                    break
            if not found:
                raise Exception(
                    f"run {run.external_id}: tried to create a data set for experiment type “{run.experiment_type.name}”, but attributo “{existing_attributo.attributo.name}” not found in this run"
                )
    if not new_data_set.attributo_values:
        raise Exception(
            f"run {run.external_id}: found no attributo values to create the data set"
        )
    existing_data_sets: list[orm.DataSet] = list(
        (
            await session.scalars(
                select(orm.DataSet, orm.ExperimentType)
                .join(orm.DataSet.experiment_type)
                .where(
                    orm.ExperimentType.beamtime_id
                    == current_experiment_type.beamtime_id
                ),
            )
        ).all()
    )
    have_equal = False
    for existing_ds in existing_data_sets:
        if data_sets_are_equal(existing_ds, new_data_set):
            have_equal = True
    if not have_equal:
        session.add(new_data_set)


@router.post(
    "/api/runs/{runExternalId}",
    tags=["runs"],
    response_model_exclude_defaults=True,
)
async def create_or_update_run(
    runExternalId: RunExternalId,  # noqa: N803
    input_: JsonCreateOrUpdateRun,
    session: Annotated[AsyncSession, Depends(get_orm_db)],
) -> JsonCreateOrUpdateRunOutput:
    beamtime_id = input_.beamtime_id
    run_logger = logger.bind(run_external_id=runExternalId, beamtime_id=beamtime_id)
    run_logger.info("creating (or updating) run")

    async with session.begin():
        latest_config = await retrieve_latest_config(session, beamtime_id)
        run_logger.info(
            f"retrieved latest user config: {latest_config.use_online_crystfel}",
        )
        experiment_type_id = latest_config.current_experiment_type_id
        if experiment_type_id is None:
            run_logger.error("no experiment type set, cannot create run")
            raise HTTPException(
                status_code=400,
                detail="Cannot create run, no experiment type set!",
            )
        run_logger.info(f"experiment type set to {experiment_type_id}")
        run_in_db = (
            await session.scalars(
                select(orm.Run)
                .where(
                    (orm.Run.external_id == runExternalId)
                    & (orm.Run.beamtime_id == beamtime_id),
                )
                .options(selectinload(orm.Run.files)),
            )
        ).one_or_none()
        run_was_created = run_in_db is None
        if run_in_db is None:
            run_logger.info("run not in DB, creating a new one")
            run_in_db = orm.Run(
                external_id=runExternalId,
                experiment_type_id=experiment_type_id,
                beamtime_id=beamtime_id,
                started=(
                    datetime.datetime.now(datetime.timezone.utc)
                    if input_.started is None
                    else utc_int_to_utc_datetime(input_.started)
                    if input_.is_utc
                    else local_int_to_utc_datetime(input_.started)
                ),
                stopped=(
                    None
                    if input_.stopped is None
                    else utc_int_to_utc_datetime(input_.stopped)
                    if input_.is_utc
                    else local_int_to_utc_datetime(input_.stopped)
                ),
                modified=datetime.datetime.now(datetime.timezone.utc),
            )

            attributi_by_id: dict[int, orm.Attributo] = {
                orm_attributo.id: orm_attributo
                for orm_attributo in (
                    await session.scalars(
                        select(orm.Attributo).where(
                            orm.Attributo.id.in_(
                                input_attributo.attributo_id
                                for input_attributo in input_.attributi
                            )
                            & (orm.Attributo.associated_table == AssociatedTable.RUN),
                        ),
                    )
                )
            }
            attributo_ids_already_in_run: set[int] = set()
            for new_attributo in input_.attributi:
                attributo_type = attributi_by_id.get(new_attributo.attributo_id)
                if attributo_type is None:
                    raise HTTPException(
                        status_code=400,
                        detail=f"attributo with ID {new_attributo.attributo_id} not found in list of run attributi",
                    )
                validation_result = validate_json_attributo_return_error(
                    new_attributo,
                    attributo_type,
                )
                run_logger.info(
                    f"validating type of {new_attributo.attributo_id}: type is {attributo_type.json_schema}: {validation_result}",
                )
                if validation_result is not None:
                    raise HTTPException(
                        status_code=400,
                        detail=f"error validating attributi: {validation_result}",
                    )
                run_in_db.attributo_values.append(
                    json_attributo_to_run_orm_attributo(new_attributo),
                )
                attributo_ids_already_in_run.add(new_attributo.attributo_id)
            if latest_config.auto_pilot:
                latest_run = await retrieve_latest_run(session, beamtime_id)
                if latest_run is not None:
                    for latest_run_attributo in latest_run.attributo_values:
                        if (
                            latest_run_attributo.attributo_id
                            not in attributo_ids_already_in_run
                            and (
                                await latest_run_attributo.awaitable_attrs.attributo
                            ).group
                            == ATTRIBUTO_GROUP_MANUAL
                        ):
                            run_in_db.attributo_values.append(
                                duplicate_run_attributo(latest_run_attributo),
                            )
            if input_.files is not None:
                for run_file in input_.files:
                    run_in_db.files.append(
                        orm.RunHasFiles(glob=run_file.glob, source=run_file.source)
                    )
            session.add(run_in_db)
            # we might have a new run, and added a chemical to it, but the chemical relationship hasn't been loaded
            # for that. That we do here by flushing
            await session.flush()
        else:
            run_logger.info("run in DB, updating attributes")
            await update_attributi_from_json(
                session,
                db_item=run_in_db,
                new_attributi=input_.attributi,
                attributi_by_id={
                    a.id: a
                    for a in (
                        await session.scalars(
                            select(orm.Attributo).where(
                                orm.Attributo.id.in_(
                                    a.attributo_id for a in input_.attributi
                                )
                                & (
                                    orm.Attributo.associated_table
                                    == AssociatedTable.RUN
                                ),
                            ),
                        )
                    )
                },
            )
            if input_.files is not None and input_.files:
                # This only adds new files. If we wanted to replace, we need another flag
                for f in input_.files:
                    run_in_db.files.append(
                        orm.RunHasFiles(glob=f.glob, source=f.source)
                    )
            if input_.started is not None:
                run_in_db.started = utc_int_to_utc_datetime(input_.started)
            if input_.stopped is not None:
                run_in_db.stopped = utc_int_to_utc_datetime(input_.stopped)

        if input_.create_data_set:
            await _create_data_set_for_run(session, latest_config, run_in_db)

        async def _inner_create_new_event(text: str) -> None:
            run_logger.error(text)
            await safe_create_new_event(
                logger,
                session,
                beamtime_id,
                f"run {runExternalId}: {text}",
                EventLogLevel.INFO,
                "API",
            )

        if not run_was_created:
            run_logger.info("CrystFEL online not needed, run is updated, not created")
            indexing_result_id = None
        elif not latest_config.use_online_crystfel:
            run_logger.info(
                "CrystFEL online deactivated (or never explicitly activated), not creating indexing job",
            )
            indexing_result_id = None
        else:
            run_logger.info("adding CrystFEL online job")

            run_indexing_metadata = await determine_run_indexing_metadata(
                session,
                run_in_db,
            )

            if isinstance(run_indexing_metadata, str):
                message = f"cannot start CrystFEL online: {run_indexing_metadata}"
                await _inner_create_new_event(message)
                raise HTTPException(
                    status_code=400,
                    detail=message,
                )

            run_logger.info(
                f"creating CrystFEL online job for chemical {run_indexing_metadata.chemical.id}",
            )
            latest_user_config = await retrieve_latest_config(session, beamtime_id)
            current_online_indexing_parameters = await latest_user_config.awaitable_attrs.current_online_indexing_parameters
            if current_online_indexing_parameters is None:
                current_online_indexing_parameters = (
                    default_online_indexing_parameters()
                )
            # We _could_ re-use the same indexing parameters from the configuration each time. But
            # if we don't have one set in the config, then we'd have to create it here, and I was
            # too lazy to figure out the consequences.
            new_indexing_result_parameters = orm.IndexingParameters(
                is_online=True,
                cell_description=(
                    coparse_cell_description(run_indexing_metadata.cell_description)
                    if run_indexing_metadata.cell_description is not None
                    else None
                ),
                command_line=current_online_indexing_parameters.command_line,
                geometry_file=current_online_indexing_parameters.geometry_file,
                source=current_online_indexing_parameters.source,
            )
            session.add(new_indexing_result_parameters)
            # Better to explicitly flush, creating the run and giving us the ID
            await session.flush()
            new_indexing_result = orm.IndexingResult(
                created=datetime.datetime.now(datetime.timezone.utc),
                run_id=run_in_db.id,
                stream_file=None,
                # program version will be determined by the job itself and sent back
                program_version="",
                frames=0,
                hits=0,
                indexed_frames=0,
                # autodetect geometry file for online indexing
                geometry_file=None,
                geometry_hash="",
                generated_geometry_file=None,
                unit_cell_histograms_file_id=None,
                job_id=None,
                job_status=DBJobStatus.QUEUED,
                job_error=None,
                job_latest_log="",
                job_started=None,
                job_stopped=None,
                indexing_parameters_id=new_indexing_result_parameters.id,
            )
            session.add(new_indexing_result)
            await session.flush()
            indexing_result_id = new_indexing_result.id

    return JsonCreateOrUpdateRunOutput(
        run_created=run_was_created,
        indexing_result_id=indexing_result_id,
        error_message=None,
        run_internal_id=run_in_db.id,
        files=[
            JsonRunFile(id=f.id, glob=f.glob, source=f.source) for f in run_in_db.files
        ],
    )


@router.patch("/api/runs", tags=["runs"], response_model_exclude_defaults=True)
async def update_run(
    input_: JsonUpdateRun,
    session: Annotated[AsyncSession, Depends(get_orm_db)],
) -> JsonUpdateRunOutput:
    async with session.begin():
        run_id = RunInternalId(input_.id)
        current_run = (
            await session.scalars(
                select(orm.Run)
                .where(orm.Run.id == run_id)
                .options(selectinload(orm.Run.files))
            )
        ).one()
        if input_.files is not None:
            # clearing doesn't work because implicit IO
            # current_run.files.clear()
            for f in current_run.files:
                await session.delete(f)
            await session.refresh(current_run)
            for new_run_file in input_.files:
                current_run.files.append(
                    orm.RunHasFiles(glob=new_run_file.glob, source=new_run_file.source)
                )
        await update_attributi_from_json(
            session,
            db_item=current_run,
            new_attributi=input_.attributi,
            attributi_by_id={
                a.id: a
                for a in (
                    await session.scalars(
                        select(orm.Attributo).where(
                            orm.Attributo.id.in_(
                                a.attributo_id for a in input_.attributi
                            )
                            & (orm.Attributo.associated_table == AssociatedTable.RUN),
                        ),
                    )
                )
            },
        )
        current_run.experiment_type_id = input_.experiment_type_id
        await session.commit()

    return JsonUpdateRunOutput(
        result=True,
        files=[
            JsonRunFile(id=f.id, source=f.source, glob=f.glob)
            for f in current_run.files
        ],
    )


def encode_attributo_value(
    attributo_id: int,
    attributo_value: AttributoValue,
) -> JsonAttributoValue:
    return JsonAttributoValue(
        attributo_id=attributo_id,
        attributo_value_str=(
            attributo_value if isinstance(attributo_value, str) else None
        ),
        attributo_value_int=(
            attributo_value if isinstance(attributo_value, int) else None
        ),
        attributo_value_float=(
            attributo_value if isinstance(attributo_value, float) else None
        ),
        attributo_value_bool=(
            attributo_value if isinstance(attributo_value, bool) else None
        ),
        attributo_value_datetime=(
            utc_datetime_to_utc_int(attributo_value)
            if isinstance(attributo_value, datetime.datetime)
            else None
        ),
        attributo_value_datetime_local=(
            utc_datetime_to_local_int(attributo_value)
            if isinstance(attributo_value, datetime.datetime)
            else None
        ),
        # we cannot thoroughly test the array for type-correctness (or we dont' want to, rather)
        attributo_value_list_str=(  # pyright: ignore
            attributo_value
            if isinstance(attributo_value, list)
            and (not attributo_value or isinstance(attributo_value[0], str))
            else None
        ),
        attributo_value_list_float=(  # pyright: ignore
            attributo_value
            if isinstance(attributo_value, list)
            and (not attributo_value or isinstance(attributo_value[0], float | int))
            else None
        ),
        attributo_value_list_bool=(  # pyright: ignore
            attributo_value
            if isinstance(attributo_value, list)
            and (not attributo_value or isinstance(attributo_value[0], bool))
            else None
        ),
    )


@dataclass(frozen=True)
class _RunHasAttributoValueToBeUsedInSet:
    integer_value: None | int
    float_value: None | float
    string_value: None | str
    bool_value: None | bool
    datetime_value: None | datetime.datetime
    list_value: None | list[Any]
    chemical_value: None | int


def _to_dataclass(o: orm.RunHasAttributoValue) -> _RunHasAttributoValueToBeUsedInSet:
    return _RunHasAttributoValueToBeUsedInSet(
        integer_value=o.integer_value,
        float_value=o.float_value,
        string_value=o.string_value,
        bool_value=o.bool_value,
        datetime_value=o.datetime_value,
        list_value=o.list_value,
        chemical_value=o.chemical_value,
    )


def _encode_dataclass(
    id_: AttributoId,
    r: _RunHasAttributoValueToBeUsedInSet,
) -> JsonAttributoValue:
    return JsonAttributoValue(
        attributo_id=id_,
        attributo_value_str=r.string_value,
        attributo_value_int=r.integer_value,
        attributo_value_chemical=r.chemical_value,
        attributo_value_datetime=(
            utc_datetime_to_utc_int(r.datetime_value)
            if r.datetime_value is not None
            else None
        ),
        attributo_value_datetime_local=(
            utc_datetime_to_local_int(r.datetime_value)
            if r.datetime_value is not None
            else None
        ),
        attributo_value_float=r.float_value,
        attributo_value_bool=r.bool_value,
        attributo_value_list_str=(
            r.list_value if r.list_value and isinstance(r.list_value[0], str) else None
        ),
        attributo_value_list_float=(
            r.list_value
            if r.list_value and isinstance(r.list_value[0], float | int)
            else None
        ),
        attributo_value_list_bool=(
            r.list_value if r.list_value and isinstance(r.list_value[0], bool) else None
        ),
    )


@router.post("/api/runs-bulk", tags=["runs"], response_model_exclude_defaults=True)
async def read_runs_bulk(
    input_: JsonReadRunsBulkInput,
    session: Annotated[AsyncSession, Depends(get_orm_db)],
) -> JsonReadRunsBulkOutput:
    beamtime_id = input_.beamtime_id
    attributi = list(
        (
            await session.scalars(
                select(orm.Attributo)
                .where(orm.Attributo.beamtime_id == beamtime_id)
                .order_by(orm.Attributo.name),
            )
        ).all(),
    )
    chemicals = (
        await session.scalars(
            select(orm.Chemical)
            .where(orm.Chemical.beamtime_id == beamtime_id)
            .options(selectinload(orm.Chemical.files)),
        )
    ).all()
    all_runs = (
        await session.scalars(
            select(orm.Run).where(
                # Important! Since we're using the external run ID for this request, we need to constrain the
                # beamtime ID here, since two beamtimes will surely have overlapping run IDs
                (orm.Run.external_id.in_(input_.external_run_ids))
                & (orm.Run.beamtime_id == beamtime_id),
            ),
        )
    ).all()
    bulk_attributi: dict[AttributoId, set[_RunHasAttributoValueToBeUsedInSet]] = {
        a.id: set() for a in attributi
    }
    for r in all_runs:
        for a in r.attributo_values:
            bulk_attributi[a.attributo_id].add(_to_dataclass(a))
    return JsonReadRunsBulkOutput(
        chemicals=[encode_chemical(s) for s in chemicals],
        attributi=[
            encode_attributo(a)
            for a in attributi
            if a.associated_table == AssociatedTable.RUN
        ],
        attributi_values=[
            JsonAttributoBulkValue(
                attributo_id=attributo_id,
                values=[_encode_dataclass(attributo_id, v) for v in values],
            )
            for attributo_id, values in bulk_attributi.items()
        ],
        experiment_types=[
            encode_experiment_type(a)
            for a in (
                await session.scalars(
                    select(orm.ExperimentType).where(
                        orm.ExperimentType.beamtime_id == beamtime_id,
                    ),
                )
            ).all()
        ],
        experiment_type_ids=list(set(r.experiment_type_id for r in all_runs)),
    )


@router.patch("/api/runs-bulk", tags=["runs"], response_model_exclude_defaults=True)
async def update_runs_bulk(
    input_: JsonUpdateRunsBulkInput,
    session: Annotated[AsyncSession, Depends(get_orm_db)],
) -> JsonUpdateRunsBulkOutput:
    async with session.begin():
        for run in await session.scalars(
            select(orm.Run).where(
                (orm.Run.beamtime_id == input_.beamtime_id)
                & (orm.Run.external_id.in_(input_.external_run_ids)),
            ),
        ):
            if input_.new_experiment_type_id is not None:
                run.experiment_type_id = input_.new_experiment_type_id
                await update_attributi_from_json(
                    session,
                    db_item=run,
                    new_attributi=input_.attributi,
                    attributi_by_id={
                        a.id: a
                        for a in (
                            await session.scalars(
                                select(orm.Attributo).where(
                                    orm.Attributo.id.in_(
                                        a.attributo_id for a in input_.attributi
                                    )
                                    & (
                                        orm.Attributo.associated_table
                                        == AssociatedTable.RUN
                                    ),
                                ),
                            )
                        )
                    },
                )
        return JsonUpdateRunsBulkOutput(result=True)


async def _find_schedule_entry(
    session: AsyncSession,
    beamtime_id: BeamtimeId,
) -> None | orm.BeamtimeSchedule:
    now = datetime.datetime.now(get_local_tz())
    minutes_since_midnight_now = now.hour * 60 + now.minute
    for schedule_entry in await session.scalars(
        select(orm.BeamtimeSchedule).where(
            (orm.BeamtimeSchedule.beamtime_id == beamtime_id)
            & (orm.BeamtimeSchedule.date == now.strftime("%Y-%m-%d")),
        ),
    ):
        entry_match = _SHIFT_RE.search(schedule_entry.shift)
        if entry_match is not None:
            minutes_since_midnight_from = int(entry_match.group(1)) * 60 + int(
                entry_match.group(2),
            )
            minutes_since_midnight_to = int(entry_match.group(3)) * 60 + int(
                entry_match.group(4),
            )
            if (
                minutes_since_midnight_from
                <= minutes_since_midnight_now
                <= minutes_since_midnight_to
            ):
                return schedule_entry
    return None


def encode_data_set_with_fom(
    ds: orm.DataSet,
    fom: None | IndexingResultSummary,
    beamtime_id: BeamtimeId,
) -> JsonDataSetWithFom:
    return JsonDataSetWithFom(
        data_set=encode_orm_data_set_to_json(ds, beamtime_id),
        fom=encode_indexing_fom_to_json(fom if fom is not None else empty_indexing_fom),
    )


@router.get(
    "/api/runs/{beamtimeId}",
    tags=["runs"],
    response_model_exclude_defaults=True,
)
async def read_runs(
    beamtimeId: BeamtimeId,  # noqa: N803
    session: Annotated[AsyncSession, Depends(get_orm_db)],
    date: None | str = None,
    filter: None | str = None,  # noqa: A002
    runRanges: None | str = None,  # noqa: N803
) -> JsonReadRuns:
    attributi = list(
        (
            await session.scalars(
                select(orm.Attributo)
                .where(orm.Attributo.beamtime_id == beamtimeId)
                .order_by(orm.Attributo.name),
            )
        ).all(),
    )
    chemicals = (
        await session.scalars(
            select(orm.Chemical)
            .where(orm.Chemical.beamtime_id == beamtimeId)
            .options(selectinload(orm.Chemical.files)),
        )
    ).all()
    experiment_types = (
        await session.scalars(
            select(orm.ExperimentType).where(
                orm.ExperimentType.beamtime_id == beamtimeId,
            ),
        )
    ).all()
    run_ids: set[int] = set()
    if runRanges is not None and runRanges.strip():
        for single_range in runRanges.split(","):
            try:
                from_to = single_range.split("-", maxsplit=2)
                from_: int
                to_: int
                if len(from_to) == 1:
                    from_ = to_ = int(from_to[0])
                else:
                    from_ = int(from_to[0])
                    to_ = int(from_to[1])
                for run_id in range(from_, to_ + 1):
                    run_ids.add(run_id)
            except:
                raise HTTPException(
                    status_code=400,
                    detail=f"run range “{range}” is not of the form “from-to” (where from, to are integers)",
                )
    all_runs = (
        await session.scalars(
            select(orm.Run)
            .where(
                (orm.Run.beamtime_id == beamtimeId)
                & (orm.Run.external_id.in_(run_ids) if run_ids else true())
            )
            # Sort by inverse chronological order
            .order_by(orm.Run.external_id.desc())
            .options(selectinload(orm.Run.files)),
        )
    ).all()
    all_events = (
        await session.scalars(
            select(orm.EventLog)
            .where(
                (orm.EventLog.beamtime_id == beamtimeId)
                & (orm.EventLog.level == EventLogLevel.USER),
            )
            .order_by(orm.EventLog.created.desc())
            .options(selectinload(orm.EventLog.files)),
        )
    ).all()

    attributo_name_to_id: dict[str, AttributoId] = {
        a.name: AttributoId(a.id)
        for a in attributi
        if a.associated_table == AssociatedTable.RUN
    }

    try:
        run_filter = compile_run_filter(filter.strip() if filter is not None else "")
        runs = [
            run
            for run in all_runs
            if run_filter(
                FilterInput(
                    run=run,
                    chemical_names={s.name: s.id for s in chemicals},
                    attributo_name_to_id=attributo_name_to_id,
                ),
            )
        ]
    except FilterParseError as e:
        raise Exception(f"error in filter string: {e}")

    date_filter = date.strip() if date is not None else ""
    runs = (
        [run for run in runs if run_has_started_date(run, date_filter)]
        if date_filter
        else runs
    )
    events = (
        [event for event in all_events if event_has_date(event, date_filter)]
        if date_filter
        else all_events
    )

    return JsonReadRuns(
        filter_dates=extract_runs_and_event_dates(all_runs, all_events),
        attributi=[encode_attributo(a) for a in attributi],
        events=[encode_event(e) for e in events],
        chemicals=[encode_chemical(a) for a in chemicals],
        experiment_types=[encode_experiment_type(a) for a in experiment_types],
        runs=[
            JsonRun(
                id=r.id,
                external_id=r.external_id,
                attributi=[encode_run_attributo_value(v) for v in r.attributo_values],
                started=utc_datetime_to_utc_int(r.started),
                started_local=utc_datetime_to_local_int(r.started),
                stopped=(
                    utc_datetime_to_utc_int(r.stopped)
                    if r.stopped is not None
                    else None
                ),
                stopped_local=(
                    utc_datetime_to_local_int(r.stopped)
                    if r.stopped is not None
                    else None
                ),
                files=[
                    JsonRunFile(id=f.id, glob=f.glob, source=f.source) for f in r.files
                ],
                summary=encode_indexing_fom_to_json(empty_indexing_fom),
                experiment_type_id=r.experiment_type_id,
            )
            for r in runs
        ],
    )


@router.get(
    "/api/runs-overview/{beamtimeId}",
    tags=["runs"],
    response_model_exclude_defaults=True,
)
async def read_runs_overview(
    beamtimeId: BeamtimeId,  # noqa: N803
    session: Annotated[AsyncSession, Depends(get_orm_db)],
) -> JsonReadRunsOverview:
    attributi = list(
        (
            await session.scalars(
                select(orm.Attributo)
                .where(orm.Attributo.beamtime_id == beamtimeId)
                .order_by(orm.Attributo.name),
            )
        ).all(),
    )
    chemicals = (
        await session.scalars(
            select(orm.Chemical)
            .where(orm.Chemical.beamtime_id == beamtimeId)
            .options(selectinload(orm.Chemical.files)),
        )
    ).all()
    experiment_types = (
        await session.scalars(
            select(orm.ExperimentType).where(
                orm.ExperimentType.beamtime_id == beamtimeId,
            ),
        )
    ).all()
    data_sets = (
        await session.scalars(
            select(orm.DataSet, orm.ExperimentType)
            .join(orm.DataSet.experiment_type)
            .where(orm.ExperimentType.beamtime_id == beamtimeId),
        )
    ).all()
    latest_run = (
        await session.scalars(
            select(orm.Run)
            .where(orm.Run.beamtime_id == beamtimeId)
            # Sort by inverse chronological order
            .order_by(orm.Run.started.desc())
            .limit(1)
            .options(
                selectinload(orm.Run.indexing_results).selectinload(
                    orm.IndexingResult.indexing_parameters,
                ),
            )
            .options(selectinload(orm.Run.files)),
        )
    ).one_or_none()
    events = (
        await session.scalars(
            select(orm.EventLog)
            .where(
                (orm.EventLog.beamtime_id == beamtimeId)
                & (orm.EventLog.level == EventLogLevel.USER),
            )
            .order_by(orm.EventLog.created.desc())
            .options(selectinload(orm.EventLog.files)),
        )
    ).all()

    attributo_types: dict[AttributoId, AttributoType] = {
        AttributoId(a.id): schema_dict_to_attributo_type(a.json_schema)
        for a in attributi
    }
    run_attributi_map: dict[AttributoId, None | orm.RunHasAttributoValue] = {
        ra.attributo_id: ra
        for ra in (latest_run.attributo_values if latest_run is not None else [])
    }
    data_set_attributi_maps: dict[
        int,
        dict[AttributoId, None | orm.DataSetHasAttributoValue],
    ] = {
        ds.id: {dsa.attributo_id: dsa for dsa in ds.attributo_values}
        for ds in data_sets
    }
    data_set_for_latest_run: None | orm.DataSet = next(
        iter(
            ds
            for ds in data_sets
            if latest_run is not None
            and ds.experiment_type_id == latest_run.experiment_type_id
            and run_matches_dataset(
                attributo_types,
                run_attributi_map,
                data_set_attributi_maps[ds.id],
            )
        ),
        None,
    )

    if latest_run is not None and data_set_for_latest_run is not None:
        # Now we know the run and its data set. Unforunately, we have to
        # now query _all_ runs, so we can show full-dataset statistics.
        other_runs = (
            await session.scalars(
                select(orm.Run)
                .where(
                    (orm.Run.beamtime_id == beamtimeId)
                    & (orm.Run.experiment_type_id == latest_run.experiment_type_id),
                )
                .options(
                    selectinload(orm.Run.indexing_results).selectinload(
                        orm.IndexingResult.indexing_parameters
                    )
                ),
            )
        ).all()

        other_runs_in_ds = [
            r
            for r in other_runs
            if run_matches_dataset(
                attributo_types,
                {ra.attributo_id: ra for ra in r.attributo_values},
                data_set_attributi_maps[data_set_for_latest_run.id],
            )
        ]

        foms_in_this_ds: list[IndexingResultSummary] = []
        for r in other_runs_in_ds:
            try:
                max_ir = max(
                    (
                        ir
                        for ir in r.indexing_results
                        if ir.indexing_parameters.is_online
                    ),
                    key=lambda ir: ir.indexed_frames,
                )
                foms_in_this_ds.append(fom_for_indexing_result(max_ir))
            except:  # noqa: S110
                # No indexing results in this run. Fine.
                pass

    user_configuration = await retrieve_latest_config(session, beamtimeId)
    live_stream_file = (
        await session.scalars(
            select(orm.File).where(
                orm.File.file_name == live_stream_image_name(beamtimeId),
            ),
        )
    ).one_or_none()
    latest_indexing_results = [
        o
        for o in (latest_run.indexing_results if latest_run is not None else [])
        if o.indexing_parameters.is_online
    ]
    if latest_run is not None and latest_indexing_results:
        latest_indexing_result_orm = latest_indexing_results[0]
        latest_statistics_orm = (
            await latest_indexing_result_orm.awaitable_attrs.statistics
        )
        target_frames_count_attributi = [
            a.id for a in attributi if a.name == "target_frame_count"
        ]
        total_frames_attributo = (
            target_frames_count_attributi[0] if target_frames_count_attributi else None
        )
        latest_indexing_result = JsonRunAnalysisIndexingResult(
            indexing_result_id=latest_indexing_result_orm.id,
            run_id=latest_run.id,
            foms=encode_indexing_fom_to_json(
                fom_for_indexing_result(latest_indexing_result_orm),
            ),
            frames=latest_indexing_result_orm.frames,
            total_frames=next(
                iter(
                    a.integer_value
                    for a in latest_run.attributo_values
                    if a.attributo_id == total_frames_attributo
                ),
                None,
            ),
            running=latest_indexing_result_orm.job_status
            in (DBJobStatus.RUNNING, DBJobStatus.QUEUED),
            indexing_statistics=[
                JsonIndexingStatistic(
                    time=(stat.time - latest_run.started).seconds,
                    frames=stat.frames,
                    hits=stat.hits,
                    indexed=stat.indexed_frames,
                    crystals=stat.indexed_crystals,
                )
                for stat in latest_statistics_orm
            ],
        )
    else:
        latest_indexing_result = None
    found_schedule_entry = await _find_schedule_entry(session, beamtimeId)
    r = latest_run
    this_run_fom = (
        fom_for_indexing_result(latest_indexing_results[0])
        if latest_indexing_results
        else empty_indexing_fom
    )
    latest_run_json = (
        JsonRun(
            id=r.id,
            external_id=r.external_id,
            attributi=[encode_run_attributo_value(v) for v in r.attributo_values],
            started=utc_datetime_to_utc_int(r.started),
            started_local=utc_datetime_to_local_int(r.started),
            stopped=(
                utc_datetime_to_utc_int(r.stopped) if r.stopped is not None else None
            ),
            stopped_local=(
                utc_datetime_to_local_int(r.stopped) if r.stopped is not None else None
            ),
            files=[JsonRunFile(id=f.id, glob=f.glob, source=f.source) for f in r.files],
            summary=encode_indexing_fom_to_json(this_run_fom),
            experiment_type_id=r.experiment_type_id,
        )
        if r is not None
        else None
    )
    return JsonReadRunsOverview(
        current_beamtime_user=(
            None if found_schedule_entry is None else found_schedule_entry.users
        ),
        latest_run=latest_run_json,
        latest_indexing_result=latest_indexing_result,
        live_stream=(
            None
            if live_stream_file is None
            else JsonLiveStream(
                file_id=live_stream_file.id,
                modified=utc_datetime_to_utc_int(live_stream_file.modified),
                modified_local=utc_datetime_to_local_int(live_stream_file.modified),
            )
        ),
        attributi=[encode_attributo(a) for a in attributi],
        events=[encode_event(e) for e in events],
        chemicals=[encode_chemical(a) for a in chemicals],
        user_config=encode_user_configuration(user_configuration),
        experiment_types=[encode_experiment_type(a) for a in experiment_types],
        foms_for_this_data_set=(
            encode_data_set_with_fom(
                data_set_for_latest_run,
                summary_from_foms(foms_in_this_ds),
                beamtimeId,
            )
            if data_set_for_latest_run
            else None
        ),
    )


@router.get(
    "/api/run-bulk-import-template/{beamtimeId}.xlsx",
    tags=["runs"],
    include_in_schema=False,
    response_model_exclude_defaults=True,
)
async def bulk_import_spreadsheet_template(
    beamtimeId: BeamtimeId,  # noqa: N803
    session: Annotated[AsyncSession, Depends(get_orm_db)],
) -> Response:
    workbook_output = Workbook()
    ws = workbook_output.active

    assert ws is not None

    ws["A1"] = "run id"
    ws["B1"] = "experiment type"
    ws["C1"] = "started"
    ws["D1"] = "stopped"
    ws["E1"] = "files"

    col_idx = 6
    for attributo in await session.scalars(
        select(orm.Attributo).where(
            (orm.Attributo.beamtime_id == beamtimeId)
            & (orm.Attributo.associated_table == AssociatedTable.RUN)
        )
    ):
        ws.cell(column=col_idx, row=1, value=attributo.name)
        col_idx += 1

    workbook = workbook_output
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)

    def iterworkbook() -> Generator[bytes, None, None]:
        yield from workbook_bytes

    return StreamingResponse(
        iterworkbook(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{beamtimeId}-import.xlsx"'
        },
    )


@router.get(
    "/api/run-bulk-import/{beamtimeId}",
    tags=["runs"],
    response_model_exclude_defaults=True,
)
async def bulk_import_info(
    beamtimeId: BeamtimeId,  # noqa: N803
    session: Annotated[AsyncSession, Depends(get_orm_db)],
) -> JsonRunsBulkImportInfo:
    return JsonRunsBulkImportInfo(
        run_attributi=[
            encode_attributo(a)
            for a in await session.scalars(
                select(orm.Attributo).where(
                    (orm.Attributo.beamtime_id == beamtimeId)
                    & (orm.Attributo.associated_table == AssociatedTable.RUN)
                )
            )
        ],
        experiment_types=[
            et.name
            for et in await session.scalars(
                select(orm.ExperimentType).where(
                    orm.ExperimentType.beamtime_id == beamtimeId
                )
            )
        ],
        chemicals=[
            encode_chemical(c)
            for c in await session.scalars(
                select(orm.Chemical)
                .where(orm.Chemical.beamtime_id == beamtimeId)
                .options(selectinload(orm.Chemical.files))
            )
        ],
    )


@router.post(
    "/api/run-bulk-import/{beamtimeId}",
    tags=["runs"],
    response_model_exclude_defaults=True,
)
async def bulk_import(
    beamtimeId: BeamtimeId,  # noqa: N803
    simulate: bool,
    create_data_sets: bool,
    file: UploadFile,
    session: Annotated[AsyncSession, Depends(get_orm_db)],
) -> JsonRunsBulkImportOutput:
    async with session.begin():
        logger.info("starting bulk import")
        wb = load_workbook(filename=BytesIO(file.file.read()), data_only=True)

        parsed_wb = parse_run_spreadsheet_workbook(wb)

        if isinstance(parsed_wb, ConversionError):
            logger.error(
                f"there were errors parsing spreadsheet: {parsed_wb.joined_messages()}"
            )
            return JsonRunsBulkImportOutput(
                errors=parsed_wb.error_messages,
                warnings=[],
                number_of_runs=0,
                data_sets=[],
                simulated=simulate,
                create_data_sets=create_data_sets,
            )

        attributi = list(
            (
                await session.scalars(
                    select(orm.Attributo).where(
                        (orm.Attributo.beamtime_id == beamtimeId)
                        & (orm.Attributo.associated_table == AssociatedTable.RUN)
                    )
                )
            ).all()
        )

        chemicals = list(
            (
                await session.scalars(
                    select(orm.Chemical).where(orm.Chemical.beamtime_id == beamtimeId)
                )
            ).all()
        )

        existing_runs = list(
            (
                await session.scalars(
                    select(orm.Run).where(orm.Run.beamtime_id == beamtimeId)
                )
            ).all()
        )

        ets = list(
            (
                await session.scalars(
                    select(orm.ExperimentType)
                    .where(orm.ExperimentType.beamtime_id == beamtimeId)
                    .options(
                        selectinload(orm.ExperimentType.attributi).selectinload(
                            orm.ExperimentHasAttributo.attributo
                        )
                    )
                )
            ).all()
        )

        created_runs = create_runs_from_spreadsheet(
            spreadsheet=parsed_wb,
            beamtime_id=beamtimeId,
            attributi=attributi,
            chemicals=chemicals,
            experiment_types=ets,
            existing_runs=existing_runs,
        )

        if isinstance(created_runs, SpreadsheetValidationErrors):
            logger.error(
                f"there were errors validating spreadsheet: {created_runs.errors}"
            )
            return JsonRunsBulkImportOutput(
                errors=created_runs.errors,
                warnings=[],
                number_of_runs=0,
                data_sets=[],
                simulated=simulate,
                create_data_sets=create_data_sets,
            )

        for run in created_runs.runs:
            session.add(run)

        existing_data_sets: list[orm.DataSet] = list(
            (
                await session.scalars(
                    select(orm.DataSet, orm.ExperimentType)
                    .join(orm.DataSet.experiment_type)
                    .where(orm.ExperimentType.beamtime_id == beamtimeId),
                )
            ).all()
        )
        data_sets: list[orm.DataSet]
        if create_data_sets:
            data_set_creation_result = create_data_set_for_runs(
                ets, created_runs.runs, existing_data_sets
            )
            if isinstance(data_set_creation_result, ConversionError):
                await session.rollback()
                return JsonRunsBulkImportOutput(
                    errors=data_set_creation_result.error_messages,
                    warnings=created_runs.warnings,
                    number_of_runs=len(created_runs.runs),
                    data_sets=[],
                    simulated=simulate,
                    create_data_sets=create_data_sets,
                )
            data_sets = data_set_creation_result
            for ds in data_sets:
                session.add(ds)
        else:
            data_sets = []

        await session.flush()

        if simulate:
            await session.rollback()
        else:
            await session.commit()

        return JsonRunsBulkImportOutput(
            errors=[],
            warnings=created_runs.warnings,
            number_of_runs=len(created_runs.runs),
            data_sets=[
                JsonDataSet(
                    id=ds.id,
                    experiment_type_id=ds.experiment_type_id,
                    attributi=[
                        encode_data_set_attributo_value(dsa)
                        for dsa in ds.attributo_values
                    ],
                    beamtime_id=beamtimeId,
                )
                for ds in data_sets
            ],
            simulated=simulate,
            create_data_sets=create_data_sets,
        )


@router.delete(
    "/api/runs/{beamtimeId}/{runId}",
    tags=["runs"],
    response_model_exclude_defaults=True,
)
async def delete_run(
    beamtimeId: BeamtimeId,  # noqa: N803
    runId: int,  # noqa: N803
    session: Annotated[AsyncSession, Depends(get_orm_db)],
) -> JsonDeleteRunOutput:
    async with session.begin():
        await session.execute(
            delete(orm.Run).where(
                (orm.Run.beamtime_id == beamtimeId) & (orm.Run.external_id == runId)
            )
        )
        await session.commit()
    return JsonDeleteRunOutput(result=True)

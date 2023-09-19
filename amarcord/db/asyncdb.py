import datetime
import hashlib
import io
import itertools
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from typing import Final
from typing import Iterable
from typing import cast

import numpy.typing as npt
import sqlalchemy as sa
import structlog.stdlib
from openpyxl import Workbook
from PIL import Image
from pint import UnitRegistry

from amarcord import magic
from amarcord.amici.crystfel.util import CrystFELCellFile
from amarcord.amici.crystfel.util import coparse_cell_description
from amarcord.amici.crystfel.util import parse_cell_description
from amarcord.db.associated_table import AssociatedTable
from amarcord.db.async_dbcontext import AsyncDBContext
from amarcord.db.async_dbcontext import Connection
from amarcord.db.attributi import ATTRIBUTO_STARTED
from amarcord.db.attributi import ATTRIBUTO_STOPPED
from amarcord.db.attributi import AttributoConversionFlags
from amarcord.db.attributi import attributo_sort_key
from amarcord.db.attributi import attributo_type_to_schema
from amarcord.db.attributi import attributo_type_to_string
from amarcord.db.attributi import schema_json_to_attributo_type
from amarcord.db.attributi_map import AttributiMap
from amarcord.db.attributo_id import AttributoId
from amarcord.db.attributo_name_and_role import AttributoNameAndRole
from amarcord.db.attributo_type import AttributoType
from amarcord.db.attributo_type import AttributoTypeChemical
from amarcord.db.attributo_type import AttributoTypeDecimal
from amarcord.db.attributo_value import AttributoValue
from amarcord.db.chemical_type import ChemicalType
from amarcord.db.data_set import DBDataSet
from amarcord.db.db_job_status import DBJobStatus
from amarcord.db.db_merge_result import DBMergeResultInput
from amarcord.db.db_merge_result import DBMergeResultOutput
from amarcord.db.db_merge_result import DBMergeRuntimeStatus
from amarcord.db.db_merge_result import DBMergeRuntimeStatusDone
from amarcord.db.db_merge_result import DBMergeRuntimeStatusError
from amarcord.db.db_merge_result import DBMergeRuntimeStatusRunning
from amarcord.db.dbattributo import DBAttributo
from amarcord.db.event_log_level import EventLogLevel
from amarcord.db.experiment_type import DBExperimentType
from amarcord.db.indexing_result import DBIndexingFOM
from amarcord.db.indexing_result import DBIndexingResultDone
from amarcord.db.indexing_result import DBIndexingResultInput
from amarcord.db.indexing_result import DBIndexingResultOutput
from amarcord.db.indexing_result import DBIndexingResultRunning
from amarcord.db.indexing_result import DBIndexingResultRuntimeStatus
from amarcord.db.indexing_result import DBIndexingResultStatistic
from amarcord.db.merge_model import MergeModel
from amarcord.db.merge_parameters import DBMergeParameters
from amarcord.db.merge_result import MergeResult
from amarcord.db.merge_result import MergeResultFom
from amarcord.db.merge_result import MergeResultOuterShell
from amarcord.db.merge_result import MergeResultShell
from amarcord.db.merge_result import RefinementResult
from amarcord.db.migrations.alembic_utilities import upgrade_to_head_connection
from amarcord.db.polarisation import Polarisation
from amarcord.db.refinement_result import DBRefinementResultOutput
from amarcord.db.scale_intensities import ScaleIntensities
from amarcord.db.schedule_entry import BeamtimeScheduleEntry
from amarcord.db.table_classes import DBChemical
from amarcord.db.table_classes import DBEvent
from amarcord.db.table_classes import DBFile
from amarcord.db.table_classes import DBRun
from amarcord.db.tables import DBTables
from amarcord.db.user_configuration import UserConfiguration
from amarcord.db.user_configuration import initial_user_configuration
from amarcord.json_schema import coparse_schema_type
from amarcord.pint_util import valid_pint_unit
from amarcord.util import datetime_to_local
from amarcord.util import group_by
from amarcord.util import sha256_file

_UNIT_REGISTRY = UnitRegistry()
LIVE_STREAM_IMAGE: Final = "live-stream-image"

logger = structlog.stdlib.get_logger(__name__)


class MergeResultError(str):
    pass


@dataclass(frozen=True)
class CreateFileResult:
    id: int
    type_: str
    size_in_bytes: int


ATTRIBUTO_GROUP_MANUAL = "manual"


def _evaluate_indexing_result_runtime_status(
    db_row: Any,
) -> DBIndexingResultRuntimeStatus:
    return (
        DBIndexingResultDone(
            stream_file=Path(db_row["stream_file"]),
            job_error=db_row["job_error"],
            fom=DBIndexingFOM(
                db_row["hit_rate"],
                db_row["indexing_rate"],
                db_row["indexed_frames"],
                db_row["detector_shift_x_mm"],
                db_row["detector_shift_y_mm"],
            ),
        )
        if db_row["job_status"] == DBJobStatus.DONE
        else DBIndexingResultRunning(
            stream_file=Path(db_row["stream_file"]),
            job_id=db_row["job_id"],
            fom=DBIndexingFOM(
                db_row["hit_rate"],
                db_row["indexing_rate"],
                db_row["indexed_frames"],
                db_row["detector_shift_x_mm"],
                db_row["detector_shift_y_mm"],
            ),
        )
        if db_row["job_status"] == DBJobStatus.RUNNING
        else None
    )


def _values_from_shell_foms(
    mrd: MergeResultShell, merge_result_id: int
) -> dict[str, float | int]:
    return {
        "merge_result_id": merge_result_id,
        "one_over_d_centre": mrd.one_over_d_centre,
        "nref": mrd.nref,
        "d_over_a": mrd.d_over_a,
        "min_res": mrd.min_res,
        "max_res": mrd.max_res,
        "cc": mrd.cc,
        "ccstar": mrd.ccstar,
        "r_split": mrd.r_split,
        "reflections_possible": mrd.reflections_possible,
        "completeness": mrd.completeness,
        "measurements": mrd.measurements,
        "redundancy": mrd.redundancy,
        "snr": mrd.snr,
        "mean_i": mrd.mean_i,
    }


def _runtime_status_sql_values(rs: DBMergeRuntimeStatus) -> dict[str, Any]:
    match rs:
        case None:
            return {"recent_log": "", "job_status": DBJobStatus.QUEUED}
        case DBMergeRuntimeStatusRunning(job_id, started, recent_log):
            return {
                "started": started,
                "job_id": job_id,
                "recent_log": recent_log,
                "job_status": DBJobStatus.RUNNING,
            }
        case DBMergeRuntimeStatusError(error, started, stopped, recent_log):
            return {
                "job_error": error,
                "started": started,
                "stopped": stopped,
                "recent_log": recent_log,
                "job_status": DBJobStatus.DONE,
            }
        case DBMergeRuntimeStatusDone(started, stopped, result, recent_log):
            return {
                "job_status": DBJobStatus.DONE,
                "recent_log": recent_log,
                "started": started,
                "stopped": stopped,
                "mtz_file_id": result.mtz_file_id,
                "fom_snr": result.fom.snr,
                "fom_wilson": result.fom.wilson,
                "fom_ln_k": result.fom.ln_k,
                "fom_discarded_reflections": result.fom.discarded_reflections,
                "fom_one_over_d_from": result.fom.one_over_d_from,
                "fom_one_over_d_to": result.fom.one_over_d_to,
                "fom_redundancy": result.fom.redundancy,
                "fom_completeness": result.fom.completeness,
                "fom_measurements_total": result.fom.measurements_total,
                "fom_reflections_total": result.fom.reflections_total,
                "fom_reflections_possible": result.fom.reflections_possible,
                "fom_r_split": result.fom.r_split,
                "fom_r1i": result.fom.r1i,
                "fom_2": result.fom.r2,
                "fom_cc": result.fom.cc,
                "fom_ccstar": result.fom.ccstar,
                "fom_ccano": result.fom.ccano,
                "fom_crdano": result.fom.crdano,
                "fom_rano": result.fom.rano,
                "fom_rano_over_r_split": result.fom.rano_over_r_split,
                "fom_d1sig": result.fom.d1sig,
                "fom_d2sig": result.fom.d2sig,
                "fom_outer_resolution": result.fom.outer_shell.resolution,
                "fom_outer_ccstar": result.fom.outer_shell.ccstar,
                "fom_outer_r_split": result.fom.outer_shell.r_split,
                "fom_outer_cc": result.fom.outer_shell.cc,
                "fom_outer_unique_reflections": result.fom.outer_shell.unique_reflections,
                "fom_outer_completeness": result.fom.outer_shell.completeness,
                "fom_outer_redundancy": result.fom.outer_shell.redundancy,
                "fom_outer_snr": result.fom.outer_shell.snr,
                "fom_outer_min_res": result.fom.outer_shell.min_res,
                "fom_outer_max_res": result.fom.outer_shell.max_res,
            }


class AsyncDB:
    def __init__(self, dbcontext: AsyncDBContext, tables: DBTables) -> None:
        self.dbcontext = dbcontext
        self.tables = tables

    async def migrate(self) -> None:
        async with self.begin() as conn:
            await conn.run_sync(upgrade_to_head_connection)

    def read_only_connection(self) -> Connection:
        return self.dbcontext.read_only_connection()

    def begin(self) -> Connection:
        return self.dbcontext.begin()

    async def dispose(self) -> None:
        await self.dbcontext.dispose()

    async def replace_beamtime_schedule(
        self, conn: Connection, schedule: list[BeamtimeScheduleEntry]
    ) -> None:
        await conn.execute(self.tables.beamtime_schedule.delete())
        await conn.execute(self.tables.beamtime_schedule_has_chemical.delete())
        for entry in schedule:
            chemical_ids = entry.chemicals
            inserted_schedule_id = (
                await conn.execute(
                    self.tables.beamtime_schedule.insert().values(
                        users=entry.users,
                        date=entry.date,
                        shift=entry.shift,
                        td_support=entry.td_support,
                        comment=entry.comment,
                    )
                )
            ).inserted_primary_key[0]

            for chemical_id in chemical_ids:
                await conn.execute(
                    self.tables.beamtime_schedule_has_chemical.insert().values(
                        beamtime_schedule_id=inserted_schedule_id,
                        chemical_id=chemical_id,
                    )
                )

    async def retrieve_beamtime_schedule(
        self, conn: Connection
    ) -> list[BeamtimeScheduleEntry]:
        schedule_chemicals_select = await conn.execute(
            sa.select(
                [
                    self.tables.beamtime_schedule_has_chemical.c.beamtime_schedule_id,
                    self.tables.beamtime_schedule_has_chemical.c.chemical_id,
                ]
            )
        )

        schedule_chemicals_fetched: list[
            tuple[int, int]
        ] = schedule_chemicals_select.fetchall()
        schedule_chemicals_by_chemical: dict[int, list[int]] = {
            k: [*map(lambda v: v[1], values)]
            for k, values in itertools.groupby(
                sorted(schedule_chemicals_fetched, key=lambda x: x[0]),
                lambda x: x[0],
            )
        }

        schedule = await conn.execute(self.tables.beamtime_schedule.select())
        return [
            BeamtimeScheduleEntry(
                users=se.users,
                comment=se.comment,
                date=se.date,
                shift=se.shift,
                td_support=se.td_support,
                chemicals=schedule_chemicals_by_chemical.get(se.id, []),
            )
            for se in schedule.fetchall()
        ]

    async def retrieve_file_id_by_name(
        self, conn: Connection, file_name: str
    ) -> int | None:
        result = await conn.execute(
            sa.select(
                [
                    self.tables.file.c.id,
                ]
            ).where(self.tables.file.c.file_name == file_name)
        )

        db_result = result.fetchone()

        return db_result[0] if db_result else None

    async def retrieve_file(
        self, conn: Connection, file_id: int, with_contents: bool
    ) -> DBFile:
        result = await conn.execute(
            sa.select(
                [
                    self.tables.file.c.file_name,
                    self.tables.file.c.description,
                    self.tables.file.c.original_path,
                    self.tables.file.c.type,
                    self.tables.file.c.size_in_bytes,
                ]
                + ([self.tables.file.c.contents] if with_contents else [])
            ).where(self.tables.file.c.id == file_id)
        )

        result_row = result.fetchone()
        return DBFile(
            id=file_id,
            description=result_row["description"],
            type_=result_row["type"],
            original_path=result_row["original_path"],
            file_name=result_row["file_name"],
            size_in_bytes=result_row["size_in_bytes"],
            contents=result_row["contents"] if with_contents else None,
        )

    async def retrieve_configuration(self, conn: Connection) -> UserConfiguration:
        result = (
            await conn.execute(
                sa.select(
                    [
                        self.tables.configuration.c.auto_pilot,
                        self.tables.configuration.c.use_online_crystfel,
                        self.tables.configuration.c.current_experiment_type_id,
                    ]
                ).order_by(self.tables.configuration.c.id.desc())
            )
        ).fetchone()

        return (
            UserConfiguration(
                auto_pilot=result[0],
                use_online_crystfel=result[1] if result[1] is not None else False,
                current_experiment_type_id=result[2],
            )
            if result is not None
            else initial_user_configuration()
        )

    async def update_configuration(
        self, conn: Connection, configuration: UserConfiguration
    ) -> None:
        await conn.execute(
            self.tables.configuration.insert().values(
                auto_pilot=configuration.auto_pilot,
                use_online_crystfel=configuration.use_online_crystfel,
                current_experiment_type_id=configuration.current_experiment_type_id,
                created=datetime.datetime.utcnow(),
            )
        )

    async def retrieve_attributi(
        self, conn: Connection, associated_table: AssociatedTable | None
    ) -> list[DBAttributo]:
        ac = self.tables.attributo.c
        select_stmt = sa.select(
            [
                ac.name,
                ac.description,
                ac.group,
                ac.json_schema,
                ac.associated_table,
            ]
        ).order_by(ac.associated_table)

        if associated_table is not None:
            select_stmt = select_stmt.where(ac.associated_table == associated_table)

        result = await conn.execute(select_stmt)
        return [
            DBAttributo(
                name=AttributoId(a["name"]),
                description=a["description"],
                group=a["group"],
                associated_table=a["associated_table"],
                attributo_type=schema_json_to_attributo_type(a["json_schema"]),
            )
            for a in result
        ]

    async def retrieve_refinement_results(
        self, conn: Connection
    ) -> list[DBRefinementResultOutput]:
        rr = self.tables.refinement_result.c
        return [
            DBRefinementResultOutput(
                id=row["id"],
                merge_result_id=row["merge_result_id"],
                pdb_file_id=row["pdb_file_id"],
                mtz_file_id=row["mtz_file_id"],
                r_free=row["r_free"],
                r_work=row["r_work"],
                rms_bond_angle=row["rms_bond_angle"],
                rms_bond_length=row["rms_bond_length"],
            )
            for row in await conn.execute(
                sa.select(
                    [
                        rr.id,
                        rr.merge_result_id,
                        rr.pdb_file_id,
                        rr.mtz_file_id,
                        rr.r_free,
                        rr.r_work,
                        rr.rms_bond_angle,
                        rr.rms_bond_length,
                    ]
                )
            )
        ]

    async def create_refinement_result(
        self,
        conn: Connection,
        merge_result_id: int,
        pdb_file_id: int,
        mtz_file_id: int,
        rfree: float,
        rwork: float,
        rms_bond_angle: float,
        rms_bond_length: float,
    ) -> int:
        result: int = (
            await conn.execute(
                self.tables.refinement_result.insert().values(
                    merge_result_id=merge_result_id,
                    pdb_file_id=pdb_file_id,
                    mtz_file_id=mtz_file_id,
                    r_free=rfree,
                    r_work=rwork,
                    rms_bond_angle=rms_bond_angle,
                    rms_bond_length=rms_bond_length,
                )
            )
        ).inserted_primary_key[0]
        return result

    async def create_chemical(
        self,
        conn: Connection,
        name: str,
        responsible_person: str,
        type_: ChemicalType,
        attributi: AttributiMap,
    ) -> int:
        result: int = (
            await conn.execute(
                self.tables.chemical.insert().values(
                    name=name,
                    modified=datetime.datetime.utcnow(),
                    attributi=attributi.to_json(),
                    type=type_,
                    responsible_person=responsible_person,
                )
            )
        ).inserted_primary_key[0]
        return result

    async def update_chemical(
        self,
        conn: Connection,
        id_: int,
        name: str,
        responsible_person: str,
        type_: ChemicalType,
        attributi: AttributiMap,
    ) -> None:
        await conn.execute(
            sa.update(self.tables.chemical)
            .values(
                name=name,
                responsible_person=responsible_person,
                modified=datetime.datetime.utcnow(),
                attributi=attributi.to_json(),
                type=type_,
            )
            .where(self.tables.chemical.c.id == id_)
        )

    async def _retrieve_files(
        self,
        conn: Connection,
        association_column: sa.Column[Any],
        where_clause: Any | None = None,
    ) -> dict[int, list[DBFile]]:
        select_stmt = (
            sa.select(
                [
                    association_column,
                    self.tables.file.c.id,
                    self.tables.file.c.description,
                    self.tables.file.c.file_name,
                    self.tables.file.c.type,
                    self.tables.file.c.size_in_bytes,
                    self.tables.file.c.original_path,
                ]
            )
            .select_from(
                association_column.table.join(
                    self.tables.file,
                    association_column.table.c.file_id == self.tables.file.c.id,
                )
            )
            .order_by(association_column)
        )
        if where_clause is not None:
            select_stmt = select_stmt.where(where_clause)
        file_results = (await conn.execute(select_stmt)).fetchall()

        result: dict[int, list[DBFile]] = {
            key: [
                DBFile(
                    id=row["id"],
                    description=row["description"],
                    type_=row["type"],
                    file_name=row["file_name"],
                    size_in_bytes=row["size_in_bytes"],
                    original_path=row["original_path"],
                    contents=None,
                )
                for row in group
            ]
            for key, group in itertools.groupby(
                file_results, key=lambda row: row[association_column.name]  # type: ignore
            )
        }
        return result

    async def retrieve_chemicals(
        self, conn: Connection, attributi: list[DBAttributo]
    ) -> list[DBChemical]:
        chemical_to_files = await self._retrieve_files(
            conn, self.tables.chemical_has_file.c.chemical_id
        )

        select_stmt = sa.select(
            [
                self.tables.chemical.c.id,
                self.tables.chemical.c.name,
                self.tables.chemical.c.responsible_person,
                self.tables.chemical.c.type,
                self.tables.chemical.c.attributi,
            ]
        ).order_by(self.tables.chemical.c.name)

        result = await conn.execute(select_stmt)

        return [
            DBChemical(
                id=a["id"],
                name=a["name"],
                responsible_person=a["responsible_person"],
                type_=a["type"],
                attributi=AttributiMap.from_types_and_json(
                    types=attributi,
                    # chemical IDs not needed since chemicals cannot refer to themselves (yet!)
                    chemical_ids=[],
                    raw_attributi=a["attributi"],
                ),
                files=chemical_to_files.get(a["id"], []),
            )
            for a in result
        ]

    async def delete_event(
        self,
        conn: Connection,
        id_: int,
    ) -> None:
        await conn.execute(
            sa.delete(self.tables.event_log).where(self.tables.event_log.c.id == id_)
        )

    async def delete_file(
        self,
        conn: Connection,
        id_: int,
    ) -> None:
        await conn.execute(
            sa.delete(self.tables.file).where(self.tables.file.c.id == id_)
        )

    async def retrieve_events(self, conn: Connection) -> list[DBEvent]:
        ec = self.tables.event_log.c
        event_to_files = await self._retrieve_files(
            conn, self.tables.event_has_file.c.event_id
        )
        return [
            DBEvent(
                row["id"],
                row["created"],
                row["level"],
                row["source"],
                row["text"],
                files=event_to_files.get(row["id"], []),
            )
            for row in await conn.execute(
                sa.select([ec.id, ec.created, ec.level, ec.source, ec.text]).order_by(
                    ec.created.desc()
                )
            )
        ]

    async def create_event(
        self, conn: Connection, level: EventLogLevel, source: str, text: str
    ) -> int:
        result: int = (
            await conn.execute(
                self.tables.event_log.insert().values(
                    created=datetime.datetime.utcnow(),
                    level=level,
                    source=source,
                    text=text,
                )
            )
        ).inserted_primary_key[0]
        return result

    async def retrieve_runs(
        self, conn: Connection, attributi: list[DBAttributo]
    ) -> list[DBRun]:
        run_to_files = await self._retrieve_files(
            conn, self.tables.run_has_file.c.run_id
        )

        select_stmt = sa.select(
            [
                self.tables.run.c.id,
                self.tables.run.c.experiment_type_id,
                self.tables.run.c.attributi,
            ]
        ).order_by(self.tables.run.c.id.desc())

        result = await conn.execute(select_stmt)

        chemical_ids = await self.retrieve_chemical_ids(conn)

        return [
            DBRun(
                id=a["id"],
                experiment_type_id=a["experiment_type_id"],
                attributi=AttributiMap.from_types_and_json(
                    types=attributi,
                    chemical_ids=chemical_ids,
                    raw_attributi=a["attributi"],
                ),
                files=run_to_files.get(a["id"], []),
            )
            for a in result
        ]

    async def delete_chemical(
        self, conn: Connection, id_: int, delete_in_dependencies: bool
    ) -> None:
        attributi = await self.retrieve_attributi(conn, AssociatedTable.RUN)

        chemical_ids = await self.retrieve_chemical_ids(conn)

        # check if we have a chemical attributo in runs and then do integrity checking
        chemical_attributi = [
            x for x in attributi if isinstance(x.attributo_type, AttributoTypeChemical)
        ]

        if chemical_attributi:
            for r in await self.retrieve_runs(conn, attributi):
                run_attributi = r.attributi
                changed = False
                for chemical_attributo in chemical_attributi:
                    run_chemical = r.attributi.select_chemical_id(
                        chemical_attributo.name
                    )
                    if run_chemical == id_:
                        if delete_in_dependencies:
                            run_attributi.remove_with_type(chemical_attributo.name)
                            changed = True
                        else:
                            raise Exception(f"run {r.id} still has chemical {id_}")
                if changed:
                    await self.update_run_attributi(conn, r.id, run_attributi)

            for ds in await self.retrieve_data_sets(
                conn, chemical_ids=chemical_ids, attributi=attributi
            ):
                changed = False
                for chemical_attributo in chemical_attributi:
                    ds_chemical = ds.attributi.select_chemical_id(
                        chemical_attributo.name
                    )
                    if ds_chemical == id_:
                        if delete_in_dependencies:
                            ds.attributi.remove_with_type(chemical_attributo.name)
                            changed = True
                        else:
                            raise Exception(
                                f"data set {ds.id} still has chemical {id_}"
                            )
                if changed:
                    await self.delete_data_set(conn, ds.id)

        await conn.execute(
            sa.delete(self.tables.chemical).where(self.tables.chemical.c.id == id_)
        )

    async def create_attributo(
        self,
        conn: Connection,
        name: str,
        description: str,
        group: str,
        associated_table: AssociatedTable,
        type_: AttributoType,
    ) -> None:
        # I honestly don't remember why I added this
        # if not re.fullmatch(ATTRIBUTO_NAME_REGEX, name, re.IGNORECASE):
        #     raise ValueError(
        #         f'attributo name "{name}" contains invalid characters (maybe a number at the beginning or a dash?)"
        #     )
        if associated_table == AssociatedTable.CHEMICAL and isinstance(
            type_, AttributoTypeChemical
        ):
            raise ValueError(f"{name}: chemicals can't have attributi of type chemical")
        if isinstance(type_, AttributoTypeDecimal) and type_.standard_unit:
            if type_.suffix is None:
                raise ValueError(f"{name}: got a standard unit, but no suffix")
            if not valid_pint_unit(type_.suffix):
                raise ValueError(f"{name}: unit {type_.suffix} not a valid unit")
        await conn.execute(
            self.tables.attributo.insert().values(
                name=name,
                description=description,
                associated_table=associated_table,
                group=group,
                json_schema=coparse_schema_type(attributo_type_to_schema(type_)),
            )
        )

    async def delete_attributo(
        self,
        conn: Connection,
        name: AttributoId,
    ) -> None:
        attributi = await self.retrieve_attributi(conn, associated_table=None)

        found_attributo = next((x for x in attributi if x.name == name), None)
        if found_attributo is None:
            raise Exception(f'couldn\'t find attributo "{name}"')

        await conn.execute(
            sa.delete(self.tables.attributo).where(self.tables.attributo.c.name == name)
        )

        if found_attributo.associated_table == AssociatedTable.CHEMICAL:
            # This is the tricky bit: we need to retrieve the chemicals with the old attributi list. The chemicals haven't
            # been converted to the new format, so using the new attributi list would make that fail validation.
            for s in await self.retrieve_chemicals(conn, attributi):
                # Then remove the attributo from the chemical and the accompanying types, and update.
                s.attributi.remove_with_type(name)
                await self.update_chemical(
                    conn=conn,
                    id_=s.id,
                    name=s.name,
                    responsible_person=s.responsible_person,
                    type_=s.type_,
                    attributi=s.attributi,
                )
        elif found_attributo.associated_table == AssociatedTable.RUN:
            # Explanation, see above for chemicals
            for r in await self.retrieve_runs(conn, attributi):
                r.attributi.remove_with_type(name)
                await self.update_run_attributi(conn, r.id, r.attributi)
        else:
            raise Exception(
                f"unimplemented: is there a new associated table {found_attributo.associated_table}?"
            )

    async def update_file(
        self, conn: Connection, id_: int, contents_location: Path
    ) -> None:
        mime = magic.from_file(str(contents_location), mime=True)  # type: ignore

        sha256 = sha256_file(contents_location)

        size_in_bytes = contents_location.stat().st_size
        with contents_location.open("rb") as f:
            await conn.execute(
                sa.update(self.tables.file)
                .values(
                    modified=datetime.datetime.utcnow(),
                    contents=f.read(),
                    type=mime,
                    sha256=sha256,
                    size_in_bytes=size_in_bytes,
                )
                .where(self.tables.file.c.id == id_)
            )

    async def update_image_from_nparray(
        self,
        conn: Connection,
        id_: int,
        file_name: str,
        description: str,
        contents: npt.ArrayLike,  # pyright: ignore [reportUnknownParameterType]
        format_: str,
    ) -> None:
        img_byte_arr_io = io.BytesIO()
        Image.fromarray(contents).save(img_byte_arr_io, format=format_)
        img_byte_arr = img_byte_arr_io.getvalue()

        return await self.update_file_from_bytes(
            conn, id_, file_name, description, original_path=None, contents=img_byte_arr
        )

    async def create_image_from_nparray(
        self,
        conn: Connection,
        file_name: str,
        description: str,
        contents: npt.ArrayLike,  # pyright: ignore [reportUnknownParameterType]
        format_: str,
        deduplicate: bool,
    ) -> CreateFileResult:
        img_byte_arr_io = io.BytesIO()
        Image.fromarray(contents).save(img_byte_arr_io, format=format_)
        img_byte_arr = img_byte_arr_io.getvalue()

        return await self.create_file_from_bytes(
            conn,
            file_name,
            description,
            original_path=None,
            contents=img_byte_arr,
            deduplicate=deduplicate,
        )

    async def duplicate_file(
        self, conn: Connection, id_: int, new_file_name: str | None
    ) -> CreateFileResult:
        original_file = await self.retrieve_file(conn, id_, with_contents=True)
        return await self.create_file_from_bytes(
            conn,
            file_name=original_file.file_name
            if new_file_name is None
            else new_file_name,
            description=original_file.description,
            original_path=Path(original_file.original_path)
            if original_file.original_path is not None
            else None,
            contents=cast(bytes, original_file.contents),
            deduplicate=False,
        )

    async def update_file_from_bytes(
        self,
        conn: Connection,
        id_: int,
        file_name: str,
        description: str,
        original_path: Path | None,
        contents: bytes,
    ) -> None:
        mime = magic.from_buffer(contents, mime=True)

        sha256 = hashlib.sha256(contents).hexdigest()

        await conn.execute(
            self.tables.file.update()
            .values(
                type=mime,
                modified=datetime.datetime.utcnow(),
                contents=contents,
                file_name=file_name,
                original_path=str(original_path),
                description=description,
                sha256=sha256,
                size_in_bytes=len(contents),
            )
            .where(self.tables.file.c.id == id_)
        )

    async def create_file_from_bytes(
        self,
        conn: Connection,
        file_name: str,
        description: str,
        original_path: Path | None,
        contents: bytes,
        deduplicate: bool,
    ) -> CreateFileResult:
        mime = magic.from_buffer(contents, mime=True)

        sha256 = hashlib.sha256(contents).hexdigest()

        if deduplicate:
            existing_file = (
                await conn.execute(
                    sa.select(
                        self.tables.file.c.id,
                        self.tables.file.c.type,
                        self.tables.file.c.size_in_bytes,
                    ).where(
                        (self.tables.file.c.sha256 == sha256)
                        & (self.tables.file.c.file_name == file_name)
                    )
                )
            ).fetchone()

            if existing_file is not None:
                logger.info(f"file {file_name} already found, not creating another one")
                return CreateFileResult(
                    existing_file["id"],
                    existing_file["type"],
                    existing_file["size_in_bytes"],
                )

        return CreateFileResult(
            id=(
                await conn.execute(
                    self.tables.file.insert().values(
                        type=mime,
                        modified=datetime.datetime.utcnow(),
                        contents=contents,
                        file_name=file_name,
                        original_path=str(original_path),
                        description=description,
                        sha256=sha256,
                        size_in_bytes=len(contents),
                    )
                )
            ).inserted_primary_key[0],
            type_=mime,
            size_in_bytes=len(contents),
        )

    async def create_file(
        self,
        conn: Connection,
        file_name: str,
        description: str,
        original_path: Path | None,
        contents_location: Path,
        deduplicate: bool,
    ) -> CreateFileResult:
        mime = magic.from_file(str(contents_location), mime=True)  # type: ignore

        sha256 = sha256_file(contents_location)

        if deduplicate:
            existing_file = (
                await conn.execute(
                    sa.select(
                        self.tables.file.c.id,
                        self.tables.file.c.type,
                        self.tables.file.c.size_in_bytes,
                    ).where(
                        (self.tables.file.c.sha256 == sha256)
                        & (self.tables.file.c.file_name == file_name)
                    )
                )
            ).fetchone()

            if existing_file is not None:
                logger.info(f"file {file_name} already found, not creating another one")
                return CreateFileResult(
                    existing_file["id"],
                    existing_file["type"],
                    existing_file["size_in_bytes"],
                )

        size_in_bytes = contents_location.stat().st_size
        with contents_location.open("rb") as f:
            return CreateFileResult(
                id=(
                    await conn.execute(
                        self.tables.file.insert().values(
                            type=mime,
                            modified=datetime.datetime.utcnow(),
                            contents=f.read(),
                            file_name=file_name,
                            original_path=str(original_path),
                            description=description,
                            sha256=sha256,
                            size_in_bytes=size_in_bytes,
                        )
                    )
                ).inserted_primary_key[0],
                type_=mime,
                size_in_bytes=size_in_bytes,
            )

    async def update_attributo(
        self,
        conn: Connection,
        name: AttributoId,
        conversion_flags: AttributoConversionFlags,
        new_attributo: DBAttributo,
    ) -> None:
        current_attributi = list(
            await self.retrieve_attributi(conn, associated_table=None)
        )

        current_attributo = next((x for x in current_attributi if x.name == name), None)

        if current_attributo is None:
            raise Exception(
                f"couldn't find attributo for table {new_attributo.associated_table} and name {name}"
            )

        existing_attributo = next(
            (a for a in current_attributi if a.name == new_attributo.name), None
        )
        if new_attributo.name != name and existing_attributo is not None:
            raise Exception(
                f"cannot rename {name} to {new_attributo.name} because we already have an attributo of that "
                + "name"
            )

        # first, change the attributo itself, then its actual values (if possible)
        await conn.execute(
            sa.update(self.tables.attributo)
            .values(
                name=new_attributo.name,
                description=new_attributo.description,
                group=new_attributo.group,
                json_schema=coparse_schema_type(
                    attributo_type_to_schema(new_attributo.attributo_type)
                ),
                associated_table=new_attributo.associated_table,
            )
            .where(self.tables.attributo.c.name == name)
        )

        if new_attributo.associated_table == AssociatedTable.CHEMICAL:
            # If we're changing the table from run to chemical, we have to remove the attributo from runs
            if current_attributo.associated_table != AssociatedTable.CHEMICAL:
                for r in await self.retrieve_runs(conn, current_attributi):
                    r.attributi.remove_with_type(current_attributo.name)
                    await self.update_run_attributi(conn, r.id, r.attributi)
            else:
                # Update column in chemical(s)
                for s in await self.retrieve_chemicals(conn, current_attributi):
                    s.attributi.convert_attributo(
                        conversion_flags=conversion_flags,
                        old_name=name,
                        new_name=new_attributo.name,
                        after_type=new_attributo.attributo_type,
                    )
                    await self.update_chemical(
                        conn=conn,
                        id_=s.id,
                        type_=s.type_,
                        name=s.name,
                        responsible_person=s.responsible_person,
                        attributi=s.attributi,
                    )
        elif new_attributo.associated_table == AssociatedTable.RUN:
            # If we're changing the table from chemical to run, we have to remove the attributo from chemicals
            if current_attributo.associated_table != AssociatedTable.RUN:
                for s in await self.retrieve_chemicals(conn, current_attributi):
                    s.attributi.remove_with_type(current_attributo.name)
                    await self.update_chemical(
                        conn=conn,
                        id_=s.id,
                        type_=s.type_,
                        name=s.name,
                        responsible_person=s.responsible_person,
                        attributi=s.attributi,
                    )
            else:
                for r in await self.retrieve_runs(conn, current_attributi):
                    r.attributi.convert_attributo(
                        conversion_flags=conversion_flags,
                        old_name=name,
                        new_name=new_attributo.name,
                        after_type=new_attributo.attributo_type,
                    )
                    await self.update_run_attributi(conn, r.id, r.attributi)
        else:
            raise Exception(
                f"unimplemented: is there a new associated table {new_attributo.associated_table}?"
            )

        chemical_ids = await self.retrieve_chemical_ids(conn)
        for ds in await self.retrieve_data_sets(conn, chemical_ids, current_attributi):
            ds.attributi.convert_attributo(
                conversion_flags=conversion_flags,
                old_name=name,
                new_name=new_attributo.name,
                after_type=new_attributo.attributo_type,
            )
            await self.update_data_set_attributi(conn, ds.id, ds.attributi)

    async def add_file_to_chemical(
        self, conn: Connection, file_id: int, chemical_id: int
    ) -> None:
        await conn.execute(
            sa.insert(self.tables.chemical_has_file).values(
                file_id=file_id, chemical_id=chemical_id
            )
        )

    async def add_file_to_event(
        self, conn: Connection, file_id: int, event_id: int
    ) -> None:
        await conn.execute(
            sa.insert(self.tables.event_has_file).values(
                file_id=file_id, event_id=event_id
            )
        )

    async def remove_files_from_chemical(
        self, conn: Connection, chemical_id: int
    ) -> None:
        await conn.execute(
            sa.delete(self.tables.chemical_has_file).where(
                self.tables.chemical_has_file.c.chemical_id == chemical_id
            )
        )

    async def retrieve_run_ids(self, conn: Connection) -> list[int]:
        sel = await conn.execute(
            sa.select([self.tables.run.c.id]).order_by(self.tables.run.c.id)
        )
        return [s[0] for s in sel.fetchall()]

    async def create_indexing_result(
        self, conn: Connection, r: DBIndexingResultInput
    ) -> int:
        fom = (
            r.runtime_status.fom
            if isinstance(
                r.runtime_status,
                (DBIndexingResultRunning, DBIndexingResultDone),
            )
            else None
        )
        return (  # type: ignore
            await conn.execute(
                sa.insert(self.tables.indexing_result).values(
                    created=r.created,
                    run_id=r.run_id,
                    stream_file=str(r.runtime_status.stream_file)
                    if isinstance(
                        r.runtime_status,
                        (DBIndexingResultRunning, DBIndexingResultDone),
                    )
                    else None,
                    hits=r.hits,
                    frames=r.frames,
                    not_indexed_frames=r.not_indexed_frames,
                    hit_rate=fom.hit_rate if fom is not None else 0.0,
                    indexing_rate=fom.indexing_rate if fom is not None else 0.0,
                    indexed_frames=fom.indexed_frames if fom is not None else 0.0,
                    cell_description=coparse_cell_description(r.cell_description)
                    if r.cell_description is not None
                    else None,
                    point_group=r.point_group,
                    chemical_id=r.chemical_id,
                    job_id=r.runtime_status.job_id
                    if isinstance(r.runtime_status, DBIndexingResultRunning)
                    else None,
                    job_status=DBJobStatus.RUNNING
                    if isinstance(r.runtime_status, DBIndexingResultRunning)
                    else DBJobStatus.QUEUED
                    if r.runtime_status is None
                    else DBJobStatus.DONE,
                    job_error=r.runtime_status.job_error
                    if isinstance(r.runtime_status, DBIndexingResultDone)
                    else None,
                )
            )
        ).inserted_primary_key[0]

    async def create_run(
        self,
        conn: Connection,
        run_id: int,
        attributi: list[DBAttributo],
        attributi_map: AttributiMap,
        experiment_type_id: int,
        keep_manual_attributes_from_previous_run: bool,
    ) -> None:
        final_attributi_map: AttributiMap
        if keep_manual_attributes_from_previous_run:
            latest_run = await self.retrieve_latest_run(conn, attributi)
            if latest_run is not None:
                final_attributi_map = latest_run.attributi.create_sub_map_for_group(
                    ATTRIBUTO_GROUP_MANUAL
                )
                final_attributi_map.remove_but_keep_type(ATTRIBUTO_STOPPED)
                final_attributi_map.extend_with_attributi_map(attributi_map)
            else:
                final_attributi_map = attributi_map
        else:
            final_attributi_map = attributi_map
        await conn.execute(
            sa.insert(self.tables.run).values(
                id=run_id,
                experiment_type_id=experiment_type_id,
                attributi=final_attributi_map.to_json(),
                modified=datetime.datetime.utcnow(),
            )
        )

    async def retrieve_latest_run(
        self, conn: Connection, attributi: list[DBAttributo]
    ) -> DBRun | None:
        maximum_id = (
            await conn.execute(sa.select([sa.func.max(self.tables.run.c.id)]))
        ).fetchone()

        if maximum_id is None:
            return None

        return await self.retrieve_run(conn, maximum_id[0], attributi)

    async def retrieve_chemical(
        self, conn: Connection, id_: int, attributi: list[DBAttributo]
    ) -> DBChemical | None:
        rc = self.tables.chemical.c
        r = (
            await conn.execute(
                sa.select(
                    [rc.id, rc.name, rc.type, rc.responsible_person, rc.attributi]
                ).where(rc.id == id_)
            )
        ).fetchone()
        files = await self._retrieve_files(
            conn,
            self.tables.chemical_has_file.c.chemical_id,
            (self.tables.chemical_has_file.c.chemical_id == id_),
        )
        if r is None:
            return None

        return DBChemical(
            id=id_,
            name=r["name"],
            responsible_person=r["responsible_person"],
            type_=r["type"],
            attributi=AttributiMap.from_types_and_json(
                attributi, chemical_ids=[], raw_attributi=r["attributi"]
            ),
            files=files.get(id_, []),
        )

    async def retrieve_run(
        self, conn: Connection, id_: int, attributi: list[DBAttributo]
    ) -> DBRun | None:
        rc = self.tables.run.c
        r = (
            await conn.execute(
                sa.select(rc.id, rc.experiment_type_id, rc.attributi).where(
                    rc.id == id_
                )
            )
        ).fetchone()
        files = await self._retrieve_files(
            conn,
            self.tables.run_has_file.c.run_id,
            (self.tables.run_has_file.c.run_id == id_),
        )
        if r is None:
            return None
        return DBRun(
            id=id_,
            experiment_type_id=r["experiment_type_id"],
            attributi=AttributiMap.from_types_and_json(
                attributi, await self.retrieve_chemical_ids(conn), r["attributi"]
            ),
            files=files.get(id_, []),
        )

    async def update_run_experiment_type(
        self, conn: Connection, id_: int, experiment_type_id: int
    ) -> None:
        await conn.execute(
            sa.update(self.tables.run)
            .values(experiment_type_id=experiment_type_id)
            .where(self.tables.run.c.id == id_)
        )

    async def update_run_attributi(
        self, conn: Connection, id_: int, attributi: AttributiMap
    ) -> None:
        await conn.execute(
            sa.update(self.tables.run)
            .values(attributi=attributi.to_json())
            .where(self.tables.run.c.id == id_)
        )

    async def retrieve_chemical_ids(self, conn: Connection) -> list[int]:
        return [
            r[0] for r in await conn.execute(sa.select([self.tables.chemical.c.id]))
        ]

    async def create_experiment_type(
        self,
        conn: Connection,
        name: str,
        experiment_attributi: Iterable[AttributoNameAndRole],
    ) -> int:
        existing_attributi_names = {
            a.name for a in await self.retrieve_attributi(conn, associated_table=None)
        }

        not_found = [
            a.attributo_name
            for a in experiment_attributi
            if a.attributo_name not in existing_attributi_names
        ]

        if not_found:
            raise Exception(
                "couldn't find the following attributi: " + ", ".join(not_found)
            )

        experiment_type_id = (
            await conn.execute(
                self.tables.experiment_type.insert().values({"name": name})
            )
        ).inserted_primary_key[0]

        await conn.execute(
            self.tables.experiment_has_attributo.insert().values(
                [
                    {
                        "experiment_type_id": experiment_type_id,
                        "attributo_name": a.attributo_name,
                        "chemical_role": a.chemical_role,
                    }
                    for a in experiment_attributi
                ]
            )
        )

        return experiment_type_id  # type: ignore

    async def delete_experiment_type(self, conn: Connection, id_: int) -> None:
        await conn.execute(
            self.tables.experiment_type.delete().where(
                self.tables.experiment_type.c.id == id_
            )
        )

    async def retrieve_experiment_types(
        self, conn: Connection
    ) -> list[DBExperimentType]:
        result: list[DBExperimentType] = []
        etc = self.tables.experiment_has_attributo.c
        et = self.tables.experiment_type
        for key, group in itertools.groupby(
            await conn.execute(
                sa.select([et.c.id, et.c.name, etc.attributo_name, etc.chemical_role])
                .join(et, et.c.id == etc.experiment_type_id)
                .order_by(etc.experiment_type_id)
            ),
            key=lambda row: row["id"],  # type: ignore
        ):
            group_list = list(group)
            result.append(
                DBExperimentType(
                    id=key,
                    name=group_list[0]["name"],
                    attributi=[
                        AttributoNameAndRole(
                            attributo_name=row["attributo_name"],
                            chemical_role=row["chemical_role"],
                        )
                        for row in group_list
                    ],
                )
            )
        return result

    async def create_data_set(
        self, conn: Connection, experiment_type_id: int, attributi: AttributiMap
    ) -> int:
        matching_experiment_type: DBExperimentType | None = next(
            (
                x
                for x in await self.retrieve_experiment_types(conn)
                if x.id == experiment_type_id
            ),
            None,
        )
        if matching_experiment_type is None:
            raise Exception(
                f'couldn\'t find experiment type with ID "{experiment_type_id}"'
            )

        existing_attributo_names = [
            a.attributo_name for a in matching_experiment_type.attributi
        ]

        superfluous_attributi = attributi.names().difference(existing_attributo_names)

        if superfluous_attributi:
            raise Exception(
                "the following attributi are not part of the data set definition: "
                + ", ".join(superfluous_attributi)
            )

        if not attributi.items():
            raise Exception("You have to assign at least one value to an attributo")

        data_set_id: int = (
            await conn.execute(
                self.tables.data_set.insert().values(
                    experiment_type_id=experiment_type_id,
                    attributi=attributi.to_json(),
                )
            )
        ).inserted_primary_key[0]

        return data_set_id

    async def delete_data_set(self, conn: Connection, id_: int) -> None:
        await conn.execute(
            self.tables.data_set.delete().where(self.tables.data_set.c.id == id_)
        )

    async def retrieve_data_set(
        self,
        conn: Connection,
        data_set_id: int,
        chemical_ids: list[int],
        attributi: list[DBAttributo],
    ) -> None | DBDataSet:
        dc = self.tables.data_set.c
        r = (
            await conn.execute(
                sa.select([dc.experiment_type_id, dc.attributi]).where(
                    dc.id == data_set_id
                )
            )
        ).fetchone()
        if r is None:
            return None
        return DBDataSet(
            id=data_set_id,
            experiment_type_id=r["experiment_type_id"],
            attributi=AttributiMap.from_types_and_json(
                attributi, chemical_ids=chemical_ids, raw_attributi=r["attributi"]
            ),
        )

    async def retrieve_data_sets(
        self,
        conn: Connection,
        chemical_ids: list[int],
        attributi: list[DBAttributo],
    ) -> list[DBDataSet]:
        dc = self.tables.data_set.c
        return [
            DBDataSet(
                id=r["id"],
                experiment_type_id=r["experiment_type_id"],
                attributi=AttributiMap.from_types_and_json(
                    attributi, chemical_ids=chemical_ids, raw_attributi=r["attributi"]
                ),
            )
            for r in await conn.execute(
                sa.select([dc.id, dc.experiment_type_id, dc.attributi])
            )
        ]

    async def update_data_set_attributi(
        self, conn: Connection, id_: int, attributi: AttributiMap
    ) -> None:
        dc = self.tables.data_set.c
        await conn.execute(
            sa.update(self.tables.data_set)
            .values(attributi=attributi.to_json())
            .where(dc.id == id_)
        )

    async def retrieve_bulk_run_attributi(
        self, conn: Connection, attributi: list[DBAttributo], run_ids: Iterable[int]
    ) -> dict[AttributoId, set[AttributoValue]]:
        runs = await self.retrieve_runs(conn, attributi)

        uninteresting_attributi = (ATTRIBUTO_STOPPED, ATTRIBUTO_STARTED)
        interesting_attributi = [
            a for a in attributi if a.name not in uninteresting_attributi
        ]

        attributi_values: dict[AttributoId, set[AttributoValue]] = {
            a.name: set() for a in interesting_attributi
        }
        for run in (run for run in runs if run.id in run_ids):
            for a in (a for a in interesting_attributi):
                run_attributo_value = run.attributi.select(a.name)
                if run_attributo_value is not None:
                    attributi_values[a.name].add(run_attributo_value)
        return attributi_values

    async def update_bulk_run_attributi(
        self,
        conn: Connection,
        attributi: list[DBAttributo],
        run_ids: set[int],
        attributi_values: AttributiMap,
    ) -> None:
        runs = await self.retrieve_runs(conn, attributi)

        for run in (run for run in runs if run.id in run_ids):
            for attributo_id, attributo_value in attributi_values.items():
                run.attributi.append_single(attributo_id, attributo_value)
            await self.update_run_attributi(conn, run.id, run.attributi)

    async def retrieve_indexing_results(
        self, conn: Connection, job_status_filter: None | DBJobStatus = None
    ) -> list[DBIndexingResultOutput]:
        ir = self.tables.indexing_result

        return [
            DBIndexingResultOutput(
                id=r["id"],
                created=r["created"],
                run_id=r["run_id"],
                frames=r["frames"],
                hits=r["hits"],
                not_indexed_frames=r["not_indexed_frames"],
                cell_description=parse_cell_description(r["cell_description"])
                if r["cell_description"] is not None
                else None,
                point_group=r["point_group"],
                chemical_id=r["chemical_id"],
                runtime_status=_evaluate_indexing_result_runtime_status(r),
            )
            for r in await conn.execute(
                sa.select(
                    [
                        ir.c.id,
                        ir.c.created,
                        ir.c.run_id,
                        ir.c.stream_file,
                        ir.c.frames,
                        ir.c.hits,
                        ir.c.not_indexed_frames,
                        ir.c.cell_description,
                        ir.c.point_group,
                        ir.c.chemical_id,
                        ir.c.hit_rate,
                        ir.c.indexing_rate,
                        ir.c.indexed_frames,
                        ir.c.detector_shift_x_mm,
                        ir.c.detector_shift_y_mm,
                        ir.c.job_id,
                        ir.c.job_status,
                        ir.c.job_error,
                    ]
                ).where(
                    ir.c.job_status == job_status_filter
                    if job_status_filter is not None
                    else True
                )
            )
        ]

    async def update_indexing_result(
        self,
        conn: Connection,
        indexing_result_id: int,
        new_result: DBIndexingResultInput,
    ) -> None:
        runtime_status = new_result.runtime_status
        await conn.execute(
            sa.update(self.tables.indexing_result)
            .values(
                {
                    "stream_file": None,
                    "cell_description": coparse_cell_description(
                        new_result.cell_description
                    )
                    if new_result.cell_description is not None
                    else None,
                    "point_group": new_result.point_group,
                    "chemical_id": new_result.chemical_id,
                    "frames": new_result.frames,
                    "hits": new_result.hits,
                    "not_indexed_frames": new_result.not_indexed_frames,
                    "job_id": None,
                    "job_status": DBJobStatus.QUEUED,
                    "job_error": None,
                }
                if runtime_status is None
                else {
                    "stream_file": str(runtime_status.stream_file),
                    "cell_description": coparse_cell_description(
                        new_result.cell_description
                    )
                    if new_result.cell_description is not None
                    else None,
                    "point_group": new_result.point_group,
                    "chemical_id": new_result.chemical_id,
                    "frames": new_result.frames,
                    "hits": new_result.hits,
                    "not_indexed_frames": new_result.not_indexed_frames,
                    "hit_rate": runtime_status.fom.hit_rate,
                    "indexing_rate": runtime_status.fom.indexing_rate,
                    "indexed_frames": runtime_status.fom.indexed_frames,
                    "job_id": runtime_status.job_id,
                    "job_status": DBJobStatus.RUNNING,
                    "job_error": None,
                }
                if isinstance(runtime_status, DBIndexingResultRunning)
                else {
                    "stream_file": str(runtime_status.stream_file),
                    "cell_description": coparse_cell_description(
                        new_result.cell_description
                    )
                    if new_result.cell_description is not None
                    else None,
                    "point_group": new_result.point_group,
                    "chemical_id": new_result.chemical_id,
                    "frames": new_result.frames,
                    "hits": new_result.hits,
                    "not_indexed_frames": new_result.not_indexed_frames,
                    "hit_rate": runtime_status.fom.hit_rate,
                    "indexing_rate": runtime_status.fom.indexing_rate,
                    "indexed_frames": runtime_status.fom.indexed_frames,
                    "job_id": None,
                    "job_status": DBJobStatus.DONE,
                    "job_error": runtime_status.job_error,
                }
            )
            .where(self.tables.indexing_result.c.id == indexing_result_id)
        )

    async def add_indexing_result_statistic(
        self, conn: Connection, s: DBIndexingResultStatistic
    ) -> None:
        await conn.execute(
            self.tables.indexing_result_has_statistic.insert().values(
                indexing_result_id=s.indexing_result_id,
                time=s.time,
                frames=s.frames,
                hits=s.hits,
                indexed_frames=s.indexed_frames,
                indexed_crystals=s.indexed_crystals,
            )
        )

    async def retrieve_indexing_result_statistics(
        self,
        conn: Connection,
        indexing_result_id: None | int,
    ) -> list[DBIndexingResultStatistic]:
        irhs = self.tables.indexing_result_has_statistic.c
        statement = sa.select(
            irhs.indexing_result_id,
            irhs.time,
            irhs.frames,
            irhs.hits,
            irhs.indexed_frames,
            irhs.indexed_crystals,
        ).order_by(irhs.time)
        if indexing_result_id is not None:
            statement = statement.where(irhs.indexing_result_id == indexing_result_id)
        return [
            DBIndexingResultStatistic(
                indexing_result_id=s["indexing_result_id"],
                time=s["time"],
                frames=s["frames"],
                hits=s["hits"],
                indexed_frames=s["indexed_frames"],
                indexed_crystals=s["indexed_crystals"],
            )
            for s in (await conn.execute(statement)).fetchall()
        ]

    async def update_indexing_result_status(
        self,
        conn: Connection,
        indexing_result_id: int,
        runtime_status: DBIndexingResultRuntimeStatus,
    ) -> None:
        await conn.execute(
            sa.update(self.tables.indexing_result)
            .values(
                {
                    "stream_file": None,
                    "job_id": None,
                    "job_status": DBJobStatus.QUEUED,
                    "job_error": None,
                }
                if runtime_status is None
                else {
                    "stream_file": str(runtime_status.stream_file),
                    "hit_rate": runtime_status.fom.hit_rate,
                    "indexing_rate": runtime_status.fom.indexing_rate,
                    "indexed_frames": runtime_status.fom.indexed_frames,
                    "job_id": runtime_status.job_id,
                    "job_status": DBJobStatus.RUNNING,
                    "job_error": None,
                }
                if isinstance(runtime_status, DBIndexingResultRunning)
                else {
                    "stream_file": str(runtime_status.stream_file),
                    "hit_rate": runtime_status.fom.hit_rate,
                    "indexing_rate": runtime_status.fom.indexing_rate,
                    "indexed_frames": runtime_status.fom.indexed_frames,
                    "job_id": None,
                    "job_status": DBJobStatus.DONE,
                    "job_error": runtime_status.job_error,
                }
            )
            .where(self.tables.indexing_result.c.id == indexing_result_id)
        )

    async def retrieve_indexing_result(
        self, conn: Connection, result_id: int
    ) -> None | DBIndexingResultOutput:
        ir = self.tables.indexing_result

        r = (
            await conn.execute(
                sa.select(
                    [
                        ir.c.id,
                        ir.c.created,
                        ir.c.run_id,
                        ir.c.stream_file,
                        ir.c.frames,
                        ir.c.hits,
                        ir.c.not_indexed_frames,
                        ir.c.hit_rate,
                        ir.c.indexing_rate,
                        ir.c.indexed_frames,
                        ir.c.detector_shift_x_mm,
                        ir.c.detector_shift_y_mm,
                        ir.c.cell_description,
                        ir.c.point_group,
                        ir.c.chemical_id,
                        ir.c.job_id,
                        ir.c.job_status,
                        ir.c.job_error,
                    ]
                ).where(ir.c.id == result_id)
            )
        ).fetchone()

        if r is None:
            return None

        return DBIndexingResultOutput(
            id=r["id"],
            created=r["created"],
            run_id=r["run_id"],
            frames=r["frames"],
            hits=r["hits"],
            not_indexed_frames=r["not_indexed_frames"],
            cell_description=parse_cell_description(r["cell_description"])
            if r["cell_description"] is not None
            else None,
            point_group=r["point_group"],
            chemical_id=r["chemical_id"],
            runtime_status=_evaluate_indexing_result_runtime_status(r),
        )

    async def retrieve_indexing_result_for_run(
        self, conn: Connection, run_id: int
    ) -> None | DBIndexingResultOutput:
        ir = self.tables.indexing_result

        r = (
            await conn.execute(
                sa.select(
                    [
                        ir.c.id,
                        ir.c.created,
                        ir.c.run_id,
                        ir.c.stream_file,
                        ir.c.frames,
                        ir.c.hits,
                        ir.c.not_indexed_frames,
                        ir.c.hit_rate,
                        ir.c.indexing_rate,
                        ir.c.indexed_frames,
                        ir.c.detector_shift_x_mm,
                        ir.c.detector_shift_y_mm,
                        ir.c.cell_description,
                        ir.c.point_group,
                        ir.c.chemical_id,
                        ir.c.job_id,
                        ir.c.job_status,
                        ir.c.job_error,
                    ]
                ).where(ir.c.run_id == run_id)
            )
        ).fetchone()

        if r is None:
            return None

        return DBIndexingResultOutput(
            id=r["id"],
            created=r["created"],
            run_id=r["run_id"],
            frames=r["frames"],
            hits=r["hits"],
            not_indexed_frames=r["not_indexed_frames"],
            cell_description=parse_cell_description(r["cell_description"])
            if r["cell_description"] is not None
            else None,
            point_group=r["point_group"],
            chemical_id=r["chemical_id"],
            runtime_status=_evaluate_indexing_result_runtime_status(r),
        )

    async def create_merge_result(
        self, conn: Connection, merge_result: DBMergeResultInput
    ) -> MergeResultError | int:
        if not merge_result.indexing_results:
            return MergeResultError("list of indexing results was empty")

        point_groups = set(
            ir.point_group
            for ir in merge_result.indexing_results
            if ir.point_group is not None
        )
        cell_descriptions = set(
            ir.cell_description
            for ir in merge_result.indexing_results
            if ir.cell_description is not None
        )

        if len(point_groups) > 1:
            return MergeResultError(
                "we have more than one possible point group, namely: "
                + ", ".join(point_groups)
            )

        if len(cell_descriptions) > 1:
            return MergeResultError(
                "we have more than one possible cell description, namely: "
                + ", ".join(coparse_cell_description(cd) for cd in cell_descriptions)
            )

        if not cell_descriptions:
            return MergeResultError(
                "we need a cell description for merging, but found none"
            )

        # Explicit type necessary here, mypy will try to be too intelligent with type inference.
        mrp = merge_result.parameters
        merge_result_dict: dict[str, Any] = {
            "created": merge_result.created,
            "point_group": next(iter(point_groups), None),
            "cell_description": next(
                (coparse_cell_description(c) for c in cell_descriptions), None
            ),
            "input_merge_model": mrp.merge_model,
            "input_scale_intensities": mrp.scale_intensities,
            "input_post_refinement": mrp.post_refinement,
            "input_iterations": mrp.iterations,
            "input_polarisation_angle": int(
                mrp.polarisation.angle.to(
                    _UNIT_REGISTRY.degrees
                ).m  # pyright: ignore [reportUnknownArgumentType]
            )
            if mrp.polarisation is not None
            else None,
            "input_polarisation_percent": mrp.polarisation.percentage
            if mrp.polarisation is not None
            else None,
            "input_start_after": mrp.start_after,
            "input_stop_after": mrp.stop_after,
            "input_rel_b": mrp.rel_b,
            "input_no_pr": mrp.no_pr,
            "input_force_bandwidth": mrp.force_bandwidth,
            "input_force_radius": mrp.force_radius,
            "input_force_lambda": mrp.force_lambda,
            "input_no_delta_cc_half": mrp.no_delta_cc_half,
            "input_max_adu": mrp.max_adu,
            "input_min_measurements": mrp.min_measurements,
            "input_logs": mrp.logs,
            "input_min_res": mrp.min_res,
            "input_push_res": mrp.push_res,
            "input_w": mrp.w,
            "negative_handling": mrp.negative_handling,
        } | _runtime_status_sql_values(merge_result.runtime_status)

        mr_id = (
            await conn.execute(
                self.tables.merge_result.insert().values(merge_result_dict)
            )
        ).inserted_primary_key[0]
        for ir in merge_result.indexing_results:
            await conn.execute(
                self.tables.merge_result_has_indexing_result.insert().values(
                    merge_result_id=mr_id, indexing_result_id=ir.id
                )
            )
        return mr_id  # type: ignore

    async def retrieve_merge_results(
        self,
        conn: Connection,
        job_status_filter: None | DBJobStatus = None,
        merge_result_id_filter: None | int = None,
    ) -> list[DBMergeResultOutput]:
        indexing_results_by_id: dict[int, DBIndexingResultOutput] = {
            ir.id: ir for ir in await self.retrieve_indexing_results(conn)
        }
        refinement_results_by_merge_id: dict[
            int, list[DBRefinementResultOutput]
        ] = group_by(
            await self.retrieve_refinement_results(conn),
            lambda rr: rr.merge_result_id,
        )
        merge_result_to_indexing_result: dict[int, list[int]] = {
            merge_result_id: [row["indexing_result_id"] for row in indexing_results]
            for merge_result_id, indexing_results in group_by(
                await conn.execute(
                    sa.select(
                        [
                            self.tables.merge_result_has_indexing_result.c.merge_result_id,
                            self.tables.merge_result_has_indexing_result.c.indexing_result_id,
                        ]
                    )
                ),
                lambda row: row["merge_result_id"],  # type: ignore
            ).items()
        }

        def convert_merge_result_row(
            row: Any, detailed_foms: list[Any]
        ) -> DBMergeResultOutput:
            runtime_status: DBMergeRuntimeStatus
            match row["job_status"]:
                case DBJobStatus.QUEUED:
                    runtime_status = None
                case DBJobStatus.RUNNING:
                    runtime_status = DBMergeRuntimeStatusRunning(
                        job_id=row["job_id"],
                        started=row["started"],
                        recent_log=row["recent_log"],
                    )
                case _:
                    if row["job_error"]:
                        # pylint: disable=redefined-variable-type
                        runtime_status = DBMergeRuntimeStatusError(
                            error=row["job_error"],
                            started=row["started"],
                            stopped=row["stopped"],
                            recent_log=row["recent_log"],
                        )
                    else:
                        runtime_status = DBMergeRuntimeStatusDone(
                            started=row["started"],
                            stopped=row["stopped"],
                            recent_log=row["recent_log"],
                            result=MergeResult(
                                detailed_foms=detailed_foms,
                                refinement_results=[
                                    RefinementResult(
                                        pdb_file_id=rr.pdb_file_id,
                                        mtz_file_id=rr.mtz_file_id,
                                        r_free=rr.r_free,
                                        r_work=rr.r_work,
                                        rms_bond_angle=rr.rms_bond_angle,
                                        rms_bond_length=rr.rms_bond_length,
                                    )
                                    for rr in refinement_results_by_merge_id.get(
                                        row["id"], []
                                    )
                                ],
                                mtz_file_id=row["mtz_file_id"],
                                fom=MergeResultFom(
                                    snr=row["fom_snr"],
                                    wilson=row["fom_wilson"],
                                    ln_k=row["fom_ln_k"],
                                    discarded_reflections=row[
                                        "fom_discarded_reflections"
                                    ],
                                    one_over_d_from=row["fom_one_over_d_from"],
                                    one_over_d_to=row["fom_one_over_d_to"],
                                    redundancy=row["fom_redundancy"],
                                    completeness=row["fom_completeness"],
                                    measurements_total=row["fom_measurements_total"],
                                    reflections_total=row["fom_reflections_total"],
                                    reflections_possible=row[
                                        "fom_reflections_possible"
                                    ],
                                    r_split=row["fom_r_split"],
                                    r1i=row["fom_r1i"],
                                    r2=row["fom_2"],
                                    cc=row["fom_cc"],
                                    ccstar=row["fom_ccstar"],
                                    ccano=row["fom_ccano"],
                                    crdano=row["fom_crdano"],
                                    rano=row["fom_rano"],
                                    rano_over_r_split=row["fom_rano_over_r_split"],
                                    d1sig=row["fom_d1sig"],
                                    d2sig=row["fom_d2sig"],
                                    outer_shell=MergeResultOuterShell(
                                        resolution=row["fom_outer_resolution"],
                                        ccstar=row["fom_outer_ccstar"],
                                        r_split=row["fom_outer_r_split"],
                                        cc=row["fom_outer_cc"],
                                        unique_reflections=row[
                                            "fom_outer_unique_reflections"
                                        ],
                                        completeness=row["fom_outer_completeness"],
                                        redundancy=row["fom_outer_redundancy"],
                                        snr=row["fom_outer_snr"],
                                        min_res=row["fom_outer_min_res"],
                                        max_res=row["fom_outer_max_res"],
                                    ),
                                ),
                            ),
                        )

            return DBMergeResultOutput(
                id=row["id"],
                created=row["created"],
                indexing_results=[
                    indexing_results_by_id[ir_id]
                    for ir_id in merge_result_to_indexing_result.get(row["id"], [])
                ],
                parameters=DBMergeParameters(
                    point_group=row["point_group"],
                    cell_description=cast(
                        CrystFELCellFile,
                        parse_cell_description(row["cell_description"]),
                    ),
                    negative_handling=row["negative_handling"],
                    merge_model=MergeModel(row["input_merge_model"]),
                    scale_intensities=ScaleIntensities(row["input_scale_intensities"]),
                    post_refinement=row["input_post_refinement"],
                    iterations=row["input_iterations"],
                    polarisation=Polarisation(
                        angle=row["input_polarisation_angle"]
                        * _UNIT_REGISTRY.degrees,  # pyright: ignore [reportUnknownArgumentType]
                        percentage=row["input_polarisation_percent"],
                    )
                    if row["input_polarisation_angle"] is not None
                    and row["input_polarisation_percent"] is not None
                    else None,
                    start_after=row["input_start_after"],
                    stop_after=row["input_stop_after"],
                    rel_b=row["input_rel_b"],
                    no_pr=row["input_no_pr"],
                    force_bandwidth=row["input_force_bandwidth"],
                    force_radius=row["input_force_radius"],
                    force_lambda=row["input_force_lambda"],
                    no_delta_cc_half=row["input_no_delta_cc_half"],
                    max_adu=row["input_max_adu"],
                    min_measurements=row["input_min_measurements"],
                    logs=row["input_logs"],
                    min_res=row["input_min_res"],
                    push_res=row["input_push_res"],
                    w=row["input_w"],
                ),
                runtime_status=runtime_status,
            )

        mr = self.tables.merge_result.c
        mdfoms = self.tables.merge_result_shell_fom.c

        merge_results = await conn.execute(
            sa.select(
                [
                    mr.id,
                    mr.created,
                    mr.recent_log,
                    mr.negative_handling,
                    mr.job_status,
                    mr.started,
                    mr.stopped,
                    mr.point_group,
                    mr.cell_description,
                    mr.job_id,
                    mr.job_error,
                    mr.mtz_file_id,
                    mr.input_merge_model,
                    mr.input_scale_intensities,
                    mr.input_post_refinement,
                    mr.input_iterations,
                    mr.input_polarisation_angle,
                    mr.input_polarisation_percent,
                    mr.input_start_after,
                    mr.input_stop_after,
                    mr.input_rel_b,
                    mr.input_no_pr,
                    mr.input_force_bandwidth,
                    mr.input_force_radius,
                    mr.input_force_lambda,
                    mr.input_no_delta_cc_half,
                    mr.input_max_adu,
                    mr.input_min_measurements,
                    mr.input_logs,
                    mr.input_min_res,
                    mr.input_push_res,
                    mr.input_w,
                    mr.fom_snr,
                    mr.fom_wilson,
                    mr.fom_ln_k,
                    mr.fom_discarded_reflections,
                    mr.fom_one_over_d_from,
                    mr.fom_one_over_d_to,
                    mr.fom_redundancy,
                    mr.fom_completeness,
                    mr.fom_measurements_total,
                    mr.fom_reflections_total,
                    mr.fom_reflections_possible,
                    mr.fom_r_split,
                    mr.fom_r1i,
                    mr.fom_2,
                    mr.fom_cc,
                    mr.fom_ccstar,
                    mr.fom_ccano,
                    mr.fom_crdano,
                    mr.fom_rano,
                    mr.fom_rano_over_r_split,
                    mr.fom_d1sig,
                    mr.fom_d2sig,
                    mr.fom_outer_resolution,
                    mr.fom_outer_ccstar,
                    mr.fom_outer_r_split,
                    mr.fom_outer_cc,
                    mr.fom_outer_unique_reflections,
                    mr.fom_outer_completeness,
                    mr.fom_outer_redundancy,
                    mr.fom_outer_snr,
                    mr.fom_outer_min_res,
                    mr.fom_outer_max_res,
                ]
            ).where(
                sa.and_(
                    mr.job_status == job_status_filter  # pyright: ignore
                    if job_status_filter is not None
                    else True,
                    mr.id == merge_result_id_filter  # pyright: ignore
                    if merge_result_id_filter is not None
                    else True,
                )
            )
        )

        async def fetch_shell_foms(merge_result_id: int) -> list[MergeResultShell]:
            merge_result_shells = await conn.execute(
                sa.select(
                    [
                        mdfoms.one_over_d_centre,
                        mdfoms.nref,
                        mdfoms.d_over_a,
                        mdfoms.min_res,
                        mdfoms.max_res,
                        mdfoms.cc,
                        mdfoms.ccstar,
                        mdfoms.r_split,
                        mdfoms.reflections_possible,
                        mdfoms.completeness,
                        mdfoms.measurements,
                        mdfoms.redundancy,
                        mdfoms.snr,
                        mdfoms.mean_i,
                    ]
                ).where(mdfoms.merge_result_id == merge_result_id)
            )
            return [
                MergeResultShell(
                    one_over_d_centre=s["one_over_d_centre"],
                    nref=s["nref"],
                    d_over_a=s["d_over_a"],
                    min_res=s["min_res"],
                    max_res=s["max_res"],
                    cc=s["cc"],
                    ccstar=s["ccstar"],
                    r_split=s["r_split"],
                    reflections_possible=s["reflections_possible"],
                    completeness=s["completeness"],
                    measurements=s["measurements"],
                    redundancy=s["redundancy"],
                    snr=s["snr"],
                    mean_i=s["mean_i"],
                )
                for s in merge_result_shells
            ]

        return [
            convert_merge_result_row(
                row=row, detailed_foms=await fetch_shell_foms(row["id"])
            )
            for row in merge_results
        ]

    async def update_merge_result_status(
        self,
        conn: Connection,
        merge_result_id: int,
        runtime_status: DBMergeRuntimeStatus,
    ) -> None:
        await conn.execute(
            sa.update(self.tables.merge_result)
            .values(_runtime_status_sql_values(runtime_status))
            .where(self.tables.merge_result.c.id == merge_result_id)
        )

        if isinstance(runtime_status, DBMergeRuntimeStatusDone):
            detailed_forms = runtime_status.result.detailed_foms
            assert detailed_forms is not None
            for shell in detailed_forms:
                await conn.execute(
                    sa.insert(self.tables.merge_result_shell_fom).values(
                        _values_from_shell_foms(shell, merge_result_id=merge_result_id)
                    )
                )


# Any until openpyxl has official types
def attributo_value_to_spreadsheet_cell(
    chemical_id_to_name: dict[int, str],
    attributo_type: AttributoType,
    attributo_value: AttributoValue,
) -> Any:
    if attributo_value is None:
        return None
    if isinstance(attributo_type, AttributoTypeChemical):
        if not isinstance(attributo_value, int):
            raise TypeError(
                f"chemical IDs have to have type int, got {attributo_value}"
            )
        return chemical_id_to_name.get(
            attributo_value, f"invalid chemical ID {attributo_value}"
        )
    if isinstance(attributo_value, datetime.datetime):
        return datetime_to_local(attributo_value)
    if isinstance(
        attributo_value,
        (str, int, float, bool),
    ):
        return attributo_value
    assert isinstance(attributo_value, list)
    return str(attributo_value)


@dataclass(frozen=True)
class WorkbookOutput:
    workbook: Workbook
    files: set[int]


async def create_workbook(
    db: AsyncDB, conn: Connection, with_events: bool
) -> WorkbookOutput:
    wb = Workbook(iso_dates=True)

    runs_sheet = wb.active
    runs_sheet.title = "Runs"
    attributi_sheet = wb.create_sheet("Attributi")
    chemicals_sheet = wb.create_sheet("chemicals")

    attributi = await db.retrieve_attributi(conn, associated_table=None)
    attributi.sort(key=attributo_sort_key)

    for attributo_column, attributo_header_name in enumerate(
        (
            "Table",
            "Name",
            "Group",
            "Description",
            "Type",
        ),
        start=1,
    ):
        # pyright thinks I cannot access .cell on the worksheet
        cell = attributi_sheet.cell(  # pyright: ignore
            row=1, column=attributo_column, value=attributo_header_name
        )
        new_font = copy(cell.font)  # pyright: ignore
        new_font.bold = True
        cell.font = new_font  # pyright: ignore

    for attributo_row_idx, attributo in enumerate(attributi, start=2):
        attributi_sheet.cell(  # pyright: ignore
            row=attributo_row_idx,
            column=1,
            value=attributo.associated_table.value.capitalize(),
        )
        attributi_sheet.cell(  # pyright: ignore[reportGeneralTypeIssues]
            row=attributo_row_idx, column=2, value=attributo.name
        )
        attributi_sheet.cell(  # pyright: ignore[reportGeneralTypeIssues]
            row=attributo_row_idx, column=3, value=attributo.group
        )
        attributi_sheet.cell(  # pyright: ignore[reportGeneralTypeIssues]
            row=attributo_row_idx, column=4, value=attributo.description
        )
        attributi_sheet.cell(  # pyright: ignore
            row=attributo_row_idx,
            column=5,
            value=attributo_type_to_string(attributo.attributo_type),
        )

    chemical_attributi = [
        a for a in attributi if a.associated_table == AssociatedTable.CHEMICAL
    ]
    for chemical_column, chemical_header_name in enumerate(
        [AttributoId("Name")]
        + [a.name for a in chemical_attributi]
        + [AttributoId("File IDs")],
        start=1,
    ):
        cell = chemicals_sheet.cell(  # pyright: ignore
            row=1, column=chemical_column, value=str(chemical_header_name)
        )
        new_font = copy(cell.font)  # pyright: ignore
        new_font.bold = True
        cell.font = new_font  # pyright: ignore

    files_to_include: set[int] = set()
    chemicals = await db.retrieve_chemicals(conn, attributi)
    for chemical_row_idx, chemical in enumerate(chemicals, start=2):
        chemicals_sheet.cell(  # pyright: ignore
            row=chemical_row_idx,
            column=1,
            value=chemical.name,
        )
        for chemical_column_idx, chemical_attributo in enumerate(
            chemical_attributi,
            start=2,
        ):
            chemicals_sheet.cell(  # pyright: ignore
                row=chemical_row_idx,
                column=chemical_column_idx,
                value=attributo_value_to_spreadsheet_cell(
                    chemical_id_to_name={},
                    attributo_type=chemical_attributo.attributo_type,
                    attributo_value=chemical.attributi.select(chemical_attributo.name),
                ),
            )
        if chemical.files:
            chemicals_sheet.cell(  # pyright: ignore
                row=chemical_row_idx,
                column=2 + len(chemical_attributi),
                value=", ".join(str(f.id) for f in chemical.files),
            )
            files_to_include.update(cast(int, f.id) for f in chemical.files)

    run_attributi = [a for a in attributi if a.associated_table == AssociatedTable.RUN]
    for run_column, run_header_name in enumerate(
        [AttributoId("ID")] + [a.name for a in run_attributi],
        start=1,
    ):
        cell = runs_sheet.cell(
            row=1, column=run_column, value=str(run_header_name)
        )  # pyright: ignore
        new_font = copy(cell.font)  # pyright: ignore
        new_font.bold = True
        cell.font = new_font  # pyright: ignore

    chemical_id_to_name: dict[int, str] = {s.id: s.name for s in chemicals}
    events = await db.retrieve_events(conn)
    event_iterator = 0
    run_row_idx = 2
    for run in await db.retrieve_runs(conn, attributi):
        started = run.attributi.select_datetime(ATTRIBUTO_STARTED)
        event_start = event_iterator
        while (
            with_events
            and started is not None
            and event_iterator < len(events)
            and events[event_iterator].created >= started
        ):
            event_iterator += 1

        for event in events[event_start:event_iterator]:
            runs_sheet.cell(row=run_row_idx, column=2, value=event.created)
            event_text = f"{event.source}: {event.text}"
            if event.files:
                event_text += (
                    " (file IDs: " + ", ".join(str(f.id) for f in event.files) + ")"
                )
                files_to_include.update(cast(int, f.id) for f in event.files)
            runs_sheet.cell(row=run_row_idx, column=4, value=event_text)
            run_row_idx += 1

        runs_sheet.cell(
            row=run_row_idx,
            column=1,
            value=run.id,
        )
        for run_column_idx, run_attributo in enumerate(run_attributi, start=2):
            runs_sheet.cell(
                row=run_row_idx,
                column=run_column_idx,
                value=attributo_value_to_spreadsheet_cell(
                    chemical_id_to_name=chemical_id_to_name,
                    attributo_type=run_attributo.attributo_type,
                    attributo_value=run.attributi.select(run_attributo.name),
                ),
            )
        run_row_idx += 1

    return WorkbookOutput(wb, files_to_include)

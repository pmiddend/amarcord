import datetime
from copy import copy
from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from sqlalchemy.sql import select

from amarcord.db import orm
from amarcord.db.associated_table import AssociatedTable
from amarcord.db.attributi import attributo_type_to_string
from amarcord.db.attributi import attributo_value_from_chemical_orm
from amarcord.db.attributi import attributo_value_from_run_or_ds_orm
from amarcord.db.attributi import schema_dict_to_attributo_type
from amarcord.db.attributo_type import AttributoType
from amarcord.db.attributo_type import AttributoTypeChemical
from amarcord.db.attributo_value import AttributoValue
from amarcord.db.beamtime_id import BeamtimeId
from amarcord.util import datetime_to_local


@dataclass(frozen=True)
class WorkbookOutput:
    workbook: Workbook
    files: list[orm.File]


# Any until openpyxl has official types
def attributo_value_to_spreadsheet_cell(
    chemical_id_to_name: dict[int, str],
    attributo_type: AttributoType,
    attributo_value: AttributoValue,
) -> Any:
    if attributo_value is None:
        return None
    if isinstance(attributo_type, AttributoTypeChemical):
        if not isinstance(attributo_value, int):
            raise TypeError(
                f"chemical IDs have to have type int, got {attributo_value}"
            )
        return chemical_id_to_name.get(
            attributo_value, f"invalid chemical ID {attributo_value}"
        )
    if isinstance(attributo_value, datetime.datetime):
        return datetime_to_local(attributo_value)
    if isinstance(
        attributo_value,
        (str, int, float, bool),
    ):
        return attributo_value
    assert isinstance(attributo_value, list)
    return str(attributo_value)


async def create_workbook(
    session: AsyncSession, beamtime_id: BeamtimeId, with_events: bool
) -> WorkbookOutput:
    wb = Workbook(iso_dates=True)

    runs_sheet = wb.active
    runs_sheet.title = "Runs"
    attributi_sheet = wb.create_sheet("Attributi")
    chemicals_sheet = wb.create_sheet("Chemicals")

    attributi = list(
        (
            await session.scalars(
                select(orm.Attributo)
                .where(orm.Attributo.beamtime_id == beamtime_id)
                .order_by(orm.Attributo.name)
            )
        ).all()
    )

    for attributo_column, attributo_header_name in enumerate(
        (
            "Table",
            "Name",
            "Group",
            "Description",
            "Type",
        ),
        start=1,
    ):
        # pyright thinks I cannot access .cell on the worksheet
        cell = attributi_sheet.cell(  # pyright: ignore
            row=1, column=attributo_column, value=attributo_header_name
        )
        new_font = copy(cell.font)  # pyright: ignore
        cell.font = new_font  # pyright: ignore

    for attributo_row_idx, attributo in enumerate(attributi, start=2):
        attributi_sheet.cell(  # pyright: ignore
            row=attributo_row_idx,
            column=1,
            value=attributo.associated_table.value.capitalize(),
        )
        attributi_sheet.cell(  # pyright: ignore
            row=attributo_row_idx, column=2, value=attributo.name
        )
        attributi_sheet.cell(  # pyright: ignore
            row=attributo_row_idx, column=3, value=attributo.group
        )
        attributi_sheet.cell(  # pyright: ignore
            row=attributo_row_idx, column=4, value=attributo.description
        )
        attributi_sheet.cell(  # pyright: ignore
            row=attributo_row_idx,
            column=5,
            value=attributo_type_to_string(
                schema_dict_to_attributo_type(attributo.json_schema)
            ),
        )

    chemical_attributi = [
        a for a in attributi if a.associated_table == AssociatedTable.CHEMICAL
    ]
    for chemical_column, chemical_header_name in enumerate(
        ["Name"] + [a.name for a in chemical_attributi] + ["File IDs"],
        start=1,
    ):
        cell = chemicals_sheet.cell(  # pyright: ignore
            row=1, column=chemical_column, value=str(chemical_header_name)
        )
        new_font = copy(cell.font)  # pyright: ignore
        cell.font = new_font  # pyright: ignore

    files_to_include: list[orm.File] = []
    chemicals = (
        await session.scalars(
            select(orm.Chemical)
            .where(orm.Chemical.beamtime_id == beamtime_id)
            .options(selectinload(orm.Chemical.attributo_values))
        )
    ).all()
    for chemical_row_idx, chemical in enumerate(chemicals, start=2):
        chemicals_sheet.cell(  # pyright: ignore
            row=chemical_row_idx,
            column=1,
            value=chemical.name,
        )
        for chemical_column_idx, chemical_attributo in enumerate(
            chemical_attributi,
            start=2,
        ):
            chemicals_sheet.cell(  # pyright: ignore
                row=chemical_row_idx,
                column=chemical_column_idx,
                value=attributo_value_to_spreadsheet_cell(
                    chemical_id_to_name={},
                    attributo_type=schema_dict_to_attributo_type(
                        chemical_attributo.json_schema
                    ),
                    attributo_value=next(
                        iter(
                            attributo_value_from_chemical_orm(x)
                            for x in chemical.attributo_values
                            if x.attributo_id == chemical_attributo.id
                        ),
                        None,
                    ),
                ),
            )
        if chemical.files:
            chemicals_sheet.cell(  # pyright: ignore
                row=chemical_row_idx,
                column=2 + len(chemical_attributi),
                value=", ".join(str(f.id) for f in chemical.files),
            )
            files_to_include.extend(chemical.files)

    run_attributi = [a for a in attributi if a.associated_table == AssociatedTable.RUN]
    for run_column, run_header_name in enumerate(
        ["ID", "started", "stopped"] + [a.name for a in run_attributi],
        start=1,
    ):
        cell = runs_sheet.cell(
            row=1, column=run_column, value=str(run_header_name)
        )  # pyright: ignore
        new_font = copy(cell.font)  # pyright: ignore
        cell.font = new_font  # pyright: ignore

    chemical_id_to_name: dict[int, str] = {s.id: s.name for s in chemicals}
    events = (
        await session.scalars(
            select(orm.EventLog)
            .where(orm.EventLog.beamtime_id == beamtime_id)
            .order_by(orm.EventLog.created)
            .options(selectinload(orm.EventLog.files))
        )
    ).all()
    event_iterator = 0
    run_row_idx = 2
    runs = (
        await session.scalars(
            select(orm.Run)
            .where(orm.Run.beamtime_id == beamtime_id)
            .order_by(orm.Run.started)
            .options(selectinload(orm.Run.attributo_values))
        )
    ).all()

    for run in runs:
        started = run.started
        event_start = event_iterator
        # important here: we select events that started before the run we're outputting now.
        # this means we also output events before the first run.
        while (
            with_events
            and event_iterator < len(events)
            and events[event_iterator].created <= started
        ):
            event_iterator += 1

        for event in events[event_start:event_iterator]:
            runs_sheet.cell(row=run_row_idx, column=2, value=event.created)
            event_text = f"{event.source}: {event.text}"
            if event.files:
                event_text += (
                    " (file IDs: " + ", ".join(str(f.id) for f in event.files) + ")"
                )
                files_to_include.extend(event.files)
            runs_sheet.cell(row=run_row_idx, column=4, value=event_text)
            run_row_idx += 1

        runs_sheet.cell(
            row=run_row_idx,
            column=1,
            value=run.external_id,
        )
        runs_sheet.cell(
            row=run_row_idx,
            column=2,
            value=run.started,
        )
        runs_sheet.cell(
            row=run_row_idx,
            column=3,
            value=run.stopped,
        )
        for run_column_idx, run_attributo in enumerate(run_attributi, start=4):
            runs_sheet.cell(
                row=run_row_idx,
                column=run_column_idx,
                value=attributo_value_to_spreadsheet_cell(
                    chemical_id_to_name=chemical_id_to_name,
                    attributo_type=schema_dict_to_attributo_type(
                        run_attributo.json_schema
                    ),
                    attributo_value=next(
                        iter(
                            attributo_value_from_run_or_ds_orm(x)
                            for x in run.attributo_values
                            if x.attributo_id == run_attributo.id
                        ),
                        None,
                    ),
                ),
            )
        run_row_idx += 1

    # Events after the last run must be treated specially
    while (
        runs
        and with_events
        and event_iterator < len(events)
        and events[event_iterator].created > runs[0].started
    ):
        event = events[event_iterator]
        runs_sheet.cell(row=run_row_idx, column=2, value=event.created)
        event_text = f"{event.source}: {event.text}"
        if event.files:
            event_text += (
                " (file IDs: " + ", ".join(str(f.id) for f in event.files) + ")"
            )
            files_to_include.extend(event.files)
        runs_sheet.cell(row=run_row_idx, column=4, value=event_text)
        run_row_idx += 1
        event_iterator += 1

    return WorkbookOutput(wb, files_to_include)
